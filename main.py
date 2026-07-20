from fastapi.staticfiles import StaticFiles
from fastapi import FastAPI, Form, HTTPException, UploadFile, File, Query, Body, Header
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware
import hmac
import os
import base64
import hashlib
from typing import Optional, List
from pydantic import BaseModel
from fastapi.responses import JSONResponse
import requests

from argo_document import argo_document_bloque1, salida_to_dict
from argo_master import build_master_output
from argo_history import save_pipeline_to_history, read_history
from argo_dashboard import build_dashboard_output
from utils_operacion import generar_id_operacion, escribir_log_operacion
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re

from argo_excel_report import generar_reporte_ejecutivo
from datetime import datetime
from argo_control import argo_control_validar_v2
from argo_historial import (
    normalizar_operacion_para_historial,
    guardar_operacion_historial,
    obtener_dashboard_desde_historial,
    obtener_clientes_desde_historial,
    obtener_historial,
    aprobar_operacion as aprobar_operacion_hist
)

from argo_models import AprobarOperacionRequest
from argo_supabase_historial import aprobar_operacion_supabase
from argo_dashboard_pro import construir_dashboard_pro, generar_pdf_dashboard_pro


class ActualizarIncidenciaRequest(BaseModel):
    id_operacion: str
    estado_incidencia: str = "ABIERTA"
    severidad: str = "MEDIA"
    asignado_a: str = ""
    comentario: str = ""


# redeploy master dashboard historial loader fix
app = FastAPI()

# 🔷 Primero crear carpeta
if not os.path.exists("outputs"):
    os.makedirs("outputs")

# 🔷 Luego exponerla
app.mount("/outputs", StaticFiles(directory="outputs"), name="outputs")

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_KEY")
SUPABASE_SERVICE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY")

def supabase_config_ok() -> bool:
    print("DEBUG_SUPABASE_URL_PRESENTE:", bool(SUPABASE_URL))
    print("DEBUG_SUPABASE_SERVICE_KEY_PRESENTE:", bool(SUPABASE_SERVICE_KEY))
    return bool(SUPABASE_URL and SUPABASE_SERVICE_KEY)

def _headers():
    return {
        "apikey": SUPABASE_SERVICE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
        "Content-Type": "application/json"
    }

SUPABASE_BUCKET = "argo-files"

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

async def argo_ocr(file):
    try:
        filename = getattr(file, "filename", "archivo.jpg")

        if hasattr(file, "read"):
            contenido = await file.read()
        else:
            contenido = file

        resultado = {
            "ok": True,
            "estado": "REVISION",
            "severidad_maxima": "MEDIA",
            "conteo": {
                "faltantes": 3,
                "alertas": 0
            },
            "faltantes": [],
            "faltantes_priorizados": [],
            "consolidado": {
                "cliente": "Fives Cinetic Mexico S A De C V",
                "proveedor": "DEMO",
                "paqueteria": None,
                "tracking": None,
                "descripcion": "DETECCION OCR",
                "cantidad_bultos": None,
                "peso_total": None,
                "peso_unidad": None,
                "direccion_origen": None,
                "direccion_destino": None
            },
            "errores": [],
            "procesados": 1,
            "total_archivos": 1
        }

        return resultado

    except Exception as e:
        print("ERROR OCR:", str(e))
        return {
            "ok": False,
            "error": str(e)
        }
        
from fastapi import Request
from fastapi.responses import JSONResponse

from argo_supabase_historial import (
    guardar_operacion_supabase,
    obtener_dashboard_supabase,
    aprobar_operacion_supabase,
    obtener_clientes_supabase,
)

@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    return JSONResponse(
        status_code=500,
        content={
            "ok": False,
            "error_type": type(exc).__name__,
            "detail": str(exc),
            "path": str(request.url.path),
        },
    )

TEMPLATE_FILE = "PLANTILLA_OFICIAL_ARGO_ENTRADA.xlsx"
OUTPUT_FOLDER = "salidas"

ROUTING_V1 = {
    "routing_version": "1.0",
    "rules": [
        {
            "when_estado": "OK",
            "action": "CONTINUAR",
            "next": ["ARGO_CONTROL", "ARGO_CLASS", "ARGO_DOCUMENT"]
        },
        {
            "when_estado": "ADVERTENCIA",
            "action": "CONTINUAR_CON_BANDERA",
            "flag": "REQUIERE_CONFIRMACION",
            "next": ["ARGO_CONTROL", "ARGO_CLASS", "ARGO_DOCUMENT"]
        },
        {
            "when_estado": "REVISION",
            "action": "DETENER",
            "next": [],
            "reason": "DATOS_INSUFICIENTES_O_RIESGO_ALTO"
        }
    ]
}
def argo_control_validar(payload_entrada: dict) -> dict:
    entrada = payload_entrada.get("entrada", {}) or {}

    def normalizar(v):
        return str(v or "").strip()

    def es_no_legible(v):
        t = normalizar(v).lower()
        return t in ["", "no legible", "n/a", "na", "none", "null", "sin dato", "sin datos"]

    campos_operativos = [
        ("cantidad", "Cantidad"),
        ("peso_total", "Peso total"),
        ("descripcion", "Descripción"),
        ("marca", "Marca"),
        ("modelo", "Modelo"),
        ("no_parte", "No. parte"),
        ("no_lote", "Lote"),
        ("no_serie", "Serie"),
        ("pais_origen", "País origen"),
    ]

    validaciones = []
    verificados = 0
    no_verificables = 0

    for campo, etiqueta in campos_operativos:
        valor = entrada.get(campo)

        if es_no_legible(valor):
            estado = "NO_VERIFICABLE"
            resultado = "⚠ No verificable"
            severidad = "MEDIA"
            no_verificables += 1
        else:
            estado = "DISPONIBLE"
            resultado = "✓ Disponible para validación"
            severidad = "BAJA"
            verificados += 1

        validaciones.append({
            "campo": campo,
            "etiqueta": etiqueta,
            "valor_documental": valor,
            "estado": estado,
            "resultado": resultado,
            "severidad": severidad,
        })

    total = len(campos_operativos)
    cobertura_pct = round((verificados / total) * 100, 2) if total else 0

    alertas = []

    for v in validaciones:
        if v["estado"] == "NO_VERIFICABLE":
            alertas.append({
                "tipo": "DATO_NO_VERIFICABLE",
                "campo": v["campo"],
                "mensaje": f"{v['etiqueta']} no disponible o no legible para validación operativa.",
                "severidad": v["severidad"],
            })

    if cobertura_pct >= 85:
        estado = "OK"
        semaforo = "VERDE"
        icono = "🟢"
        severidad_maxima = "NINGUNA"
        dictamen_operativo = "Información suficiente para continuar operación."
    elif cobertura_pct >= 55:
        estado = "ADVERTENCIA"
        semaforo = "AMARILLO"
        icono = "🟡"
        severidad_maxima = "MEDIA"
        dictamen_operativo = "Operación puede continuar con datos no verificables documentados."
    else:
        estado = "REVISION"
        semaforo = "ROJO"
        icono = "🔴"
        severidad_maxima = "ALTA"
        dictamen_operativo = "Información limitada; requiere revisión operativa antes de continuar."

    return {
        "version": "2.0",
        "modulo": "ARGO_CONTROL",
        "estado": estado,
        "semaforo": semaforo,
        "icono": icono,
        "severidad_maxima": severidad_maxima,
        "dictamen_operativo": dictamen_operativo,
        "cobertura_validacion_pct": cobertura_pct,
        "conteo": {
            "campos_totales": total,
            "campos_disponibles": verificados,
            "campos_no_verificables": no_verificables,
            "alertas": len(alertas),
        },
        "validaciones_operativas": validaciones,
        "alertas": alertas,
        "entrada_ref": {
            "version": payload_entrada.get("version"),
            "modulo": payload_entrada.get("modulo"),
            "shipment_id": entrada.get("shipment_id"),
            "tracking": entrada.get("tracking"),
            "cliente": entrada.get("cliente"),
            "proveedor": entrada.get("proveedor"),
        }
    }

if not os.path.exists(OUTPUT_FOLDER):
    os.makedirs(OUTPUT_FOLDER)


@app.get("/health")
def health():
    return {"mensaje": "ARGO ENTRADA MVP funcionando"}

@app.get("/argo/dashboard/pro")
def argo_dashboard_pro(
    request: Request,
    cliente_id: Optional[str] = Query(None),
    x_cliente_id: Optional[str] = Header(None)
):

    x_usuario_email = request.headers.get("x-usuario-email")

    usuario_rbac = obtener_usuario_rbac(x_usuario_email)

    validacion_modulo = validar_modulo_usuario(
        usuario_rbac,
        "analytics_pro"
    )

    if not validacion_modulo.get("ok"):
        return JSONResponse(
            status_code=403,
            content=validacion_modulo
        )

    usuario_actual = obtener_usuario_rbac(x_usuario_email)

    if usuario_actual:

        feature = validar_feature_plan(
            usuario_actual,
            "dashboard_pro"
        )

        if not feature.get("ok"):
            return JSONResponse(
                status_code=403,
                content=feature
            )

    cliente_final = cliente_id or x_cliente_id

    try:
        try:
            base = obtener_dashboard_supabase(cliente_id=cliente_final)
        except TypeError:
            base = obtener_dashboard_supabase(cliente_final)
    except TypeError:
        base = obtener_dashboard_supabase()

    return construir_dashboard_pro(base, cliente_id=cliente_final)






# =========================================================
# DASHBOARD PRO - CENTRO DE INCIDENCIAS ACCIONABLE
# =========================================================

@app.patch("/argo/dashboard/pro/incidencia")
async def actualizar_incidencia_dashboard_pro(
    payload: ActualizarIncidenciaRequest,
    x_cliente_id: str = Header(default=None),
    x_usuario_email: str = Header(default=None),
):
    try:

        permitido, usuario_admin, motivo = validar_permiso_rbac(
            email=x_usuario_email,
            roles_permitidos=ROLES_APROBACION,
            cliente_id=x_cliente_id,
        )

        if not permitido:
            return JSONResponse(
                status_code=403,
                content={
                    "ok": False,
                    "error": motivo,
                    "codigo": "RBAC_DENY"
                }
            )

        rol = str(usuario_admin.get("rol") or "").lower()
        actor_nombre = usuario_admin.get("nombre") or x_usuario_email or "sistema"

        if not supabase_config_ok():
            raise RuntimeError("Supabase no configurado")

        url_select = (
            f"{SUPABASE_URL}/rest/v1/argo_operaciones"
            f"?id_operacion=eq.{payload.id_operacion}"
            f"&select=id_operacion"
        )

        response_select = requests.get(
            url_select,
            headers=_headers(),
            timeout=30
        )

        if response_select.status_code != 200:
            raise RuntimeError(
                f"Error buscando operación: {response_select.status_code} - {response_select.text}"
            )

        data = response_select.json()

        if not data:
            return {
                "ok": False,
                "error": "Operación no encontrada",
                "id_operacion": payload.id_operacion,
            }

        # Nota enterprise:
        # La tabla argo_operaciones aún no tiene columnas físicas de incidencia.
        # Para evitar romper producción, registramos la gestión accionable en auditoría.
        # En la fase de DB/migraciones se podrá persistir también en columnas dedicadas.

        guardar_auditoria_admin(
            accion="dashboard_pro_actualizar_incidencia",
            actor_email=x_usuario_email,
            actor_rol=rol,
            tenant=x_cliente_id,
            objetivo_email=payload.asignado_a,
            detalle={
                "id_operacion": payload.id_operacion,
                "estado_incidencia": payload.estado_incidencia,
                "severidad": payload.severidad,
                "comentario": payload.comentario,
            }
        )

        return {
            "ok": True,
            "mensaje": "Incidencia actualizada correctamente",
            "id_operacion": payload.id_operacion,
            "estado_incidencia": payload.estado_incidencia,
            "severidad": payload.severidad,
            "asignado_a": payload.asignado_a,
            "actualizado_por": actor_nombre,
        }

    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={
                "ok": False,
                "error": str(e),
                "modulo": "dashboard_pro_incidencias"
            }
        )


@app.get("/argo/dashboard/pro/incidencias")
def argo_dashboard_pro_incidencias(
    request: Request,
    cliente_id: Optional[str] = Query(None),
    x_cliente_id: Optional[str] = Header(None)
):

    x_usuario_email = request.headers.get("x-usuario-email")

    usuario_actual = obtener_usuario_rbac(x_usuario_email)

    if usuario_actual:

        feature = validar_feature_plan(
            usuario_actual,
            "dashboard_pro"
        )

        if not feature.get("ok"):
            return JSONResponse(
                status_code=403,
                content=feature
            )
    cliente_final = cliente_id or x_cliente_id

    try:
        try:
            base = obtener_dashboard_supabase(cliente_id=cliente_final)
        except TypeError:
            base = obtener_dashboard_supabase(cliente_final)
    except TypeError:
        base = obtener_dashboard_supabase()

    pro = construir_dashboard_pro(base, cliente_id=cliente_final)
    return {
        "ok": True,
        "cliente_id": cliente_final,
        "incidencias_criticas": pro.get("incidencias_criticas", []),
        "alertas_inteligentes": pro.get("alertas_inteligentes", []),
        "total_incidencias": len(pro.get("incidencias_criticas", [])),
        "total_alertas": len(pro.get("alertas_inteligentes", [])),
    }


@app.get("/argo/dashboard/pro/timeline")
def argo_dashboard_pro_timeline(
    request: Request,
    cliente_id: Optional[str] = Query(None),
    x_cliente_id: Optional[str] = Header(None)
):

    x_usuario_email = request.headers.get("x-usuario-email")

    usuario_actual = obtener_usuario_rbac(x_usuario_email)

    if usuario_actual:

        feature = validar_feature_plan(
            usuario_actual,
            "dashboard_pro"
        )

        if not feature.get("ok"):
            return JSONResponse(
                status_code=403,
                content=feature
            )
    cliente_final = cliente_id or x_cliente_id

    try:
        try:
            base = obtener_dashboard_supabase(cliente_id=cliente_final)
        except TypeError:
            base = obtener_dashboard_supabase(cliente_final)
    except TypeError:
        base = obtener_dashboard_supabase()

    pro = construir_dashboard_pro(base, cliente_id=cliente_final)
    return {
        "ok": True,
        "cliente_id": cliente_final,
        "timeline_vivo": pro.get("timeline_vivo", []),
    }

@app.get("/argo/dashboard/pro/pdf")
def argo_dashboard_pro_pdf(
    request: Request,
    cliente_id: Optional[str] = Query(None),
    x_cliente_id: Optional[str] = Header(None)
):

    x_usuario_email = request.headers.get("x-usuario-email")

    usuario_actual = obtener_usuario_rbac(x_usuario_email)

    if usuario_actual:

        feature_pdf = validar_feature_plan(
            usuario_actual,
            "export_pdf"
        )

        if not feature_pdf.get("ok"):
            return JSONResponse(
                status_code=403,
                content=feature_pdf
            )

        feature_dashboard = validar_feature_plan(
            usuario_actual,
            "dashboard_pro"
        )

        if not feature_dashboard.get("ok"):
            return JSONResponse(
                status_code=403,
                content=feature_dashboard
            )
    cliente_final = cliente_id or x_cliente_id or "cliente_demo"

    try:
        try:
            base = obtener_dashboard_supabase(cliente_id=cliente_final)
        except TypeError:
            base = obtener_dashboard_supabase(cliente_final)
    except TypeError:
        base = obtener_dashboard_supabase()

    pro = construir_dashboard_pro(base, cliente_id=cliente_final)

    safe_cliente = _safe_filename(cliente_final)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nombre_pdf = f"ARGO_DASHBOARD_PRO_{safe_cliente}_{timestamp}.pdf"
    output_path = os.path.join("outputs", nombre_pdf)

    generar_pdf_dashboard_pro(pro, output_path)

    return FileResponse(
        path=output_path,
        filename=nombre_pdf,
        media_type="application/pdf"
    )

@app.get("/descargar/{nombre_archivo}")
def descargar_archivo(nombre_archivo: str):
    ruta_salidas = os.path.join("salidas", nombre_archivo)
    ruta_outputs = os.path.join("outputs", nombre_archivo)

    if os.path.exists(ruta_salidas):
        ruta = ruta_salidas
    elif os.path.exists(ruta_outputs):
        ruta = ruta_outputs
    else:
        raise HTTPException(status_code=404, detail="Archivo no encontrado")

    return FileResponse(
        path=ruta,
        filename=nombre_archivo,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
def _safe_filename(text: str) -> str:
    text = (text or "").strip()
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"[^A-Za-z0-9 _-]", "_", text)
    text = text.replace(" ", "_")
    return text or "SIN_CLIENTE"


def _clear_sheet(ws, max_rows=200, max_cols=10):
    for r in range(1, max_rows + 1):
        for c in range(1, max_cols + 1):
            ws.cell(row=r, column=c).value = None

def subir_archivo_a_supabase(file_path: str, bucket: str = SUPABASE_BUCKET) -> dict:
    if not SUPABASE_URL or not SUPABASE_SERVICE_KEY:
        raise Exception("Faltan variables de entorno de Supabase")

    if not os.path.exists(file_path):
        raise Exception(f"Archivo no existe: {file_path}")

    file_name = os.path.basename(file_path)
    storage_path = f"outputs/{file_name}"

    upload_url = f"{SUPABASE_URL}/storage/v1/object/{bucket}/{storage_path}"

    with open(file_path, "rb") as f:
        response = requests.post(
            upload_url,
            headers={
                "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
                "apikey": SUPABASE_SERVICE_KEY,
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "x-upsert": "true",
            },
            data=f.read(),
        )

    if response.status_code not in [200, 201]:
        raise Exception(f"Error subiendo a Supabase: {response.status_code} {response.text}")

    signed_url_endpoint = f"{SUPABASE_URL}/storage/v1/object/sign/{bucket}/{storage_path}"

    signed_response = requests.post(
        signed_url_endpoint,
        headers={
            "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
            "apikey": SUPABASE_SERVICE_KEY,
            "Content-Type": "application/json",
        },
        json={
            "expiresIn": 60 * 60 * 24 * 7
        },
    )

    signed_url = None

    if signed_response.status_code in [200, 201]:
        signed_data = signed_response.json()
        signed_path = signed_data.get("signedURL") or signed_data.get("signedUrl")

        if signed_path:
            signed_url = f"{SUPABASE_URL}/storage/v1{signed_path}"

    return {
        "bucket": bucket,
        "path": storage_path,
        "file_name": file_name,
        "signed_url": signed_url,
    }

@app.post("/entrada/validar")
def entrada_validar(
    shipment_id: str = Form(...),
    cliente: str = Form(...),
    tracking: str = Form(...),
    peso_total: str = Form(...),
    unidad: str = Form(...),
    pais_origen: str = Form(...),
    proveedor: str = Form("No legible"),
    paqueteria: str = Form("No legible"),
    descripcion: str = Form("No legible"),
    marca: str = Form("No legible"),
    modelo: str = Form("No legible"),
    no_parte: str = Form("No legible"),
    no_lote: str = Form("No legible"),
    no_serie: str = Form("No legible"),
    cantidad: str = Form("No legible"),
):
    # --- Validación anti-placeholder de Swagger ("string") y vacíos ---
    def _clean_required(name: str, v: str) -> str:
        s = (v or "").strip()
        if s == "" or s.lower() == "string":
            raise HTTPException(status_code=422, detail=f"Campo requerido inválido: {name}")
        return s

    shipment_id = _clean_required("shipment_id", shipment_id)
    cliente = _clean_required("cliente", cliente)
    tracking = _clean_required("tracking", tracking)
    peso_total = _clean_required("peso_total", peso_total)
    unidad = _clean_required("unidad", unidad)
    pais_origen = _clean_required("pais_origen", pais_origen)

    fecha_recepcion = datetime.now().strftime("%m/%d/%Y")

    entrada = {
        "shipment_id": shipment_id,
        "fecha_recepcion": fecha_recepcion,
        "cliente": cliente,
        "proveedor": proveedor,
        "paqueteria": paqueteria,
        "tracking": tracking,
        "descripcion": descripcion,
        "marca": marca,
        "modelo": modelo,
        "no_parte": no_parte,
        "no_lote": no_lote,
        "no_serie": no_serie,
        "cantidad": cantidad,
        "unidad": unidad,
        "peso_total": peso_total,
        "pais_origen": pais_origen
    }

    faltantes = []
    alertas = []

    for campo, valor in entrada.items():
        if valor == "No legible":
            faltantes.append({"campo": campo, "valor": valor})

    # Estado base
    estado = "OK"
    severidad_maxima = "NINGUNA"

    if len(faltantes) > 0:
        estado = "ADVERTENCIA"
        severidad_maxima = "MEDIA"

    control_stub = argo_control_validar({
        "version": "1.0",
        "modulo": "ARGO_ENTRADA",
        "entrada": entrada
    })

    return {
        "version": "1.0",
        "modulo": "ARGO_ENTRADA",
        "estado": estado,
        "severidad_maxima": severidad_maxima,
        "conteo": {"faltantes": len(faltantes), "alertas": len(alertas)},
        "faltantes": faltantes,
        "alertas": alertas,
        "entrada": entrada,
        "control": control_stub
    }

@app.post("/generar")
def generar_excel(
    shipment_id: str = Form(...),
    cliente: str = Form(...),
    tracking: str = Form(...),
    peso_total: str = Form(...),
    unidad: str = Form(...),
    pais_origen: str = Form(...),
    proveedor: str = Form("No legible"),
    paqueteria: str = Form("No legible"),
    descripcion: str = Form("No legible"),
    marca: str = Form("No legible"),
    modelo: str = Form("No legible"),
    no_parte: str = Form("No legible"),
    no_lote: str = Form("No legible"),
    no_serie: str = Form("No legible"),
    cantidad: str = Form("No legible"),
):
    wb = load_workbook(TEMPLATE_FILE)
    ws = wb["Entrada"]

    fecha_recepcion = datetime.now().strftime("%m/%d/%Y")

    tracking_original = (tracking or "").strip()
    tracking_texto = tracking_original

    if (
        tracking_texto.lower().endswith("e")
        or "e+" in tracking_texto.lower()
        or "e-" in tracking_texto.lower()
    ):
        tracking_texto = "No legible"

    cliente = (cliente or "").strip() or "No legible"
    proveedor = (proveedor or "").strip() or "No legible"

    # =========================
    # HOJA ENTRADA
    # =========================
    ws["B2"] = (shipment_id or "").strip() or "No legible"
    ws["B3"] = fecha_recepcion
    ws["B4"] = cliente
    ws["B5"] = proveedor
    ws["B6"] = (paqueteria or "").strip() or "No legible"

    ws["B7"].number_format = "@"
    ws["B7"] = tracking_texto

    ws["B8"] = (descripcion or "").strip() or "No legible"
    ws["B9"] = (marca or "").strip() or "No legible"
    ws["B10"] = (modelo or "").strip() or "No legible"
    ws["B11"] = (no_parte or "").strip() or "No legible"
    ws["B12"] = (no_lote or "").strip() or "No legible"
    ws["B13"] = (no_serie or "").strip() or "No legible"
    ws["B14"] = (cantidad or "").strip() or "No legible"
    ws["B15"] = (unidad or "").strip() or "No legible"
    ws["B16"] = (peso_total or "").strip() or "No legible"
    ws["B17"] = (pais_origen or "").strip() or "No legible"

    # =========================
    # DATOS FALTANTES
    # =========================
    faltantes = []

    campos = [
        ("Shipment ID", ws["B2"].value),
        ("Fecha recepción", ws["B3"].value),
        ("Cliente", ws["B4"].value),
        ("Proveedor", ws["B5"].value),
        ("Paquetería", ws["B6"].value),
        ("Tracking", ws["B7"].value),
        ("Descripción", ws["B8"].value),
        ("Marca", ws["B9"].value),
        ("Modelo", ws["B10"].value),
        ("No. parte", ws["B11"].value),
        ("No. lote", ws["B12"].value),
        ("No. serie", ws["B13"].value),
        ("Cantidad", ws["B14"].value),
        ("Unidad", ws["B15"].value),
        ("Peso total", ws["B16"].value),
        ("País origen", ws["B17"].value),
    ]

    for campo, valor in campos:
        if str(valor).strip() == "No legible":
            faltantes.append((campo, "No legible"))

    ws_df = wb["Datos faltantes"]
    _clear_sheet(ws_df)

    ws_df["A1"] = "Campo"
    ws_df["B1"] = "Valor"

    fila = 2
    for campo, valor in faltantes:
        ws_df[f"A{fila}"] = campo
        ws_df[f"B{fila}"] = valor
        fila += 1

    # =========================
    # ALERTAS
    # =========================
    alertas = []

    if tracking_texto == "No legible":
        alertas.append(("Tracking inválido", "Formato incorrecto", "ALTA"))

    if ws["B16"].value == "No legible":
        alertas.append(("Peso no legible", "No se detectó peso válido", "MEDIA"))

    if ws["B4"].value == "No legible":
        alertas.append(("Cliente faltante", "No se detectó cliente", "ALTA"))

    if ws["B5"].value == "No legible":
        alertas.append(("Proveedor faltante", "No se detectó proveedor", "MEDIA"))

    ws_al = wb["Alertas"]
    _clear_sheet(ws_al)

    ws_al["A1"] = "Alerta"
    ws_al["B1"] = "Detalle"
    ws_al["C1"] = "Severidad"

    fila = 2
    for a, d, s in alertas:
        ws_al[f"A{fila}"] = a
        ws_al[f"B{fila}"] = d
        ws_al[f"C{fila}"] = s
        fila += 1

    # =========================
    # RESUMEN OPERATIVO
    # =========================
    ws_res = wb["Resumen operativo"]
    _clear_sheet(ws_res)

    ws_res["A1"] = "Item"
    ws_res["B1"] = "Valor"

    resumen = [
        ("Fecha recepción", ws["B3"].value),
        ("Cliente", ws["B4"].value),
        ("Proveedor", ws["B5"].value),
        ("Paquetería", ws["B6"].value),
        ("Tracking", ws["B7"].value),
        ("Shipment ID", ws["B2"].value),
        ("País origen", ws["B17"].value),
        ("Faltantes (#)", len(faltantes)),
        ("Alertas (#)", len(alertas)),
    ]

    hay_alta = any(a[2] == "ALTA" for a in alertas)
    if hay_alta:
        estado = "REVISIÓN"
    elif len(faltantes) > 0 or len(alertas) > 0:
        estado = "ADVERTENCIA"
    else:
        estado = "OK"

    resumen.append(("Estado", estado))

    severidad_max = "NINGUNA"
    if any(a[2] == "ALTA" for a in alertas):
        severidad_max = "ALTA"
    elif any(a[2] == "MEDIA" for a in alertas):
        severidad_max = "MEDIA"
    elif any(a[2] == "BAJA" for a in alertas):
        severidad_max = "BAJA"

    resumen.append(("Severidad máxima", severidad_max))

    fila = 2
    for item, valor in resumen:
        ws_res[f"A{fila}"] = item
        ws_res[f"B{fila}"] = valor
        fila += 1

    # =========================
    # GUARDADO
    # =========================
    cliente_archivo = _safe_filename(cliente)
    ult4 = (tracking_original[-4:] if len(tracking_original) >= 4 else tracking_original) or "XXXX"

    output_name = f"ENTRADA_{cliente_archivo}_{ult4}.xlsx"
    output_path = os.path.join(OUTPUT_FOLDER, output_name)

    wb.save(output_path)

    return FileResponse(output_path, filename=output_name)

@app.post("/argo-control")
async def ejecutar_argo_control(
    archivo_entrada: UploadFile = File(...),
    plantilla_control: UploadFile = File(...),
    modo: str = Form("excel")
):
    # Guardar archivos temporalmente
    entrada_path = f"temp_{archivo_entrada.filename}"
    control_path = f"temp_{plantilla_control.filename}"

    with open(entrada_path, "wb") as f:
        f.write(await archivo_entrada.read())

    with open(control_path, "wb") as f:
        f.write(await plantilla_control.read())

    # Ejecutar validación
    output_path, icono, estatus = argo_control_validar_v2(
        entrada_path,
        control_path
    )

    # Modo JSON (para ARGO CLASS)
    if modo.lower() == "json":

            from argo_control import extraer_resumen_control_desde_excel

            resumen = extraer_resumen_control_desde_excel(output_path)

            return {
            "ok": True,
            "modulo": "ARGO_CONTROL",
            "estatus": estatus,
            "icono": icono,
            "output_path": output_path,
            "resumen": resumen
        }

    # Mantener Excel
    return FileResponse(
        path=output_path,
        filename=output_path.split("/")[-1],
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
from pydantic import BaseModel
from typing import Any, Dict
from argo_class_engine import build_output

class ArgoClassRequest(BaseModel):
    payload: Dict[str, Any]

@app.post("/argo/class/v2026/clasificar")
def argo_class_clasificar(req: ArgoClassRequest):

    payload = req.payload

    # Si viene información de ARGO CONTROL, agregarla al payload
    if "control" in payload:
        payload["argo_control"] = payload["control"]

    return build_output(payload)

from fastapi import UploadFile, File, Form, HTTPException
import os

# IMPORTS necesarios (ajusta si tus rutas/nombres son distintos)
from argo_control import argo_control_validar_v2, extraer_resumen_control_desde_excel
from argo_class_engine import build_output


@app.post("/argo/pipeline/clasificar")
async def argo_pipeline_clasificar(
    archivo_entrada: UploadFile = File(...),
    plantilla_control: UploadFile = File(...),
    descripcion: str = Form(""),
):
    id_operacion = None
    entrada_path = None
    control_path = None

    try:
        # ID único end-to-end (AHORA sí dentro del try)
        id_operacion = generar_id_operacion()
        print(f"ID_OPERACION_PIPELINE: {id_operacion}")

        # 1) Guardar archivos temporales
        entrada_path = f"temp_{archivo_entrada.filename}"
        control_path = f"temp_{plantilla_control.filename}"

        with open(entrada_path, "wb") as f:
            f.write(await archivo_entrada.read())

        with open(control_path, "wb") as f:
            f.write(await plantilla_control.read())

        # 2) Ejecutar ARGO CONTROL (con trazabilidad)
        output_path, icono, estatus = argo_control_validar_v2(
            entrada_path,
            control_path,
            id_operacion=id_operacion
        )

        resumen = extraer_resumen_control_desde_excel(output_path)

        control_json = {
            "ok": True,
            "modulo": "ARGO_CONTROL",
            "id_operacion": id_operacion,
            "estatus": estatus,
            "icono": icono,
            "output_path": output_path,
            "resumen": resumen
        }

        # 3) Payload para ARGO CLASS (con id_operacion)
        payload_master = {
            "meta": {
            "id_operacion": id_operacion,
            "id_shipment": None,
            "id_item": None,
        },
        "descripcion": descripcion or "",

        # Paths estándar + compatibilidad (evita NameError)
        "entrada_path": entrada_path,
        "control_path": control_path,
        "archivo_entrada_path": entrada_path,
        "plantilla_control_path": control_path,

        # Resumen de CONTROL para influir certeza
        "control": {"resumen": resumen},

        # Objeto completo de CONTROL por si se ocupa
        "argo_control": control_json
    }

        # 4) Ejecutar ARGO CLASS
        salida_class = build_output(payload_master)

        # 5) Log JSON por operación (si falla NO rompe)
        try:
            escribir_log_operacion(
                id_operacion=id_operacion,
                payload={
                    "modulo": "ARGO_PIPELINE",
                    "inputs": {
                        "entrada_path": entrada_path,
                        "plantilla_control_path": control_path,
                        "descripcion": descripcion or "",
                    },
                    "outputs": {
                        "control_output_path": output_path,
                        "class_output_path": (salida_class.get("output_path") if isinstance(salida_class, dict) else None),
                    },
                    "control": {
                        "estatus": estatus,
                        "icono": icono,
                        "resumen": resumen,
                    },
                    "class": salida_class,
                },
                logs_dir="logs",
            )
        except Exception as log_err:
            print(f"WARNING LOG {id_operacion}: {log_err}")

  
                       # 6) ARGO DOCUMENT
        salida_document = argo_document_bloque1(
            input_xlsx_path=output_path,
            plantilla_path="PLANTILLA_OFICIAL_ARGO_DOCUMENT_MEJORADA_v2026.xlsx",
            outputs_dir="outputs",
            id_operacion=id_operacion,
        )
        document_json = salida_to_dict(salida_document)

        master_json = build_master_output({
            "id_operacion": id_operacion,
            "control": control_json,
            "class": salida_class,
            "document": document_json
        })

        control_public_url = ""
        document_public_url = ""

        try:
            if output_path:
                control_public_url = subir_archivo_a_supabase(output_path)

            if document_json.get("output_path"):
                document_public_url = subir_archivo_a_supabase(document_json["output_path"])

        except Exception as e:
            print("ERROR SUBIDA:", str(e))


# 🔥 ahora sí ya existen las URLs
        master_json = build_master_output(...)

        if "descargas" not in master_json:
            master_json["descargas"] = {}

        master_json["descargas"]["control_url"] = control_public_url
        master_json["descargas"]["document_url"] = document_public_url

        try:
            if output_path:
                control_public_url = subir_archivo_a_supabase(output_path)
        except Exception as supa_err:
            print(f"WARNING SUPABASE CONTROL [{id_operacion}]: {supa_err}")

        try:
            if document_json.get("output_path"):
                document_public_url = subir_archivo_a_supabase(document_json["output_path"])
        except Exception as supa_err:
            print(f"WARNING SUPABASE DOCUMENT [{id_operacion}]: {supa_err}")

        pipeline_result = {
            "ok": True,
            "modulo": "ARGO_PIPELINE",
            "id_operacion": id_operacion,
            "control": control_json,
            "class": salida_class,
            "document": document_json,
            "archivos_publicos": {
                "control_url": control_public_url,
                "document_url": document_public_url,
            }
        }

        master_json = build_master_output(pipeline_result)
        pipeline_result["master"] = master_json

        history_save_path = ""
        history_save_error = ""

        # asegurar que archivos_publicos exista en pipeline_result
        if "archivos_publicos" not in pipeline_result:
            pipeline_result["archivos_publicos"] = {}

        # ===== HISTORIAL ANTIGUO DESACTIVADO =====
        # try:
        #     history_save_path = save_pipeline_to_history(pipeline_result, logs_dir="logs")
        #     print(f"HISTORY OK [{id_operacion}] -> {history_save_path}")
        # except Exception as history_err:
        #     history_save_error = str(history_err)
        #     print(f"WARNING HISTORY [{id_operacion}]: {history_save_error}")
        #
        # pipeline_result["history_debug"] = {
        #     "saved": history_save_error == "",
        #     "path": history_save_path,
        #     "error": history_save_error,
        # }

        print("DEBUG PIPELINE_RESULT KEYS:", list(pipeline_result.keys()) if isinstance(pipeline_result, dict) else type(pipeline_result))
        print("DEBUG MASTER:", pipeline_result.get("master") if isinstance(pipeline_result, dict) else None)
        print("DEBUG OPERACIONES:", pipeline_result.get("operaciones") if isinstance(pipeline_result, dict) else None)
        
                # ===== GUARDAR EN HISTORIAL (FIX COMPLETO) =====
        try:
            print("DEBUG: entrando a guardado de historial")

            cliente_id = "cliente_bodega_2"
            cliente_nombre = "Bodega El Güero"

            pipeline_output_historial = {
                "ok": True,
                "modulo": "ARGO_PIPELINE",
                "id_operacion": id_operacion,
                "estado": control_json.get("estatus"),
                "severidad_maxima": "NINGUNA",
                "decision": {
                    "accion": "CONTINUAR",
                    "razon": "Pipeline OK"
                },
                "ocr": {},
                "generacion": {
                    "entrada": {
                        "cliente": (
                            pipeline_result.get("cliente")
                            or cliente_nombre
                        ),
                        "tracking": (
                            pipeline_result.get("tracking")
                            or pipeline_result.get("shipment_id")
                            or control_json.get("tracking")
                        )
                    },
                    "archivo_generado": pipeline_result.get("archivo_generado"),
                    "ruta_archivo": pipeline_result.get("ruta_archivo"),
                    "descarga": pipeline_result.get("descarga")
                }
            }

            registro_historial = normalizar_operacion_para_historial(
                pipeline_output=pipeline_output_historial,
                cliente_id=cliente_id,
                cliente_nombre=cliente_nombre
            )

            print("DEBUG REGISTRO:", registro_historial)

            if registro_historial.get("id_operacion"):
                guardar_operacion_supabase(registro_historial)

        except Exception as e:
            print("ERROR guardando historial:", str(e))

        return pipeline_result        

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"ARGO_PIPELINE error: {str(e)}")

    finally:
        try:
            if entrada_path and os.path.exists(entrada_path):
                os.remove(entrada_path)
            if control_path and os.path.exists(control_path):
                os.remove(control_path)
        except Exception as cleanup_err:
            print(f"WARNING CLEANUP {id_operacion}: {cleanup_err}")

@app.post("/argo-document")
async def ejecutar_argo_document(
    archivo_xlsx: UploadFile = File(...),
    id_operacion: Optional[str] = Form(None),
):
    temp_input_path = f"temp_{archivo_xlsx.filename}"

    with open(temp_input_path, "wb") as f:
        f.write(await archivo_xlsx.read())

    plantilla_path = "PLANTILLA_OFICIAL_ARGO_DOCUMENT_MEJORADA_v2026.xlsx"

    salida = argo_document_bloque1(
        input_xlsx_path=temp_input_path,
        plantilla_path=plantilla_path,
        outputs_dir="outputs",
        id_operacion=id_operacion,
    )

    try:
        os.remove(temp_input_path)
    except Exception:
        pass

    return JSONResponse(content=salida_to_dict(salida))
       
@app.get("/argo/history")
async def consultar_historial_argo(limit: int = 50):
    data = read_history(limit=limit, logs_dir="logs")
    return {
        "ok": True,
        "modulo": "ARGO_HISTORY",
        "total": len(data),
        "items": data
    }


# =========================================================
# ENDPOINTS HISTORIAL / DASHBOARD / APROBACION
# =========================================================


# =========================================================
# RBAC ENTERPRISE BACKEND
# =========================================================

ROLES_APROBACION = {"supervisor", "admin", "admin_cliente", "master_admin"}
ROLES_ADMIN_CLIENTES = {"admin", "admin_cliente", "master_admin"}


def obtener_usuario_rbac(email: str):
    if not email:
        return None

    if not supabase_config_ok():
        raise RuntimeError("Supabase no configurado")

    url = f"{SUPABASE_URL}/rest/v1/argo_usuarios?email=eq.{email}&select=email,nombre,id_cliente,rol,activo,plan_saas"

    response = requests.get(url, headers=_headers(), timeout=30)

    if response.status_code != 200:
        raise RuntimeError(f"Error consultando usuario RBAC: {response.status_code} - {response.text}")

    data = response.json()

    if not data:
        return None

    return data[0]


def validar_permiso_rbac(email: str, roles_permitidos: set, cliente_id: str = None):
    user = obtener_usuario_rbac(email)

    if not user:
        return False, None, "Usuario no encontrado"

    if user.get("activo") is False:
        return False, user, "Usuario inactivo"

    rol = str(user.get("rol") or "operador").lower()
    cliente_usuario = user.get("id_cliente")

    if rol not in roles_permitidos:
        return False, user, "Permisos insuficientes"

    if cliente_id and rol != "master_admin" and cliente_usuario != cliente_id:
        return False, user, "Usuario no pertenece al cliente solicitado"

    return True, user, "Autorizado"

@app.post("/argo/operacion/aprobar")
async def aprobar_operacion(
    payload: AprobarOperacionRequest,
    x_cliente_id: str = Header(default=None),
    x_usuario_email: str = Header(default=None),
    x_usuario_rol: str = Header(default="operador"),
):

    try:
        permitido, usuario_rbac, motivo = validar_permiso_rbac(
            email=x_usuario_email,
            roles_permitidos=ROLES_APROBACION,
            cliente_id=x_cliente_id,
        )

        if not permitido:
            return {
                "ok": False,
                "error": motivo,
                "codigo": "RBAC_DENY",
                "rol_detectado": str((usuario_rbac or {}).get("rol") or x_usuario_rol or "operador").lower(),
            }

        rol = str(usuario_rbac.get("rol") or "operador").lower()
        aprobador = usuario_rbac.get("nombre") or payload.aprobada_por or "sistema"

        resultado = aprobar_operacion_supabase(
            id_operacion=payload.id_operacion,
            usuario=aprobador
        )

        return {
            "ok": True,
            "mensaje": "Operación aprobada correctamente",
            "data": resultado,
            "rbac": {
                "email": x_usuario_email,
                "rol": rol,
                "cliente": x_cliente_id,
                "autorizado": True
            }
        }

    except Exception as e:
        return {
            "ok": False,
            "error": str(e)
        }
# =========================================
# LOGIN USUARIO
# =========================================
# =========================================
# PASSWORD SEGURA
# =========================================

PBKDF2_ITERATIONS = 100_000


def hash_password(password: str) -> str:
    salt = os.urandom(16)
    dk = hashlib.pbkdf2_hmac(
        "sha256",
        password.encode("utf-8"),
        salt,
        PBKDF2_ITERATIONS
    )
    salt_b64 = base64.b64encode(salt).decode("utf-8")
    hash_b64 = base64.b64encode(dk).decode("utf-8")
    return f"pbkdf2_sha256${PBKDF2_ITERATIONS}${salt_b64}${hash_b64}"


def verify_password(password_plano: str, password_guardado: str) -> bool:
    if not password_guardado:
        return False

    # Compatibilidad temporal con passwords viejos en texto plano
    if not password_guardado.startswith("pbkdf2_sha256$"):
        return hmac.compare_digest(password_plano, password_guardado)

    try:
        algoritmo, iteraciones_str, salt_b64, hash_b64 = password_guardado.split("$", 3)
        iteraciones = int(iteraciones_str)
        salt = base64.b64decode(salt_b64.encode("utf-8"))
        hash_real = base64.b64decode(hash_b64.encode("utf-8"))

        hash_test = hashlib.pbkdf2_hmac(
            "sha256",
            password_plano.encode("utf-8"),
            salt,
            iteraciones
        )

        return hmac.compare_digest(hash_test, hash_real)

    except Exception:
        return False





PLANES_SAAS = {
    "BASIC": {
        "nombre": "Basic",
        "modulos": ["entrada_documental", "camara_pro"],
        "limites": {
            "operaciones_mes": 250,
            "usuarios": 3,
            "clientes_tenant": 1,
            "export_pdf": False,
            "dashboard_pro": False,
        },
    },
    "PRO": {
        "nombre": "PRO",
        "modulos": [
            "entrada_documental",
            "camara_pro",
            "aprobaciones",
            "dashboard",
            "analytics_pro",
            "reportes",
        ],
        "limites": {
            "operaciones_mes": 1500,
            "usuarios": 15,
            "clientes_tenant": 3,
            "export_pdf": True,
            "dashboard_pro": True,
        },
    },
    "ENTERPRISE": {
        "nombre": "Enterprise",
        "modulos": [
            "entrada_documental",
            "camara_pro",
            "aprobaciones",
            "dashboard",
            "analytics_pro",
            "reportes",
            "admin_saas",
            "auditoria",
            "incidencias",
            "argo_connect",
        ],
        "limites": {
            "operaciones_mes": None,
            "usuarios": None,
            "clientes_tenant": None,
            "export_pdf": True,
            "dashboard_pro": True,
        },
    },
}


def normalizar_plan(plan: str) -> str:
    plan = str(plan or "ENTERPRISE").strip().upper()
    if plan not in PLANES_SAAS:
        return "ENTERPRISE"
    return plan


def obtener_plan_saas(user: dict) -> dict:
    plan_code = normalizar_plan(
        user.get("plan")
        or user.get("plan_saas")
        or user.get("tipo_plan")
        or "ENTERPRISE"
    )

    config = PLANES_SAAS[plan_code]

    return {
        "codigo": plan_code,
        "nombre": config["nombre"],
        "limites": config["limites"],
    }


def obtener_modulos_por_rol(rol: str) -> list:
    rol = str(rol or "operador").lower()

    base = ["entrada_documental"]

    if rol in {"operador"}:
        return base + ["camara_pro"]

    if rol in {"supervisor"}:
        return base + [
            "camara_pro",
            "aprobaciones",
            "dashboard",
            "analytics_pro",
            "reportes",
        ]

    if rol in {"admin", "admin_cliente", "master_admin"}:
        return base + [
            "camara_pro",
            "aprobaciones",
            "dashboard",
            "analytics_pro",
            "reportes",
            "admin_saas",
            "auditoria",
            "incidencias",
            "argo_connect",
        ]

    return base


def obtener_modulos_por_plan_y_rol(user: dict) -> list:
    rol = str(user.get("rol") or "operador").lower()
    plan = obtener_plan_saas(user)
    modulos_rol = set(obtener_modulos_por_rol(rol))
    modulos_plan = set(PLANES_SAAS[plan["codigo"]]["modulos"])

    if rol == "master_admin":
        return sorted(modulos_rol)

    return sorted(modulos_rol.intersection(modulos_plan))



# =========================================================
# SAAS LIMIT ENFORCEMENT
# =========================================================



def usuario_tiene_modulo(user: dict, modulo: str) -> bool:
    try:
        if not user:
            return False

        modulos = obtener_modulos_por_plan_y_rol(user)

        return modulo in modulos

    except Exception:
        return False


def validar_modulo_usuario(user: dict, modulo: str):
    if not user:
        return {
            "ok": False,
            "error": "Usuario no autenticado",
            "codigo": "AUTH_REQUIRED"
        }

    if user.get("activo") is False:
        return {
            "ok": False,
            "error": "Usuario inactivo",
            "codigo": "USER_INACTIVE"
        }

    permitido = usuario_tiene_modulo(user, modulo)

    if not permitido:
        return {
            "ok": False,
            "error": f"Modulo no permitido: {modulo}",
            "codigo": "MODULO_DENEGADO",
            "modulo": modulo,
            "plan": obtener_plan_saas(user).get("codigo"),
            "rol": user.get("rol"),
        }

    return {
        "ok": True
    }



def contar_operaciones_mes(cliente_id: str) -> int:
    try:

        if not cliente_id:
            return 0

        mes_actual = datetime.now().strftime("%Y-%m")

        url = (
            f"{SUPABASE_URL}/rest/v1/argo_operaciones"
            f"?cliente_id=eq.{cliente_id}"
            f"&select=id_operacion,created_at,timestamp_local,cliente_id"
        )

        response = requests.get(
            url,
            headers=_headers(),
            timeout=30
        )

        if response.status_code != 200:
            return 0

        data = response.json()

        total = 0

        for row in data:

            fecha = str(row.get("created_at") or row.get("timestamp_local") or "")

            if fecha.startswith(mes_actual):
                total += 1

        return total

    except Exception as e:
        print("ERROR CONTANDO OPERACIONES:", str(e))
        return 0



def validar_feature_plan(usuario: dict, feature: str):
    try:
        plan = obtener_plan_saas(usuario)
        codigo = plan.get("codigo")
        limites = PLANES_SAAS.get(codigo, {}).get("limites", {})

        if feature == "export_pdf":
            permitido = limites.get("export_pdf") is True
        elif feature == "dashboard_pro":
            permitido = limites.get("dashboard_pro") is True
        else:
            permitido = False

        if not permitido:
            guardar_auditoria_admin(
                accion="FEATURE_BLOCKED_BY_PLAN",
                actor_email=usuario.get("email"),
                actor_rol=usuario.get("rol"),
                tenant=usuario.get("id_cliente"),
                detalle={
                    "feature": feature,
                    "plan": codigo,
                }
            )

            return {
                "ok": False,
                "error": "Función no disponible para tu plan SaaS",
                "codigo": "FEATURE_BLOCKED",
                "feature": feature,
                "plan": codigo,
            }

        return {"ok": True, "feature": feature, "plan": codigo}

    except Exception as e:
        return {
            "ok": False,
            "error": str(e),
            "codigo": "FEATURE_VALIDATION_ERROR",
            "feature": feature,
        }





def contar_usuarios_tenant(cliente_id: str) -> int:

    try:

        if not cliente_id:
            return 0

        url = (
            f"{SUPABASE_URL}/rest/v1/argo_usuarios"
            f"?id_cliente=eq.{cliente_id}"
            f"&select=email"
        )

        response = requests.get(
            url,
            headers=_headers(),
            timeout=30
        )

        if response.status_code != 200:
            return 0

        data = response.json()

        return len(data)

    except Exception:
        return 0




def validar_licencia_saas(user: dict) -> dict:

    try:

        if not user:
            return {
                "ok": False,
                "error": "Usuario inválido",
                "codigo": "INVALID_USER"
            }

        rol = str(user.get("rol") or "").lower()

        # MASTER ADMIN SIEMPRE ENTRA
        if rol == "master_admin":
            return {
                "ok": True,
                "estado": "MASTER_ADMIN"
            }

        licencia = (
            user.get("estado_licencia")
            or user.get("licencia_estado")
            or user.get("saas_estado")
            or "ACTIVA"
        )

        licencia = str(licencia).upper()

        fecha_vencimiento = (
            user.get("fecha_vencimiento")
            or user.get("licencia_hasta")
            or user.get("saas_expira")
        )

        dias_restantes = None

        try:

            if fecha_vencimiento:

                fecha_dt = datetime.fromisoformat(
                    str(fecha_vencimiento).replace("Z", "")
                )

                dias_restantes = (
                    fecha_dt.date() - datetime.now().date()
                ).days

                # EXPIRACIÓN AUTOMÁTICA
                if dias_restantes < 0:
                    licencia = "VENCIDA"

                elif dias_restantes <= 7 and licencia == "ACTIVA":
                    licencia = "POR_VENCER"

        except Exception as e:
            print("ERROR VALIDANDO FECHA LICENCIA:", str(e))

        if licencia in ["SUSPENDIDA", "VENCIDA", "BLOQUEADA"]:

            guardar_auditoria_admin(
                accion="SAAS_ACCESO_BLOQUEADO",
                actor_email=user.get("email"),
                actor_rol=user.get("rol"),
                tenant=user.get("id_cliente"),
                detalle={
                    "estado_licencia": licencia
                }
            )

            return {
                "ok": False,
                "error": f"Licencia SaaS {licencia}",
                "codigo": "SAAS_LICENSE_BLOCKED",
                "estado_licencia": licencia
            }

        if licencia == "GRACIA":

            return {
                "ok": True,
                "estado_licencia": licencia,
                "warning": "Licencia en periodo de gracia"
            }

        return {
            "ok": True,
            "estado_licencia": licencia,
            "dias_restantes": dias_restantes,
            "fecha_vencimiento": fecha_vencimiento,
        }

    except Exception as e:

        return {
            "ok": False,
            "error": str(e),
            "codigo": "SAAS_LICENSE_ERROR"
        }



def validar_limite_usuarios_plan(user: dict) -> dict:

    try:

        if not user:
            return {
                "ok": False,
                "error": "Usuario inválido"
            }

        plan = obtener_plan_saas(user)

        limites = plan.get("limites", {}) or {}

        limite_usuarios = limites.get("usuarios")

        if limite_usuarios is None:

            return {
                "ok": True,
                "plan": plan["codigo"],
                "usuarios_actuales": 0,
                "limite_usuarios": None,
                "modo": "ILIMITADO"
            }

        cliente_id = user.get("id_cliente")

        usuarios_actuales = contar_usuarios_tenant(cliente_id)

        restante = max(limite_usuarios - usuarios_actuales, 0)

        porcentaje = 0

        if limite_usuarios > 0:
            porcentaje = round(
                (usuarios_actuales / limite_usuarios) * 100,
                2
            )

        if usuarios_actuales >= limite_usuarios:

            guardar_auditoria_admin(
                accion="LIMITE_SAAS_USUARIOS_EXCEDIDO",
                actor_email=user.get("email"),
                actor_rol=user.get("rol"),
                tenant=cliente_id,
                detalle={
                    "plan": plan.get("codigo"),
                    "limite": limite_usuarios,
                    "usuarios_actuales": usuarios_actuales,
                }
            )

            return {
                "ok": False,
                "error": "Limite de usuarios alcanzado",
                "codigo": "PLAN_USER_LIMIT_REACHED",
                "plan": plan["codigo"],
                "usuarios_actuales": usuarios_actuales,
                "limite_usuarios": limite_usuarios,
                "usuarios_restantes": restante,
                "porcentaje": porcentaje,
                "upgrade_requerido": True
            }

        return {
            "ok": True,
            "plan": plan["codigo"],
            "usuarios_actuales": usuarios_actuales,
            "limite_usuarios": limite_usuarios,
            "usuarios_restantes": restante,
            "porcentaje": porcentaje,
            "upgrade_requerido": porcentaje >= 80
        }

    except Exception as e:

        return {
            "ok": False,
            "error": str(e),
            "codigo": "PLAN_USER_LIMIT_ERROR"
        }



def validar_limite_operaciones_plan(usuario: dict):

    try:

        plan = obtener_plan_saas(usuario)

        limite = (
            plan.get("limites", {})
            .get("operaciones_mes")
        )

        if limite is None:
            return {
                "ok": True
            }

        cliente_id = usuario.get("id_cliente")

        usadas = contar_operaciones_mes(cliente_id)

        if usadas >= limite:

            guardar_auditoria_admin(
                accion="LIMITE_SAAS_OPERACIONES_EXCEDIDO",
                actor_email=usuario.get("email"),
                actor_rol=usuario.get("rol"),
                tenant=cliente_id,
                detalle={
                    "plan": plan.get("codigo"),
                    "limite": limite,
                    "usadas": usadas,
                }
            )

            return {
                "ok": False,
                "error": "Límite mensual de operaciones alcanzado",
                "codigo": "SAAS_LIMIT_REACHED",
                "plan": plan.get("codigo"),
                "limite": limite,
                "usadas": usadas,
            }

        return {
            "ok": True,
            "plan": plan.get("codigo"),
            "limite": limite,
            "usadas": usadas,
            "debug_saas": {
                "cliente_id": cliente_id,
                "email": usuario.get("email"),
                "rol": usuario.get("rol"),
            }
        }

    except Exception as e:
        return {
            "ok": False,
            "error": str(e)
        }



@app.post("/argo/login")
async def login_usuario(payload: dict = Body(...)):
    
    email = payload.get("email")
    password = payload.get("password")

    if not email or not password:
        return {
            "ok": False,
            "error": "Faltan credenciales"
        }

    if not supabase_config_ok():
        return {
            "ok": False,
            "error": "Supabase no configurado"
        }

    url = f"{SUPABASE_URL}/rest/v1/argo_usuarios?email=eq.{email}&select=*"

    response = requests.get(url, headers=_headers())

    if response.status_code != 200:
        return {
            "ok": False,
            "error": "Error en consulta"
        }

    data = response.json()

    if not data:
        return {
            "ok": False,
            "error": "Credenciales inválidas"
        }

    user = data[0]

    licencia = validar_licencia_saas(user)

    if not licencia.get("ok"):

        return licencia

    if not verify_password(password, user.get("password", "")):
        return {
            "ok": False,
            "error": "Credenciales inválidas"
        }

    return {
        "ok": True,
        "usuario": {
            "email": user["email"],
            "nombre": user["nombre"],
            "id_cliente": user["id_cliente"],
            "rol": user["rol"],
            "plan": obtener_plan_saas(user),
            "limites": obtener_plan_saas(user).get("limites"),
            "modulos": obtener_modulos_por_plan_y_rol(user),
            "consumo": validar_limite_operaciones_plan(user),
            "usuarios_plan": validar_limite_usuarios_plan(user)
        }
    }

@app.get("/argo/clientes")
async def endpoint_clientes():
    return obtener_clientes_supabase()

# =========================================================
# SAAS ADMIN - CREAR USUARIO
# =========================================================

@app.post("/argo/admin/crear_usuario")
async def crear_usuario_admin(
    payload: dict = Body(...),
    x_usuario_email: str = Header(default=None),
):

    try:

        permitido, usuario_admin, motivo = validar_permiso_rbac(
            email=x_usuario_email,
            roles_permitidos=ROLES_ADMIN_CLIENTES,
        )

        if not permitido:
            return {
                "ok": False,
                "error": motivo,
                "codigo": "RBAC_DENY"
            }

        email = str(payload.get("email") or "").strip().lower()
        password = str(payload.get("password") or "").strip()
        nombre = str(payload.get("nombre") or "").strip()
        rol = str(payload.get("rol") or "operador").strip().lower()
        activo = bool(payload.get("activo", True))

        if not email or not password or not nombre:
            return {
                "ok": False,
                "error": "Datos incompletos"
            }

        licencia = validar_licencia_saas(usuario_admin)

        if not licencia.get("ok"):

            return JSONResponse(
                status_code=403,
                content=licencia
            )

        rol_admin = str(usuario_admin.get("rol") or "").lower()

        validacion_usuarios = validar_limite_usuarios_plan(usuario_admin)

        if not validacion_usuarios.get("ok"):

            return JSONResponse(
                status_code=403,
                content=validacion_usuarios
            )
        cliente_admin = usuario_admin.get("id_cliente")

        cliente_usuario = (
            payload.get("id_cliente")
            or payload.get("cliente_id")
            or cliente_admin
        )

        # =========================================
        # AISLAMIENTO TENANT
        # =========================================

        if rol_admin != "master_admin":
            cliente_usuario = cliente_admin

        if not cliente_usuario:
            return {
                "ok": False,
                "error": "Tenant requerido para crear usuario",
                "codigo": "TENANT_REQUIRED",
            }

        # =========================================
        # VALIDAR EXISTENCIA
        # =========================================

        url_check = f"{SUPABASE_URL}/rest/v1/argo_usuarios?email=eq.{email}&select=email"

        response_check = requests.get(
            url_check,
            headers=_headers(),
            timeout=30
        )

        if response_check.status_code != 200:
            return {
                "ok": False,
                "error": "Error validando usuario"
            }

        existe = response_check.json()

        if existe:
            return {
                "ok": False,
                "error": "Usuario ya existe"
            }

        # =========================================
        # CREAR USUARIO
        # =========================================

        nuevo_usuario = {
            "email": email,
            "password": hash_password(password),
            "nombre": nombre,
            "rol": rol,
            "id_cliente": cliente_usuario,
            "activo": activo,
        }

        url_insert = f"{SUPABASE_URL}/rest/v1/argo_usuarios"

        headers = _headers()
        headers["Prefer"] = "return=representation"

        response_insert = requests.post(
            url_insert,
            headers=headers,
            json=nuevo_usuario,
            timeout=30
        )

        if response_insert.status_code not in [200, 201]:
            return {
                "ok": False,
                "error": response_insert.text
            }

        creado = response_insert.json()
        usuario_creado = creado[0]

        auditoria_creacion = guardar_auditoria_admin(
            accion="CREAR_USUARIO",
            actor_email=x_usuario_email,
            actor_rol=rol_admin,
            tenant=cliente_usuario,
            objetivo_email=email,
            detalle={
                "nombre": nombre,
                "rol": rol,
                "activo": activo,
                "plan_saas": (
                    usuario_creado.get("plan_saas")
                    or usuario_creado.get("plan")
                ),
            },
        )

        return {
            "ok": True,
            "mensaje": "Usuario creado correctamente",
            "usuario": usuario_creado,
            "auditoria": auditoria_creacion,
            "rbac": {
                "creado_por": x_usuario_email,
                "rol_admin": rol_admin,
                "tenant": cliente_usuario,
            }
        }

    except Exception as e:
        return {
            "ok": False,
            "error": str(e)
        }
# =========================================================
# SAAS ADMIN - LISTAR USUARIOS
# =========================================================

def guardar_auditoria_admin(
    accion: str,
    actor_email: str = None,
    actor_rol: str = None,
    tenant: str = None,
    objetivo_email: str = None,
    detalle: dict = None,
):
    from datetime import datetime, timezone
    import json
    import os

    evento = {
        "fecha": datetime.now(timezone.utc).isoformat(),
        "accion": accion,
        "actor_email": actor_email,
        "actor_rol": actor_rol,
        "tenant": tenant,
        "objetivo_email": objetivo_email,
        "detalle": detalle or {},
    }

    supabase_ok = False
    supabase_error = None

    try:
        if supabase_config_ok():
            url = f"{SUPABASE_URL}/rest/v1/argo_auditoria_admin"
            headers = _headers()
            headers["Prefer"] = "return=representation"

            resp = requests.post(
                url,
                headers=headers,
                json=evento,
                timeout=30
            )

            if resp.status_code in [200, 201]:
                supabase_ok = True
            else:
                supabase_error = f"{resp.status_code} - {resp.text}"

    except Exception as e:
        supabase_error = str(e)

    try:
        os.makedirs("logs", exist_ok=True)
        with open("logs/admin_audit.jsonl", "a", encoding="utf-8") as f:
            f.write(json.dumps(evento, ensure_ascii=False) + "\n")
    except Exception:
        pass

    return {
        "ok": supabase_ok,
        "evento": evento,
        "storage": "supabase" if supabase_ok else "jsonl_fallback",
        "supabase_error": supabase_error,
    }


def actualizar_usuario_rbac(email: str, cambios: dict):

    if not supabase_config_ok():
        return {
            "ok": False,
            "error": "Supabase no configurado"
        }

    try:

        url = (
            f"{SUPABASE_URL}"
            f"/rest/v1/argo_usuarios"
            f"?email=eq.{email}"
        )

        resp = requests.patch(
            url,
            headers=_headers(),
            json=cambios,
            timeout=20
        )

        if resp.status_code not in [200, 204]:
            return {
                "ok": False,
                "error": f"Error actualizando usuario: {resp.status_code} - {resp.text}"
            }

        return {
            "ok": True
        }

    except Exception as e:
        return {
            "ok": False,
            "error": str(e)
        }
@app.get("/argo/admin/usuarios")
async def listar_usuarios_admin(
    cliente_id: str = None,
    x_usuario_email: str = Header(default=None),
):
    try:

        permitido, usuario_admin, motivo = validar_permiso_rbac(
            email=x_usuario_email,
            roles_permitidos=ROLES_ADMIN_CLIENTES,
            cliente_id=cliente_id
        )

        if not permitido:
            return JSONResponse(
                status_code=403,
                content={
                    "ok": False,
                    "error": motivo
                }
            )

        rol_admin = str(usuario_admin.get("rol") or "").lower()
        cliente_admin = usuario_admin.get("id_cliente")

        tenant_final = cliente_id or cliente_admin

        # =========================================
        # HARD TENANT ISOLATION
        # =========================================

        if rol_admin != "master_admin":
            tenant_final = cliente_admin

        url = (
            f"{SUPABASE_URL}/rest/v1/argo_usuarios"
            f"?id_cliente=eq.{tenant_final}"
            f"&select=nombre,email,rol,activo,id_cliente,plan_saas"
            f"&order=nombre.asc"
        )

        response = requests.get(
            url,
            headers=_headers(),
            timeout=20
        )

        if response.status_code >= 300:
            return JSONResponse(
                status_code=500,
                content={
                    "ok": False,
                    "error": "Error consultando usuarios",
                    "detalle": response.text
                }
            )

        usuarios = response.json()

        return {
            "ok": True,
            "usuarios": usuarios,
            "total": len(usuarios),
            "tenant": tenant_final,
            "audit": {
                "consultado_por": x_usuario_email,
                "rol_admin": rol_admin
            }
        }

    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={
                "ok": False,
                "error": str(e)
            }
        )
# =========================================================
# SAAS ADMIN - ACTIVAR / DESACTIVAR USUARIO
# =========================================================

@app.patch("/argo/admin/usuario/activo")
async def cambiar_estado_usuario_admin(
    payload: dict = Body(...),
    x_usuario_email: str = Header(default=None),
):
    try:

        email_objetivo = str(payload.get("email") or "").strip().lower()
        activo = bool(payload.get("activo"))

        if not email_objetivo:
            return {
                "ok": False,
                "error": "Email objetivo requerido"
            }

        permitido, usuario_admin, motivo = validar_permiso_rbac(
            email=x_usuario_email,
            roles_permitidos=ROLES_ADMIN_CLIENTES,
        )

        if not permitido:
            return JSONResponse(
                status_code=403,
                content={
                    "ok": False,
                    "error": motivo
                }
            )

        rol_admin = str(usuario_admin.get("rol") or "").lower()
        cliente_admin = usuario_admin.get("id_cliente")

        usuario_objetivo = obtener_usuario_rbac(email_objetivo)

        if not usuario_objetivo:
            return {
                "ok": False,
                "error": "Usuario objetivo no encontrado"
            }

        if rol_admin != "master_admin" and usuario_objetivo.get("id_cliente") != cliente_admin:
            return JSONResponse(
                status_code=403,
                content={
                    "ok": False,
                    "error": "No puedes modificar usuarios de otro tenant"
                }
            )

        resultado = actualizar_usuario_rbac(
            email=email_objetivo,
            cambios={
                "activo": activo
            }
        )

        if not resultado.get("ok"):
            return resultado

        guardar_auditoria_admin(
            accion="CAMBIO_ESTADO_USUARIO" if activo else "ELIMINAR_ACCESO_USUARIO",
            actor_email=x_usuario_email,
            actor_rol=rol_admin,
            tenant=usuario_objetivo.get("id_cliente"),
            objetivo_email=email_objetivo,
            detalle={
                "activo": activo
            }
        )

        return {
            "ok": True,
            "mensaje": "Estado de usuario actualizado",
            "usuario": {
                "email": email_objetivo,
                "activo": activo
            },
            "audit": {
                "actualizado_por": x_usuario_email,
                "rol_admin": rol_admin,
                "tenant": usuario_objetivo.get("id_cliente")
            }
        }

    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={
                "ok": False,
                "error": str(e)
            }
        )

# =========================================================
# SAAS ADMIN - ELIMINAR ACCESO ENTERPRISE
# =========================================================

@app.patch("/argo/admin/usuario/eliminar_acceso")
async def eliminar_acceso_usuario_admin(
    payload: dict = Body(...),
    x_usuario_email: str = Header(default=None),
):
    try:
        email_objetivo = str(payload.get("email") or "").strip().lower()

        if not email_objetivo:
            return {"ok": False, "error": "Email requerido"}

        permitido, usuario_admin, motivo = validar_permiso_rbac(
            email=x_usuario_email,
            roles_permitidos=ROLES_ADMIN_CLIENTES,
        )

        if not permitido:
            return {"ok": False, "error": motivo}

        usuario_objetivo = obtener_usuario_rbac(email_objetivo)

        if not usuario_objetivo:
            return {"ok": False, "error": "Usuario objetivo no encontrado"}

        rol_admin = str(usuario_admin.get("rol") or "").lower()
        cliente_admin = usuario_admin.get("id_cliente")

        if rol_admin != "master_admin" and usuario_objetivo.get("id_cliente") != cliente_admin:
            return {"ok": False, "error": "Acceso denegado por aislamiento tenant"}

        if email_objetivo == str(x_usuario_email or "").strip().lower():
            return {"ok": False, "error": "No puedes eliminar tu propio acceso"}

        resultado = actualizar_usuario_rbac(
            email=email_objetivo,
            cambios={
                "activo": False,
                "rol": "operador",
            }
        )

        if not resultado.get("ok"):
            return resultado

        guardar_auditoria_admin(
            accion="ELIMINAR_ACCESO_ENTERPRISE",
            actor_email=x_usuario_email,
            actor_rol=rol_admin,
            tenant=usuario_objetivo.get("id_cliente"),
            objetivo_email=email_objetivo,
            detalle={
                "activo": False,
                "rol_anterior": usuario_objetivo.get("rol"),
                "rol_nuevo": "operador",
                "motivo": "Acceso enterprise eliminado por administrador",
            }
        )

        return {
            "ok": True,
            "mensaje": "Acceso eliminado correctamente",
            "email": email_objetivo,
        }

    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"ok": False, "error": str(e)}
        )


# =========================================================
# SAAS ADMIN - CAMBIAR ROL USUARIO
# =========================================================

@app.patch("/argo/admin/usuario/rol")
async def cambiar_rol_usuario_admin(
    payload: dict = Body(...),
    x_usuario_email: str = Header(default=None),
):
    try:

        email_objetivo = str(payload.get("email") or "").strip().lower()
        nuevo_rol = str(payload.get("rol") or "").strip().lower()

        roles_validos = [
            "operador",
            "supervisor",
            "admin_cliente",
        ]

        if nuevo_rol not in roles_validos:
            return {
                "ok": False,
                "error": "Rol inválido"
            }

        permitido, usuario_admin, motivo = validar_permiso_rbac(
            email=x_usuario_email,
            roles_permitidos=ROLES_ADMIN_CLIENTES,
        )

        if not permitido:
            return JSONResponse(
                status_code=403,
                content={
                    "ok": False,
                    "error": motivo
                }
            )

        rol_admin = str(usuario_admin.get("rol") or "").lower()
        cliente_admin = usuario_admin.get("id_cliente")

        usuario_objetivo = obtener_usuario_rbac(email_objetivo)

        if not usuario_objetivo:
            return {
                "ok": False,
                "error": "Usuario objetivo no encontrado"
            }

        if (
            rol_admin != "master_admin"
            and usuario_objetivo.get("id_cliente") != cliente_admin
        ):
            return JSONResponse(
                status_code=403,
                content={
                    "ok": False,
                    "error": "No puedes modificar usuarios de otro tenant"
                }
            )

        resultado = actualizar_usuario_rbac(
            email=email_objetivo,
            cambios={
                "rol": nuevo_rol
            }
        )

        if not resultado.get("ok"):
            return resultado

        guardar_auditoria_admin(
            accion="CAMBIO_ROL_USUARIO",
            actor_email=x_usuario_email,
            actor_rol=rol_admin,
            tenant=usuario_objetivo.get("id_cliente"),
            objetivo_email=email_objetivo,
            detalle={
                "rol_anterior": usuario_objetivo.get("rol"),
                "rol_nuevo": nuevo_rol
            }
        )

        return {
            "ok": True,
            "mensaje": "Rol actualizado",
            "usuario": {
                "email": email_objetivo,
                "rol": nuevo_rol
            },
            "audit": {
                "accion": "CAMBIO_ROL_USUARIO",
                "actualizado_por": x_usuario_email,
                "rol_admin": rol_admin,
                "tenant": cliente_admin,
            }
        }

    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={
                "ok": False,
                "error": str(e)
            }
        )



# =========================================================
# SAAS ADMIN - CAMBIAR PLAN SAAS USUARIO/TENANT
# =========================================================

@app.patch("/argo/admin/usuario/plan")
async def cambiar_plan_usuario_admin(
    payload: dict = Body(...),
    x_usuario_email: str = Header(default=None),
):
    try:
        email_objetivo = str(payload.get("email") or "").strip().lower()
        nuevo_plan = normalizar_plan(payload.get("plan_saas") or payload.get("plan") or "")

        if not email_objetivo:
            return {
                "ok": False,
                "error": "Email requerido"
            }

        if nuevo_plan not in PLANES_SAAS:
            return {
                "ok": False,
                "error": "Plan inválido"
            }

        permitido, usuario_admin, motivo = validar_permiso_rbac(
            email=x_usuario_email,
            roles_permitidos=ROLES_ADMIN_CLIENTES,
        )

        if not permitido:
            return JSONResponse(
                status_code=403,
                content={
                    "ok": False,
                    "error": motivo,
                    "codigo": "RBAC_DENY"
                }
            )

        rol_admin = str(usuario_admin.get("rol") or "").lower()
        cliente_admin = usuario_admin.get("id_cliente")

        usuario_objetivo = obtener_usuario_rbac(email_objetivo)

        if not usuario_objetivo:
            return {
                "ok": False,
                "error": "Usuario objetivo no encontrado"
            }

        if (
            rol_admin != "master_admin"
            and usuario_objetivo.get("id_cliente") != cliente_admin
        ):
            return JSONResponse(
                status_code=403,
                content={
                    "ok": False,
                    "error": "No puedes modificar usuarios de otro tenant",
                    "codigo": "TENANT_DENY"
                }
            )

        plan_anterior = (
            usuario_objetivo.get("plan_saas")
            or usuario_objetivo.get("plan")
            or usuario_objetivo.get("tipo_plan")
            or "ENTERPRISE"
        )

        resultado = actualizar_usuario_rbac(
            email_objetivo,
            {
                "plan_saas": nuevo_plan
            }
        )

        if not resultado.get("ok"):
            return resultado

        guardar_auditoria_admin(
            accion="cambiar_plan_saas_usuario",
            actor_email=x_usuario_email,
            actor_rol=rol_admin,
            tenant=usuario_objetivo.get("id_cliente"),
            objetivo_email=email_objetivo,
            detalle={
                "plan_anterior": plan_anterior,
                "plan_nuevo": nuevo_plan,
                "modulos_plan": PLANES_SAAS[nuevo_plan]["modulos"],
                "limites": PLANES_SAAS[nuevo_plan]["limites"],
            }
        )

        return {
            "ok": True,
            "mensaje": "Plan SaaS actualizado correctamente",
            "email": email_objetivo,
            "plan_anterior": plan_anterior,
            "plan_saas": nuevo_plan,
            "plan": PLANES_SAAS[nuevo_plan],
        }

    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={
                "ok": False,
                "error": str(e)
            }
        )


# tenant billing endpoint enabled
# =========================================================
# SAAS ADMIN - LICENCIA / BILLING TENANT
# =========================================================

@app.patch("/argo/admin/tenant/licencia")
async def actualizar_licencia_tenant_admin(
    payload: dict = Body(...),
    x_usuario_email: str = Header(default=None),
):
    try:
        tenant_id = str(
            payload.get("tenant")
            or payload.get("id_cliente")
            or payload.get("cliente_id")
            or ""
        ).strip()

        nuevo_estado = str(
            payload.get("estado_licencia")
            or payload.get("estado")
            or "ACTIVA"
        ).strip().upper()

        fecha_vencimiento = (
            payload.get("fecha_vencimiento")
            or payload.get("licencia_hasta")
            or payload.get("saas_expira")
        )

        if not tenant_id:
            return {
                "ok": False,
                "error": "Tenant requerido"
            }

        if nuevo_estado not in ["ACTIVA", "POR_VENCER", "SUSPENDIDA", "VENCIDA", "BLOQUEADA", "CANCELADA"]:
            return {
                "ok": False,
                "error": "Estado de licencia inválido",
                "permitidos": ["ACTIVA", "POR_VENCER", "SUSPENDIDA", "VENCIDA", "BLOQUEADA", "CANCELADA"]
            }

        permitido, usuario_admin, motivo = validar_permiso_rbac(
            email=x_usuario_email,
            roles_permitidos={"master_admin"},
        )

        if not permitido:
            return JSONResponse(
                status_code=403,
                content={
                    "ok": False,
                    "error": motivo,
                    "codigo": "MASTER_ADMIN_REQUIRED"
                }
            )

        update_data = {
            "estado_licencia": nuevo_estado
        }

        if fecha_vencimiento:
            update_data["fecha_vencimiento"] = fecha_vencimiento

        url = (
            f"{SUPABASE_URL}/rest/v1/argo_usuarios"
            f"?id_cliente=eq.{tenant_id}"
        )

        response = requests.patch(
            url,
            headers={
                **_headers(),
                "Prefer": "return=representation"
            },
            json=update_data,
            timeout=30
        )

        if response.status_code not in [200, 204]:
            return JSONResponse(
                status_code=500,
                content={
                    "ok": False,
                    "error": "No se pudo actualizar licencia tenant",
                    "detalle": response.text
                }
            )

        actualizados = response.json() if response.text else []

        guardar_auditoria_admin(
            accion="actualizar_licencia_tenant",
            actor_email=x_usuario_email,
            actor_rol="master_admin",
            tenant=tenant_id,
            detalle={
                "estado_licencia": nuevo_estado,
                "fecha_vencimiento": fecha_vencimiento,
                "usuarios_actualizados": len(actualizados),
            }
        )

        return {
            "ok": True,
            "mensaje": "Licencia tenant actualizada",
            "tenant": tenant_id,
            "estado_licencia": nuevo_estado,
            "fecha_vencimiento": fecha_vencimiento,
            "usuarios_actualizados": len(actualizados),
            "data": actualizados,
        }

    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={
                "ok": False,
                "error": str(e)
            }
        )




# =========================================================
# SAAS ADMIN - PLAN TENANT
# =========================================================

@app.patch("/argo/admin/tenant/plan")
async def actualizar_plan_tenant_admin(
    payload: dict = Body(...),
    x_usuario_email: str = Header(default=None),
):
    try:
        tenant_id = str(
            payload.get("tenant")
            or payload.get("id_cliente")
            or payload.get("cliente_id")
            or ""
        ).strip()

        nuevo_plan = str(
            payload.get("plan")
            or payload.get("plan_saas")
            or ""
        ).strip().upper()

        if not tenant_id:
            return {
                "ok": False,
                "error": "Tenant requerido"
            }

        if nuevo_plan not in PLANES_SAAS:
            return {
                "ok": False,
                "error": "Plan SaaS inválido",
                "permitidos": list(PLANES_SAAS.keys())
            }

        permitido, usuario_admin, motivo = validar_permiso_rbac(
            email=x_usuario_email,
            roles_permitidos={"master_admin"},
        )

        if not permitido:
            return JSONResponse(
                status_code=403,
                content={
                    "ok": False,
                    "error": motivo,
                    "codigo": "MASTER_ADMIN_REQUIRED"
                }
            )

        # Obtener usuarios actuales para auditoría
        consulta_url = (
            f"{SUPABASE_URL}/rest/v1/argo_usuarios"
            f"?id_cliente=eq.{tenant_id}&select=email,plan_saas,plan,tipo_plan,id_cliente"
        )

        consulta = requests.get(
            consulta_url,
            headers=_headers(),
            timeout=30
        )

        usuarios_antes = consulta.json() if consulta.status_code == 200 else []

        planes_anteriores = sorted(list(set([
            str(
                u.get("plan_saas")
                or u.get("plan")
                or u.get("tipo_plan")
                or "BASIC"
            ).upper()
            for u in usuarios_antes
            if isinstance(u, dict)
        ])))

        update_data = {
            "plan_saas": nuevo_plan
        }

        update_url = (
            f"{SUPABASE_URL}/rest/v1/argo_usuarios"
            f"?id_cliente=eq.{tenant_id}"
        )

        response = requests.patch(
            update_url,
            headers={
                **_headers(),
                "Prefer": "return=representation"
            },
            json=update_data,
            timeout=30
        )

        if response.status_code not in [200, 204]:
            return JSONResponse(
                status_code=500,
                content={
                    "ok": False,
                    "error": "No se pudo actualizar plan tenant",
                    "detalle": response.text
                }
            )

        actualizados = response.json() if response.text else []

        guardar_auditoria_admin(
            accion="actualizar_plan_tenant",
            actor_email=x_usuario_email,
            actor_rol="master_admin",
            tenant=tenant_id,
            detalle={
                "planes_anteriores": planes_anteriores,
                "plan_nuevo": nuevo_plan,
                "usuarios_actualizados": len(actualizados),
                "modulos_plan": PLANES_SAAS[nuevo_plan]["modulos"],
                "limites": PLANES_SAAS[nuevo_plan]["limites"],
            }
        )

        return {
            "ok": True,
            "mensaje": "Plan tenant actualizado",
            "tenant": tenant_id,
            "plan_saas": nuevo_plan,
            "plan": PLANES_SAAS[nuevo_plan],
            "planes_anteriores": planes_anteriores,
            "usuarios_actualizados": len(actualizados),
            "data": actualizados,
        }

    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={
                "ok": False,
                "error": str(e)
            }
        )


# =========================================================
# SAAS ADMIN - RESET PASSWORD USUARIO
# =========================================================

@app.patch("/argo/admin/usuario/reset_password")
async def reset_password_usuario_admin(
    payload: dict = Body(...),
    x_usuario_email: str = Header(default=None),
):
    try:

        email_objetivo = str(payload.get("email") or "").strip().lower()
        nuevo_password = str(payload.get("password") or "").strip()

        if not email_objetivo or not nuevo_password:
            return {
                "ok": False,
                "error": "Email y nuevo password son requeridos"
            }

        if len(nuevo_password) < 6:
            return {
                "ok": False,
                "error": "El password debe tener al menos 6 caracteres"
            }

        permitido, usuario_admin, motivo = validar_permiso_rbac(
            email=x_usuario_email,
            roles_permitidos=ROLES_ADMIN_CLIENTES,
        )

        if not permitido:
            return JSONResponse(
                status_code=403,
                content={
                    "ok": False,
                    "error": motivo
                }
            )

        rol_admin = str(usuario_admin.get("rol") or "").lower()
        cliente_admin = usuario_admin.get("id_cliente")

        usuario_objetivo = obtener_usuario_rbac(email_objetivo)

        if not usuario_objetivo:
            return {
                "ok": False,
                "error": "Usuario objetivo no encontrado"
            }

        if (
            rol_admin != "master_admin"
            and usuario_objetivo.get("id_cliente") != cliente_admin
        ):
            return JSONResponse(
                status_code=403,
                content={
                    "ok": False,
                    "error": "No puedes modificar usuarios de otro tenant"
                }
            )

        resultado = actualizar_usuario_rbac(
            email=email_objetivo,
            cambios={
                "password": hash_password(nuevo_password)
            }
        )

        if not resultado.get("ok"):
            return resultado

        guardar_auditoria_admin(
            accion="RESET_PASSWORD_USUARIO",
            actor_email=x_usuario_email,
            actor_rol=rol_admin,
            tenant=usuario_objetivo.get("id_cliente"),
            objetivo_email=email_objetivo,
            detalle={
                "password_reseteado": True
            }
        )

        return {
            "ok": True,
            "mensaje": "Password actualizado correctamente",
            "usuario": {
                "email": email_objetivo
            },
            "audit": {
                "accion": "RESET_PASSWORD_USUARIO",
                "actualizado_por": x_usuario_email,
                "rol_admin": rol_admin,
                "tenant": usuario_objetivo.get("id_cliente")
            }
        }

    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={
                "ok": False,
                "error": str(e)
            }
        )



# =========================================================
# SAAS ADMIN - CONSULTAR AUDITORIA ADMIN
# =========================================================



# =========================================================
# ENTERPRISE ACTIVITY FEED / AUDITORIA AVANZADA
# =========================================================

@app.get("/argo/admin/activity_feed")
async def activity_feed_enterprise(
    limit: int = 100,
    accion: str = None,
    actor_email: str = None,
    tenant: str = None,
    x_usuario_email: str = Header(default=None),
):
    try:
        permitido, usuario_admin, motivo = validar_permiso_rbac(
            email=x_usuario_email,
            roles_permitidos=ROLES_ADMIN_CLIENTES,
        )

        if not permitido:
            return JSONResponse(
                status_code=403,
                content={
                    "ok": False,
                    "error": motivo,
                    "codigo": "RBAC_DENY"
                }
            )

        import json
        import os
        from collections import Counter

        rol_admin = str(usuario_admin.get("rol") or "").lower()
        tenant_admin = usuario_admin.get("id_cliente")
        tenant_final = tenant or tenant_admin
        limit_final = max(1, min(int(limit or 100), 500))

        logs = []
        fuente = "supabase"

        if supabase_config_ok():
            query = (
                f"{SUPABASE_URL}/rest/v1/argo_auditoria_admin"
                f"?select=fecha,accion,actor_email,actor_rol,tenant,objetivo_email,detalle"
                f"&order=fecha.desc"
                f"&limit={limit_final}"
            )

            if rol_admin != "master_admin":
                query += f"&tenant=eq.{tenant_admin}"
            elif tenant_final:
                query += f"&tenant=eq.{tenant_final}"

            if accion:
                query += f"&accion=eq.{accion}"

            if actor_email:
                query += f"&actor_email=eq.{actor_email}"

            resp = requests.get(
                query,
                headers=_headers(),
                timeout=30
            )

            if resp.status_code == 200:
                data = resp.json() or []
                for evento in data:
                    detalle = evento.get("detalle") or {}
                    logs.append({
                        "fecha": evento.get("fecha"),
                        "accion": evento.get("accion"),
                        "actor_email": evento.get("actor_email"),
                        "actor_rol": evento.get("actor_rol"),
                        "tenant": evento.get("tenant"),
                        "objetivo_email": evento.get("objetivo_email"),
                        "detalle": detalle,
                        "id_operacion": detalle.get("id_operacion"),
                        "severidad": detalle.get("severidad"),
                        "estado_incidencia": detalle.get("estado_incidencia"),
                        "modulo": (
                            "DASHBOARD_PRO"
                            if str(evento.get("accion") or "").startswith("dashboard_pro")
                            else "ADMIN"
                        ),
                    })

        if not logs:
            fuente = "jsonl_fallback"
            path_logs = "logs/admin_audit.jsonl"

            if os.path.exists(path_logs):
                with open(path_logs, "r", encoding="utf-8") as f:
                    for line in f:
                        try:
                            evento = json.loads(line.strip())
                        except Exception:
                            continue

                        evento_tenant = evento.get("tenant")

                        if rol_admin != "master_admin":
                            if evento_tenant != tenant_admin:
                                continue
                        elif tenant_final:
                            if evento_tenant != tenant_final:
                                continue

                        if accion and evento.get("accion") != accion:
                            continue

                        if actor_email and evento.get("actor_email") != actor_email:
                            continue

                        detalle = evento.get("detalle") or {}

                        logs.append({
                            "fecha": evento.get("fecha"),
                            "accion": evento.get("accion"),
                            "actor_email": evento.get("actor_email"),
                            "actor_rol": evento.get("actor_rol"),
                            "tenant": evento.get("tenant"),
                            "objetivo_email": evento.get("objetivo_email"),
                            "detalle": detalle,
                            "id_operacion": detalle.get("id_operacion"),
                            "severidad": detalle.get("severidad"),
                            "estado_incidencia": detalle.get("estado_incidencia"),
                            "modulo": (
                                "DASHBOARD_PRO"
                                if str(evento.get("accion") or "").startswith("dashboard_pro")
                                else "ADMIN"
                            ),
                        })

        logs = sorted(logs, key=lambda x: x.get("fecha") or "", reverse=True)

        acciones = Counter([x.get("accion") or "SIN_ACCION" for x in logs])
        modulos = Counter([x.get("modulo") or "SIN_MODULO" for x in logs])
        actores = Counter([x.get("actor_email") or "SIN_ACTOR" for x in logs])

        return {
            "ok": True,
            "tenant": tenant_final,
            "total": len(logs),
            "logs": logs[:limit_final],
            "fuente": fuente,
            "resumen": {
                "acciones": dict(acciones),
                "modulos": dict(modulos),
                "actores": dict(actores),
            }
        }

    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={
                "ok": False,
                "error": str(e),
                "modulo": "activity_feed_enterprise"
            }
        )


@app.get("/argo/admin/auditoria")
async def consultar_auditoria_admin(
    limit: int = 50,
    x_usuario_email: str = Header(default=None),
):
    try:
        permitido, usuario_admin, motivo = validar_permiso_rbac(
            email=x_usuario_email,
            roles_permitidos=ROLES_ADMIN_CLIENTES,
        )

        if not permitido:
            return JSONResponse(
                status_code=403,
                content={
                    "ok": False,
                    "error": motivo
                }
            )

        import json
        import os

        rol_admin = str(usuario_admin.get("rol") or "").lower()
        tenant_admin = usuario_admin.get("id_cliente")

        path_logs = "logs/admin_audit.jsonl"

        if not os.path.exists(path_logs):
            return {
                "ok": True,
                "logs": [],
                "total": 0,
                "tenant": tenant_admin
            }

        eventos = []

        with open(path_logs, "r", encoding="utf-8") as f:
            for line in f:
                try:
                    evento = json.loads(line.strip())
                    if rol_admin == "master_admin" or evento.get("tenant") == tenant_admin:
                        eventos.append(evento)
                except Exception:
                    continue

        eventos = list(reversed(eventos))[: max(1, min(limit, 200))]

        return {
            "ok": True,
            "logs": eventos,
            "total": len(eventos),
            "tenant": tenant_admin
        }

    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={
                "ok": False,
                "error": str(e)
            }
        )




# =========================================================
# MASTER ADMIN COMMERCIAL DASHBOARD
# =========================================================

@app.get("/argo/master/dashboard")
async def argo_master_dashboard(
    request: Request
):

    try:

        x_usuario_email = request.headers.get("x-usuario-email")

        usuario = obtener_usuario_rbac(x_usuario_email)

        if not usuario:
            return {
                "ok": False,
                "error": "Usuario no encontrado"
            }

        rol = str(usuario.get("rol") or "").lower()

        if rol != "master_admin":
            return {
                "ok": False,
                "error": "Acceso restringido MASTER_ADMIN"
            }

        # =====================================================
        # CARGAR TENANTS
        # =====================================================

        url = (
            f"{SUPABASE_URL}/rest/v1/argo_usuarios"
            f"?select=*"
        )

        response = requests.get(
            url,
            headers=_headers(),
            timeout=30
        )

        if response.status_code != 200:
            return {
                "ok": False,
                "error": "No se pudo cargar informacion SaaS"
            }

        usuarios = response.json()

        tenants = {}
        total_usuarios = len(usuarios)

        for u in usuarios:

            tenant = u.get("id_cliente") or "SIN_TENANT"

            if tenant not in tenants:

                plan = obtener_plan_saas(u)

                licencia_info = validar_licencia_saas(u)

                tenants[tenant] = {
                    "tenant": tenant,
                    "plan": plan.get("codigo"),
                    "usuarios": 0,
                    "estado_licencia": (
                        licencia_info.get("estado_licencia")
                        or u.get("estado_licencia")
                        or "ACTIVA"
                    ),
                    "dias_restantes": licencia_info.get("dias_restantes"),
                    "fecha_vencimiento": licencia_info.get("fecha_vencimiento"),
                    "operaciones_mes": 0,
                }

            tenants[tenant]["usuarios"] += 1

        # =====================================================
        # OPERACIONES POR TENANT
        # =====================================================

        historial_raw = obtener_historial()

        if isinstance(historial_raw, dict):
            historial = (
                historial_raw.get("operaciones")
                or historial_raw.get("historial")
                or historial_raw.get("data")
                or []
            )
        elif isinstance(historial_raw, list):
            historial = historial_raw
        else:
            historial = []

        total_operaciones = len(historial)

        for op in historial:

            if not isinstance(op, dict):
                continue

            tenant = (
                op.get("cliente_id")
                or op.get("id_cliente")
                or "SIN_TENANT"
            )

            if tenant in tenants:
                tenants[tenant]["operaciones_mes"] += 1

        # =====================================================
        # METRICAS GLOBALES
        # =====================================================

        activos = 0
        suspendidos = 0

        planes = {
            "BASIC": 0,
            "PRO": 0,
            "ENTERPRISE": 0,
            "CUSTOM": 0,
        }

        revenue_estimado = 0

        precios = {
            "BASIC": 99,
            "PRO": 399,
            "ENTERPRISE": 1299,
            "CUSTOM": 3000,
        }

        for t in tenants.values():

            plan = str(t.get("plan") or "BASIC").upper()

            if plan not in planes:
                planes[plan] = 0

            planes[plan] += 1

            revenue_estimado += precios.get(plan, 0)

            estado = str(
                t.get("estado_licencia") or "ACTIVA"
            ).upper()

            if estado in ["SUSPENDIDA", "VENCIDA"]:
                suspendidos += 1
            else:
                activos += 1

        # =====================================================
        # TOP TENANTS
        # =====================================================

        top_tenants = sorted(
            list(tenants.values()),
            key=lambda x: x.get("operaciones_mes", 0),
            reverse=True
        )[:10]

        # executive feed enabled
# =====================================================
        # EXECUTIVE ACTIVITY FEED
        # =====================================================

        activity_feed = []
        operaciones_criticas = 0
        aprobaciones_total = 0

        for op in historial[-50:]:

            if not isinstance(op, dict):
                continue

            severidad = str(
                op.get("semaforo_operacion")
                or op.get("riesgo_global")
                or "MEDIA"
            ).upper()

            tenant = (
                op.get("cliente_nombre")
                or op.get("cliente_id")
                or "SIN_TENANT"
            )

            if severidad in ["ALTA", "CRITICA", "CRÍTICA"]:
                operaciones_criticas += 1

            if op.get("aprobada") is True:
                aprobaciones_total += 1

            activity_feed.append({
                "tenant": tenant,
                "operacion": op.get("id_operacion"),
                "riesgo": severidad,
                "aprobada": op.get("aprobada", False),
                "fecha": (
                    op.get("fecha_aprobacion")
                    or op.get("fecha")
                    or op.get("timestamp_local")
                    or op.get("created_at")
                    or "N/D"
                )
            })

        ultimas_aprobaciones = [
            x for x in activity_feed
            if x.get("aprobada")
        ][-10:]

        licencias_por_vencer = sorted(
            [
                t for t in tenants.values()
                if t.get("dias_restantes") is not None
                and t.get("dias_restantes") <= 30
            ],
            key=lambda x: x.get("dias_restantes", 999999)
        )

        tenants_en_riesgo = [
            t for t in tenants.values()
            if str(t.get("estado_licencia") or "").upper()
            in ["POR_VENCER", "SUSPENDIDA", "VENCIDA", "BLOQUEADA"]
            or (
                t.get("dias_restantes") is not None
                and t.get("dias_restantes") <= 7
            )
        ]

        if operaciones_criticas > 0 or tenants_en_riesgo:
            riesgo_global = "ALTO"
        elif licencias_por_vencer:
            riesgo_global = "MEDIO"
        else:
            riesgo_global = "BAJO"

        operaciones_por_dia_map = {
            "LUN": 0,
            "MAR": 0,
            "MIE": 0,
            "JUE": 0,
            "VIE": 0,
            "SAB": 0,
            "DOM": 0,
        }

        riesgos_map = {
            "BAJO": 0,
            "MEDIO": 0,
            "ALTO": 0,
            "CRITICO": 0,
        }

        for op in historial:
            if not isinstance(op, dict):
                continue

            fecha_op = (
                op.get("fecha_aprobacion")
                or op.get("fecha")
                or op.get("timestamp_local")
                or op.get("created_at")
                or op.get("timestamp")
            )

            try:
                if fecha_op:
                    fecha_dt = datetime.fromisoformat(
                        str(fecha_op).replace("Z", "").split(".")[0]
                    )
                    dia_idx = fecha_dt.weekday()
                    dia_nombre = ["LUN", "MAR", "MIE", "JUE", "VIE", "SAB", "DOM"][dia_idx]
                    operaciones_por_dia_map[dia_nombre] += 1
            except Exception:
                pass

            riesgo_op = str(
                op.get("semaforo_operacion")
                or op.get("riesgo_global")
                or "MEDIO"
            ).upper().replace("CRÍTICA", "CRITICO").replace("CRITICA", "CRITICO")

            if riesgo_op in ["BAJA", "OK", "OPERABLE"]:
                riesgo_op = "BAJO"
            elif riesgo_op in ["MEDIA", "REVISION", "REVISIÓN"]:
                riesgo_op = "MEDIO"
            elif riesgo_op in ["ALTA"]:
                riesgo_op = "ALTO"
            elif riesgo_op not in riesgos_map:
                riesgo_op = "MEDIO"

            riesgos_map[riesgo_op] += 1

        # =====================================================
        # KPIS DE VALIDACION OPERACIONAL
        # =====================================================

        coberturas_validacion = []
        operaciones_verdes = 0
        operaciones_amarillas = 0
        operaciones_rojas = 0
        operaciones_sin_control = 0
        campos_no_verificables_total = 0

        cobertura_por_tenant_acumulada = {}

        for op in historial:

            if not isinstance(op, dict):
                continue

            resumen_op = (
                op.get("resumen_operativo")
                or op.get("control")
                or {}
            )

            if not isinstance(resumen_op, dict):
                resumen_op = {}

            semaforo_op = str(
                resumen_op.get("semaforo")
                or resumen_op.get("estado")
                or "SIN_CONTROL"
            ).upper()

            cobertura_op = resumen_op.get("cobertura_validacion_pct")

            try:
                cobertura_num = float(cobertura_op)
            except (TypeError, ValueError):
                cobertura_num = None

            if cobertura_num is not None:
                coberturas_validacion.append(cobertura_num)

            if semaforo_op == "VERDE":
                operaciones_verdes += 1
            elif semaforo_op == "AMARILLO":
                operaciones_amarillas += 1
            elif semaforo_op == "ROJO":
                operaciones_rojas += 1
            else:
                operaciones_sin_control += 1

            conteo_op = resumen_op.get("conteo") or {}

            campos_no_verificables = (
                resumen_op.get("campos_no_verificables")
                if resumen_op.get("campos_no_verificables") is not None
                else conteo_op.get("campos_no_verificables", 0)
            )

            try:
                campos_no_verificables_total += int(
                    campos_no_verificables or 0
                )
            except (TypeError, ValueError):
                pass

            tenant_op = (
                op.get("cliente_id")
                or op.get("id_cliente")
                or "SIN_TENANT"
            )

            if tenant_op not in cobertura_por_tenant_acumulada:
                cobertura_por_tenant_acumulada[tenant_op] = {
                    "tenant": tenant_op,
                    "suma_cobertura": 0.0,
                    "operaciones_con_control": 0,
                    "operaciones_verdes": 0,
                    "operaciones_amarillas": 0,
                    "operaciones_rojas": 0,
                    "campos_no_verificables": 0,
                }

            tenant_kpi = cobertura_por_tenant_acumulada[tenant_op]

            if cobertura_num is not None:
                tenant_kpi["suma_cobertura"] += cobertura_num
                tenant_kpi["operaciones_con_control"] += 1

            if semaforo_op == "VERDE":
                tenant_kpi["operaciones_verdes"] += 1
            elif semaforo_op == "AMARILLO":
                tenant_kpi["operaciones_amarillas"] += 1
            elif semaforo_op == "ROJO":
                tenant_kpi["operaciones_rojas"] += 1

            try:
                tenant_kpi["campos_no_verificables"] += int(
                    campos_no_verificables or 0
                )
            except (TypeError, ValueError):
                pass

        cobertura_promedio_validacion = round(
            sum(coberturas_validacion) / len(coberturas_validacion),
            2
        ) if coberturas_validacion else 0

        cobertura_por_tenant = []

        for tenant_kpi in cobertura_por_tenant_acumulada.values():

            operaciones_con_control = tenant_kpi.get(
                "operaciones_con_control",
                0
            )

            cobertura_promedio_tenant = round(
                tenant_kpi.get("suma_cobertura", 0)
                / max(operaciones_con_control, 1),
                2
            ) if operaciones_con_control else 0

            cobertura_por_tenant.append({
                "tenant": tenant_kpi.get("tenant"),
                "cobertura_promedio_validacion": cobertura_promedio_tenant,
                "operaciones_con_control": operaciones_con_control,
                "operaciones_verdes": tenant_kpi.get(
                    "operaciones_verdes",
                    0
                ),
                "operaciones_amarillas": tenant_kpi.get(
                    "operaciones_amarillas",
                    0
                ),
                "operaciones_rojas": tenant_kpi.get(
                    "operaciones_rojas",
                    0
                ),
                "campos_no_verificables": tenant_kpi.get(
                    "campos_no_verificables",
                    0
                ),
            })

        cobertura_por_tenant = sorted(
            cobertura_por_tenant,
            key=lambda x: x.get(
                "cobertura_promedio_validacion",
                0
            ),
            reverse=True
        )

        # =====================================================
        # ANALYTICS REALES POR OPERADOR Y TENDENCIA
        # =====================================================

        operadores_map = {}
        tendencia_diaria_map = {}
        operaciones_con_operador = 0
        operaciones_sin_operador = 0

        for op in historial:

            if not isinstance(op, dict):
                continue

            usuario_email_op = str(
                op.get("usuario_email")
                or op.get("creado_por")
                or ""
            ).strip().lower()

            operador_nombre = str(
                op.get("operador")
                or usuario_email_op
                or ""
            ).strip()

            rol_op = str(
                op.get("rol_operador")
                or "operador"
            ).strip().lower()

            fecha_raw = (
                op.get("fecha")
                or op.get("timestamp_local")
                or op.get("created_at")
                or op.get("timestamp")
            )

            fecha_iso = None
            fecha_dt = None

            try:
                if fecha_raw:
                    fecha_dt = datetime.fromisoformat(
                        str(fecha_raw)
                        .replace("Z", "+00:00")
                    )
                    fecha_iso = fecha_dt.date().isoformat()
            except Exception:
                fecha_dt = None
                fecha_iso = None

            if fecha_iso:
                if fecha_iso not in tendencia_diaria_map:
                    tendencia_diaria_map[fecha_iso] = {
                        "fecha": fecha_iso,
                        "operaciones": 0,
                        "aprobadas": 0,
                        "verdes": 0,
                        "amarillas": 0,
                        "rojas": 0,
                    }

                tendencia_dia = tendencia_diaria_map[fecha_iso]
                tendencia_dia["operaciones"] += 1

                if op.get("aprobada") is True:
                    tendencia_dia["aprobadas"] += 1

                resumen_op = (
                    op.get("resumen_operativo")
                    or op.get("control")
                    or {}
                )

                if not isinstance(resumen_op, dict):
                    resumen_op = {}

                semaforo_tendencia = str(
                    resumen_op.get("semaforo")
                    or resumen_op.get("estado")
                    or ""
                ).upper()

                if semaforo_tendencia == "VERDE":
                    tendencia_dia["verdes"] += 1
                elif semaforo_tendencia == "AMARILLO":
                    tendencia_dia["amarillas"] += 1
                elif semaforo_tendencia == "ROJO":
                    tendencia_dia["rojas"] += 1

            if not usuario_email_op and not operador_nombre:
                operaciones_sin_operador += 1
                continue

            operaciones_con_operador += 1

            operador_clave = (
                usuario_email_op
                or operador_nombre.lower()
            )

            if operador_clave not in operadores_map:
                operadores_map[operador_clave] = {
                    "usuario_email": usuario_email_op or None,
                    "operador": operador_nombre or usuario_email_op,
                    "rol": rol_op,
                    "operaciones": 0,
                    "aprobadas": 0,
                    "verdes": 0,
                    "amarillas": 0,
                    "rojas": 0,
                    "campos_no_verificables": 0,
                    "coberturas": [],
                    "ultima_actividad": None,
                }

            operador_kpi = operadores_map[operador_clave]
            operador_kpi["operaciones"] += 1

            if op.get("aprobada") is True:
                operador_kpi["aprobadas"] += 1

            resumen_op = (
                op.get("resumen_operativo")
                or op.get("control")
                or {}
            )

            if not isinstance(resumen_op, dict):
                resumen_op = {}

            semaforo_operador = str(
                resumen_op.get("semaforo")
                or resumen_op.get("estado")
                or ""
            ).upper()

            if semaforo_operador == "VERDE":
                operador_kpi["verdes"] += 1
            elif semaforo_operador == "AMARILLO":
                operador_kpi["amarillas"] += 1
            elif semaforo_operador == "ROJO":
                operador_kpi["rojas"] += 1

            cobertura_operador = resumen_op.get(
                "cobertura_validacion_pct"
            )

            try:
                operador_kpi["coberturas"].append(
                    float(cobertura_operador)
                )
            except (TypeError, ValueError):
                pass

            conteo_operador = resumen_op.get("conteo") or {}

            no_verificables_operador = (
                resumen_op.get("campos_no_verificables")
                if resumen_op.get("campos_no_verificables") is not None
                else conteo_operador.get(
                    "campos_no_verificables",
                    0
                )
            )

            try:
                operador_kpi["campos_no_verificables"] += int(
                    no_verificables_operador or 0
                )
            except (TypeError, ValueError):
                pass

            if fecha_raw:
                fecha_texto = str(fecha_raw)

                if (
                    not operador_kpi.get("ultima_actividad")
                    or fecha_texto > operador_kpi["ultima_actividad"]
                ):
                    operador_kpi["ultima_actividad"] = fecha_texto

        productividad_por_operador = []

        for operador_kpi in operadores_map.values():

            coberturas_operador = operador_kpi.pop(
                "coberturas",
                []
            )

            cobertura_promedio_operador = round(
                sum(coberturas_operador)
                / len(coberturas_operador),
                2
            ) if coberturas_operador else 0

            total_operador = operador_kpi.get("operaciones", 0)
            aprobadas_operador = operador_kpi.get("aprobadas", 0)

            tasa_aprobacion = round(
                aprobadas_operador
                / max(total_operador, 1)
                * 100,
                2
            ) if total_operador else 0

            productividad_por_operador.append({
                **operador_kpi,
                "cobertura_promedio_validacion": (
                    cobertura_promedio_operador
                ),
                "tasa_aprobacion_pct": tasa_aprobacion,
            })

        productividad_por_operador = sorted(
            productividad_por_operador,
            key=lambda x: (
                x.get("operaciones", 0),
                x.get("cobertura_promedio_validacion", 0)
            ),
            reverse=True
        )

        top_operadores = productividad_por_operador[:10]

        tendencia_diaria = sorted(
            tendencia_diaria_map.values(),
            key=lambda x: x.get("fecha") or ""
        )[-30:]

        revenue_historico = [
            {
                "mes": "ACTUAL",
                "revenue": revenue_estimado
            }
        ]

        operaciones_por_dia = [
            {"dia": dia, "ops": total}
            for dia, total in operaciones_por_dia_map.items()
        ]

        riesgos = [
            {"riesgo": riesgo, "total": total}
            for riesgo, total in riesgos_map.items()
        ]

        aprobaciones = {
            "aprobadas": aprobaciones_total,
            "pendientes": max(total_operaciones - aprobaciones_total, 0),
            "total": total_operaciones,
        }

        analytics = {
            "revenue_historico": revenue_historico,
            "operaciones_por_dia": operaciones_por_dia,
            "riesgos": riesgos,
            "aprobaciones": aprobaciones,
            "top_tenants": top_tenants,
            "activity_feed_total": len(activity_feed),
            "productividad_por_operador": productividad_por_operador,
            "top_operadores": top_operadores,
            "tendencia_diaria": tendencia_diaria,
            "trazabilidad": {
                "operaciones_con_operador": operaciones_con_operador,
                "operaciones_sin_operador": operaciones_sin_operador,
                "cobertura_trazabilidad_pct": round(
                    operaciones_con_operador
                    / max(total_operaciones, 1)
                    * 100,
                    2
                ) if total_operaciones else 0,
            },
            "validacion_operacional": {
                "cobertura_promedio_validacion": cobertura_promedio_validacion,
                "operaciones_verdes": operaciones_verdes,
                "operaciones_amarillas": operaciones_amarillas,
                "operaciones_rojas": operaciones_rojas,
                "operaciones_sin_control": operaciones_sin_control,
                "campos_no_verificables_total": campos_no_verificables_total,
                "cobertura_por_tenant": cobertura_por_tenant,
            },
        }

        activity_feed_operativo = []
        for evento in activity_feed:
            nuevo = dict(evento)
            nuevo["tipo"] = "OPERATIVO"
            nuevo["accion"] = "OPERACION_APROBADA" if evento.get("aprobada") else "OPERACION_REGISTRADA"
            nuevo["actor_email"] = None
            nuevo["objetivo_email"] = None
            activity_feed_operativo.append(nuevo)

        activity_feed_admin = []
        activity_feed_admin_fuente = "supabase"

        try:
            import json
            import os

            eventos_admin = []

            if supabase_config_ok():
                url_feed_admin = (
                    f"{SUPABASE_URL}/rest/v1/argo_auditoria_admin"
                    f"?select=fecha,accion,actor_email,actor_rol,"
                    f"tenant,objetivo_email,detalle"
                    f"&order=fecha.desc"
                    f"&limit=100"
                )

                respuesta_feed_admin = requests.get(
                    url_feed_admin,
                    headers=_headers(),
                    timeout=30,
                )

                if respuesta_feed_admin.status_code == 200:
                    eventos_admin = respuesta_feed_admin.json() or []
                else:
                    print(
                        "WARNING activity_feed_admin supabase:",
                        respuesta_feed_admin.status_code,
                        respuesta_feed_admin.text,
                    )

            if not eventos_admin:
                activity_feed_admin_fuente = "jsonl_fallback"
                path_logs = "logs/admin_audit.jsonl"

                if os.path.exists(path_logs):
                    with open(path_logs, "r", encoding="utf-8") as f:
                        for line in f:
                            try:
                                evento = json.loads(line.strip())
                            except Exception:
                                continue

                            eventos_admin.append(evento)

            for evento in eventos_admin:
                if not isinstance(evento, dict):
                    continue

                tenant_evento = evento.get("tenant")

                if tenant_evento and tenant_evento not in tenants:
                    continue

                detalle = evento.get("detalle") or {}

                if not isinstance(detalle, dict):
                    detalle = {}

                accion_evento = str(
                    evento.get("accion")
                    or "EVENTO_ADMIN"
                )

                activity_feed_admin.append({
                    "tipo": "ADMIN",
                    "accion": accion_evento,
                    "tenant": tenant_evento or "SIN_TENANT",
                    "operacion": (
                        detalle.get("id_operacion")
                        or evento.get("objetivo_email")
                        or accion_evento
                    ),
                    "riesgo": detalle.get("severidad") or "ADMIN",
                    "aprobada": True,
                    "fecha": evento.get("fecha") or "N/D",
                    "actor_email": evento.get("actor_email"),
                    "actor_rol": evento.get("actor_rol"),
                    "objetivo_email": evento.get("objetivo_email"),
                    "detalle": detalle,
                    "fuente": activity_feed_admin_fuente,
                })

        except Exception as feed_admin_error:
            print("WARNING activity_feed_admin:", str(feed_admin_error))

        activity_feed_unificado = sorted(
            activity_feed_operativo + activity_feed_admin,
            key=lambda x: str(x.get("fecha") or ""),
            reverse=True
        )

        resumen_ejecutivo = {
            "riesgo_global": riesgo_global,
            "licencias_por_vencer_total": len(licencias_por_vencer),
            "tenants_en_riesgo_total": len(tenants_en_riesgo),
            "activity_feed_total": len(activity_feed_unificado),
            "activity_feed_operativo_total": len(activity_feed_operativo),
            "activity_feed_admin_total": len(activity_feed_admin),
            "activity_feed_admin_fuente": activity_feed_admin_fuente,
            "revenue_mensual_estimado_usd": revenue_estimado,
            "ticket_promedio_tenant_usd": round(
                revenue_estimado / max(len(tenants), 1),
                2
            ),
            "aprobaciones_total": aprobaciones_total,
            "operaciones_por_dia_total": sum(operaciones_por_dia_map.values()),
            "cobertura_promedio_validacion": cobertura_promedio_validacion,
            "operaciones_verdes": operaciones_verdes,
            "operaciones_amarillas": operaciones_amarillas,
            "operaciones_rojas": operaciones_rojas,
            "operaciones_sin_control": operaciones_sin_control,
            "campos_no_verificables_total": campos_no_verificables_total,
            "operadores_activos_total": len(
                productividad_por_operador
            ),
            "operaciones_con_operador": operaciones_con_operador,
            "operaciones_sin_operador": operaciones_sin_operador,
            "cobertura_trazabilidad_pct": round(
                operaciones_con_operador
                / max(total_operaciones, 1)
                * 100,
                2
            ) if total_operaciones else 0,
        }

        return {
            "ok": True,

            "saas": {
                "tenants_totales": len(tenants),
                "tenants_activos": activos,
                "tenants_suspendidos": suspendidos,
                "usuarios_totales": total_usuarios,
                "operaciones_totales": total_operaciones,
                "revenue_estimado_usd": revenue_estimado,
                "planes": planes,
                "operaciones_criticas": operaciones_criticas,
                "aprobaciones_total": aprobaciones_total,
                "riesgo_global": riesgo_global,
                "licencias_por_vencer_total": len(licencias_por_vencer),
                "tenants_en_riesgo_total": len(tenants_en_riesgo),
            },

            "resumen_ejecutivo": resumen_ejecutivo,

            "analytics": analytics,

            "top_tenants": top_tenants,

            "licencias_por_vencer": licencias_por_vencer,

            "tenants_en_riesgo": tenants_en_riesgo,

            "activity_feed": activity_feed_unificado[:25],

            "activity_feed_operativo": activity_feed_operativo[-15:],

            "activity_feed_admin": activity_feed_admin[:25],

            "ultimas_aprobaciones": ultimas_aprobaciones,

            "upgrade_sugeridos": [
                t for t in tenants.values()
                if t.get("operaciones_mes", 0) > 200
            ]
        }

    except Exception as e:

        return JSONResponse(
            status_code=500,
            content={
                "ok": False,
                "error": str(e)
            }
        )



@app.get("/argo/dashboard")
async def endpoint_dashboard(
    cliente_id: str = Query(default=None),
    x_cliente_id: str = Header(default=None),
    x_usuario_email: str = Header(default=None),
):
    try:

        # =========================================
        # HARD MULTI-TENANT ISOLATION ENTERPRISE
        # =========================================

        if not x_usuario_email:
            return {
                "ok": False,
                "error": "usuario requerido para consultar dashboard",
                "codigo": "TENANT_AUTH_REQUIRED"
            }

        usuario_rbac = obtener_usuario_rbac(x_usuario_email)

        if not usuario_rbac:
            return {
                "ok": False,
                "error": "Usuario no encontrado",
                "codigo": "TENANT_USER_NOT_FOUND"
            }

        if usuario_rbac.get("activo") is False:
            return {
                "ok": False,
                "error": "Usuario inactivo",
                "codigo": "TENANT_USER_INACTIVE"
            }

        rol = str(usuario_rbac.get("rol") or "operador").lower()
        cliente_usuario = usuario_rbac.get("id_cliente")

        cliente_solicitado = x_cliente_id or cliente_id or cliente_usuario

        if not cliente_solicitado:
            return {
                "ok": False,
                "error": "cliente_id requerido",
                "codigo": "TENANT_CLIENT_REQUIRED"
            }

        if rol != "master_admin" and cliente_solicitado != cliente_usuario:
            return {
                "ok": False,
                "error": "Acceso denegado al cliente solicitado",
                "codigo": "TENANT_DENY",
                "cliente_usuario": cliente_usuario,
                "cliente_solicitado": cliente_solicitado,
            }

        dashboard = obtener_dashboard_supabase(cliente_solicitado)

        return {
            "ok": True,
            "cliente_id": cliente_solicitado,
            "tenant": {
                "email": x_usuario_email,
                "rol": rol,
                "cliente_usuario": cliente_usuario,
                "aislamiento": "HARD"
            },
            "dashboard": dashboard
        }

    except Exception as e:
        return {
            "ok": False,
            "error": str(e)
        }
@app.post("/argo/ocr")
async def argo_ocr(
    request: Request,
    archivos: list[UploadFile] | None = File(default=None),
    archivo1: UploadFile = File(None),
    archivo2: UploadFile = File(None),
    archivo3: UploadFile = File(None),
    archivo4: UploadFile = File(None),
    archivo5: UploadFile = File(None),
):
    import json
    import re

    x_usuario_email = request.headers.get("x-usuario-email")

    usuario_actual = obtener_usuario_rbac(x_usuario_email)

    licencia = validar_licencia_saas(usuario_actual)

    if not licencia.get("ok"):

        return JSONResponse(
            status_code=403,
            content=licencia
        )

    validacion_modulo = validar_modulo_usuario(
        usuario_actual,
        "entrada_documental"
    )

    if not validacion_modulo.get("ok"):
        return JSONResponse(
            status_code=403,
            content=validacion_modulo
        )


    # =====================================================
    # CARGA MASIVA DE EVIDENCIAS v1.0
    # Compatibilidad:
    # - Campo dinámico repetible: archivos
    # - Campos legacy: archivo1 ... archivo5
    # =====================================================
    MAX_ARCHIVOS_POR_OPERACION = 100
    MAX_BYTES_POR_ARCHIVO = 20 * 1024 * 1024
    MAX_BYTES_TOTALES = 500 * 1024 * 1024

    archivos_dinamicos = list(archivos or [])
    archivos_legacy = [
        archivo1,
        archivo2,
        archivo3,
        archivo4,
        archivo5,
    ]

    archivos_validos = [
        archivo
        for archivo in archivos_dinamicos + archivos_legacy
        if archivo is not None
    ]

    if not archivos_validos:
        return JSONResponse(
            status_code=400,
            content={
                "ok": False,
                "error": "No se recibieron archivos",
                "codigo": "ARCHIVOS_REQUERIDOS",
            },
        )

    if len(archivos_validos) > MAX_ARCHIVOS_POR_OPERACION:
        return JSONResponse(
            status_code=413,
            content={
                "ok": False,
                "error": (
                    f"Máximo {MAX_ARCHIVOS_POR_OPERACION} archivos "
                    "por operación"
                ),
                "codigo": "LIMITE_ARCHIVOS_EXCEDIDO",
                "recibidos": len(archivos_validos),
                "maximo": MAX_ARCHIVOS_POR_OPERACION,
            },
        )

    resultados = []
    errores = []
    bytes_recibidos = 0

    # =========================
    # OCR POR ARCHIVO
    # =========================
    for file in archivos_validos:
        try:
            contenido = await file.read()
            tamano_archivo = len(contenido)
            bytes_recibidos += tamano_archivo

            if tamano_archivo == 0:
                errores.append({
                    "archivo": getattr(
                        file,
                        "filename",
                        "archivo_sin_nombre",
                    ),
                    "error": "Archivo vacío",
                    "codigo": "ARCHIVO_VACIO",
                })
                continue

            if tamano_archivo > MAX_BYTES_POR_ARCHIVO:
                errores.append({
                    "archivo": getattr(
                        file,
                        "filename",
                        "archivo_sin_nombre",
                    ),
                    "error": "El archivo supera 20 MB",
                    "codigo": "ARCHIVO_DEMASIADO_GRANDE",
                    "bytes": tamano_archivo,
                })
                continue

            if bytes_recibidos > MAX_BYTES_TOTALES:
                return JSONResponse(
                    status_code=413,
                    content={
                        "ok": False,
                        "error": (
                            "La carga completa supera el límite "
                            "operativo de 500 MB"
                        ),
                        "codigo": "CARGA_TOTAL_DEMASIADO_GRANDE",
                        "bytes_recibidos": bytes_recibidos,
                        "maximo_bytes": MAX_BYTES_TOTALES,
                    },
                )

            imagen_base64 = convertir_a_base64(contenido)

            response = client.responses.create(
                model="gpt-5.4",
                input=[
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "input_text",
                                "text": """
Eres un sistema OCR experto en logística y documentos de embarque.

Tu tarea es extraer SOLAMENTE un JSON válido, sin texto adicional, sin explicación, sin markdown.

Debes responder EXACTAMENTE con este esquema:

{
  "cliente": null,
  "proveedor": null,
  "paqueteria": null,
  "tracking": null,
  "descripcion": null,
  "cantidad_bultos": null,
  "peso_total": null,
  "peso_unidad": null,
  "direccion_origen": null,
  "direccion_destino": null
}

Reglas obligatorias:
- No inventes datos.
- Si no se ve claramente, usa null.
- cliente = consignee / ship to / deliver to / buyer / recipient si aplica.
- proveedor = shipper / vendor / supplier / remitente si aplica.
- paqueteria = UPS / FedEx / DHL / etc.
- tracking = número principal de guía.
- descripcion = descripción del producto o mercancía.
- cantidad_bultos:
  - si ves "2 OF 3" devuelve 3
  - si ves "1 OF 1" devuelve 1
  - si ves "PKGS 2" devuelve 2
- peso_total:
  - si ves "40 LBS" devuelve 40
  - si ves "12 KG" devuelve 12
- peso_unidad:
  - si ves LB o LBS devuelve "LBS"
  - si ves KG o KGS devuelve "KGS"
- Responde solo JSON válido.
"""
                            },
                            {
                                "type": "input_image",
                                "image_url": f"data:image/jpeg;base64,{imagen_base64}"
                            }
                        ]
                    }
                ]
            )

            texto = (response.output_text or "").strip()

            try:
                ocr_json = json.loads(texto)
            except Exception:
                match = re.search(r"\{.*\}", texto, re.DOTALL)
                if match:
                    try:
                        ocr_json = json.loads(match.group(0))
                    except Exception:
                        ocr_json = {
                            "cliente": None,
                            "proveedor": None,
                            "paqueteria": None,
                            "tracking": None,
                            "descripcion": None,
                            "cantidad_bultos": None,
                            "peso_total": None,
                            "peso_unidad": None,
                            "direccion_origen": None,
                            "direccion_destino": None
                        }
                else:
                    ocr_json = {
                        "cliente": None,
                        "proveedor": None,
                        "paqueteria": None,
                        "tracking": None,
                        "descripcion": None,
                        "cantidad_bultos": None,
                        "peso_total": None,
                        "peso_unidad": None,
                        "direccion_origen": None,
                        "direccion_destino": None
                    }

            resultados.append({
                "archivo": getattr(file, "filename", "archivo.jpg"),
                "ocr_raw": texto,
                "ocr_json": ocr_json
            })

        except Exception as e:
            errores.append({
                "archivo": getattr(file, "filename", "archivo.jpg") if file else None,
                "error": str(e)
            })

    # =========================
    # CONSOLIDADO INTELIGENTE
    # =========================
    consolidado = {
        "cliente": None,
        "proveedor": None,
        "paqueteria": None,
        "tracking": None,
        "descripcion": None,
        "cantidad_bultos": None,
        "peso_total": None,
        "peso_unidad": None,
        "direccion_origen": None,
        "direccion_destino": None
    }

    def extraer_peso_desde_texto(texto):
        if texto in [None, "", "null"]:
            return None, None

        t = str(texto).upper().strip()

        unidad = None
        if "LBS" in t or "LB" in t:
            unidad = "LBS"
        elif "KGS" in t or "KG" in t:
            unidad = "KGS"

        match = re.search(r"(\d+(?:\.\d+)?)", t)
        numero = None
        if match:
            try:
                valor = float(match.group(1))
                numero = int(valor) if valor.is_integer() else valor
            except Exception:
                numero = None

        return numero, unidad

    for item in resultados:
        data = item.get("ocr_json", {})
        nombre_archivo = (item.get("archivo") or "").lower()

        prioridad_cliente_proveedor = 1
        prioridad_tracking_paqueteria = 1

        if "invoice" in nombre_archivo or "packing" in nombre_archivo:
            prioridad_cliente_proveedor = 3

        if "paqueteria" in nombre_archivo or "label" in nombre_archivo or "etiqueta" in nombre_archivo:
            prioridad_tracking_paqueteria = 3

        # Peso: revisar ambos campos y conservar ambos valores
        peso_num_1, peso_uni_1 = extraer_peso_desde_texto(data.get("peso_total"))
        peso_num_2, peso_uni_2 = extraer_peso_desde_texto(data.get("peso_unidad"))

        peso_num = peso_num_1 if peso_num_1 is not None else peso_num_2
        peso_uni = peso_uni_1 if peso_uni_1 is not None else peso_uni_2

        if peso_num is not None and consolidado["peso_total"] in [None, "", "null"]:
            consolidado["peso_total"] = peso_num

        if peso_uni is not None and consolidado["peso_unidad"] in [None, "", "null"]:
            consolidado["peso_unidad"] = peso_uni

        for campo in consolidado.keys():
            valor_actual = consolidado[campo]
            valor_nuevo = data.get(campo)

            if campo in ["peso_total", "peso_unidad"]:
                continue

            if valor_nuevo in [None, "", "null"]:
                continue

            if campo in ["cliente", "proveedor"]:
                if prioridad_cliente_proveedor >= 3:
                    consolidado[campo] = valor_nuevo
                elif consolidado[campo] in [None, "", "null"]:
                    consolidado[campo] = valor_nuevo

            elif campo in ["tracking", "paqueteria"]:
                if prioridad_tracking_paqueteria >= 3:
                    consolidado[campo] = valor_nuevo
                elif consolidado[campo] in [None, "", "null"]:
                    consolidado[campo] = valor_nuevo

            elif campo == "cantidad_bultos":
                if isinstance(valor_nuevo, str):
                    texto_cb = valor_nuevo.upper().strip()
                    if "OF" in texto_cb:
                        try:
                            total = int(texto_cb.split("OF")[1].strip())
                            consolidado[campo] = total
                        except Exception:
                            if consolidado[campo] is None:
                                consolidado[campo] = valor_nuevo
                    else:
                        m = re.search(r"(\d+)", texto_cb)
                        if m and consolidado[campo] is None:
                            consolidado[campo] = int(m.group(1))
                elif consolidado[campo] is None:
                    consolidado[campo] = valor_nuevo

            elif campo == "descripcion":
                if not valor_actual or len(str(valor_nuevo)) > len(str(valor_actual)):
                    consolidado[campo] = valor_nuevo

            elif consolidado[campo] in [None, "", "null"]:
                consolidado[campo] = valor_nuevo

    # =========================
    # FALTANTES INTELIGENTES
    # =========================
    faltantes = []
    alertas = []

    def es_faltante(valor):
        if valor is None:
            return True
        if isinstance(valor, str):
            v = valor.strip().lower()
            if v in ["", "null", "no legible", "n/a", "na", "unknown"]:
                return True
        return False

    if es_faltante(consolidado.get("cliente")):
        faltantes.append({"campo": "cliente", "valor": "No detectado"})

    if es_faltante(consolidado.get("proveedor")):
        faltantes.append({"campo": "proveedor", "valor": "No detectado"})

    if es_faltante(consolidado.get("paqueteria")):
        faltantes.append({"campo": "paqueteria", "valor": "No detectado"})

    tracking = consolidado.get("tracking")
    if es_faltante(tracking) or (isinstance(tracking, str) and len(tracking.strip()) < 8):
        faltantes.append({"campo": "tracking", "valor": "No detectado"})

    descripcion = consolidado.get("descripcion")
    if es_faltante(descripcion) or (isinstance(descripcion, str) and len(descripcion.strip()) < 10):
        faltantes.append({"campo": "descripcion", "valor": "No detectado"})

    if consolidado.get("cantidad_bultos") in [None, "", "null", 0]:
        faltantes.append({"campo": "cantidad_bultos", "valor": "No detectado"})

    if consolidado.get("peso_total") in [None, "", "null", 0]:
        faltantes.append({"campo": "peso_total", "valor": "No detectado"})

    if es_faltante(consolidado.get("peso_unidad")):
        faltantes.append({"campo": "peso_unidad", "valor": "No detectado"})

    if es_faltante(consolidado.get("direccion_origen")):
        faltantes.append({"campo": "direccion_origen", "valor": "No detectado"})

    if es_faltante(consolidado.get("direccion_destino")):
        faltantes.append({"campo": "direccion_destino", "valor": "No detectado"})

    # =========================
    # PRIORIDAD
    # =========================
    faltantes_priorizados = []

    for f in faltantes:
        campo = f["campo"]

        if campo in ["cliente", "tracking"]:
            nivel = "CRITICO"
        elif campo in ["descripcion", "cantidad_bultos"]:
            nivel = "MEDIO"
        else:
            nivel = "BAJO"

        faltantes_priorizados.append({
            "campo": campo,
            "nivel": nivel
        })

    # =========================
    # ESTADO
    # =========================
    estado = "OK"
    severidad_maxima = "NINGUNA"

    if any(fp["nivel"] == "CRITICO" for fp in faltantes_priorizados):
        estado = "REVISION"
        severidad_maxima = "ALTA"
    elif len(faltantes) > 0:
        estado = "ADVERTENCIA"
        severidad_maxima = "MEDIA"

    # =========================
    # DECISION
    # =========================
    accion = "CONTINUAR"
    razon = "Sin faltantes"

    if any(fp["nivel"] == "CRITICO" for fp in faltantes_priorizados):
        accion = "CONTINUAR_CON_ALERTA"
        razon = "Faltantes críticos: operación continúa con baja certeza y requiere revisión"
    elif any(fp["nivel"] == "MEDIO" for fp in faltantes_priorizados):
        accion = "CONTINUAR_CON_ALERTA"
        razon = "Faltantes medios"
    elif any(fp["nivel"] == "BAJO" for fp in faltantes_priorizados):
        accion = "CONTINUAR"
        razon = "Solo faltantes menores"

    return {
        "ok": True,
        "modulo": "ARGO_OCR",
        "estado": estado,
        "severidad_maxima": severidad_maxima,
        "decision": {
            "accion": accion,
            "razon": razon
        },
        "conteo": {
            "faltantes": len(faltantes),
            "alertas": len(alertas)
        },
        "faltantes": faltantes,
        "faltantes_priorizados": faltantes_priorizados,
        "alertas": alertas,
        "total_archivos": len(archivos_validos),
        "procesados": len(resultados),
        "fallidos": len(errores),
        "bytes_recibidos": bytes_recibidos,
        "limite_archivos": MAX_ARCHIVOS_POR_OPERACION,
        "errores": errores,
        "consolidado": consolidado,
        "resultados": resultados
    }

@app.post("/argo/generar_desde_ocr")
async def argo_generar_desde_ocr(payload: dict = Body(...)):
    from datetime import datetime

    # =========================
    # INPUT
    # =========================
    ocr = payload.get("ocr") or payload
    decision = payload.get("decision", {})

    accion = decision.get("accion", "CONTINUAR")

    # =========================
    # ENTERPRISE: OCR NUNCA DETIENE
    # =========================
    if accion == "DETENER":
        accion = "CONTINUAR_CON_ALERTA"
        decision["accion"] = accion
        decision["razon"] = "Política enterprise: OCR no detiene; continúa con revisión obligatoria"
    # =========================
    # DATOS OCR
    # =========================
    cliente = ocr.get("cliente") or "No legible"
    proveedor = ocr.get("proveedor") or "No legible"
    paqueteria = ocr.get("paqueteria") or "No legible"
    tracking = ocr.get("tracking") or "No legible"
    descripcion = ocr.get("descripcion") or "No legible"
    cantidad_bultos = ocr.get("cantidad_bultos")
    peso_total = ocr.get("peso_total")
    peso_unidad = ocr.get("peso_unidad") or "No legible"
    direccion_origen = ocr.get("direccion_origen") or "No legible"
    direccion_destino = ocr.get("direccion_destino") or "No legible"

    # =========================
    # NORMALIZACIÓN
    # =========================
    peso_unidad_txt = str(peso_unidad).strip().upper()

    if "LBS" in peso_unidad_txt or peso_unidad_txt == "LB":
        peso_unidad_norm = "LBS"
    elif "KGS" in peso_unidad_txt or peso_unidad_txt == "KG":
        peso_unidad_norm = "KGS"
    else:
        peso_unidad_norm = "No legible"

    shipment_id = tracking if tracking != "No legible" else f"OCR-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    fecha_recepcion = datetime.now().strftime("%m/%d/%Y")

    cantidad = str(cantidad_bultos) if cantidad_bultos not in [None, "", "null"] else "No legible"
    peso_total_str = str(peso_total) if peso_total not in [None, "", "null"] else "No legible"

    # =========================
    # ENTRADA
    # =========================
    entrada = {
        "shipment_id": shipment_id,
        "fecha_recepcion": fecha_recepcion,
        "cliente": cliente,
        "proveedor": proveedor,
        "paqueteria": paqueteria,
        "tracking": tracking,
        "descripcion": descripcion,
        "marca": "No legible",
        "modelo": "No legible",
        "no_parte": "No legible",
        "no_lote": "No legible",
        "no_serie": "No legible",
        "cantidad": cantidad,
        "unidad": peso_unidad_norm,
        "peso_total": peso_total_str,
        "pais_origen": "No legible",
        "direccion_origen": direccion_origen,
        "direccion_destino": direccion_destino
    }

    # =========================
    # FALTANTES REALES PARA ESTA ETAPA
    # =========================
    campos_requeridos_generacion = [
        "cliente",
        "proveedor",
        "paqueteria",
        "tracking",
        "descripcion",
        "cantidad",
        "unidad",
        "peso_total",
        "direccion_origen",
        "direccion_destino"
    ]

    faltantes = []
    for campo in campos_requeridos_generacion:
        valor = entrada.get(campo)
        if valor == "No legible":
            faltantes.append({
                "campo": campo,
                "valor": valor
            })

    # =========================
    # ESTADO
    # =========================
    if accion == "CONTINUAR":
        estado = "OK"
        severidad_maxima = "NINGUNA"
    elif accion == "CONTINUAR_CON_ALERTA":
        estado = "ADVERTENCIA"
        severidad_maxima = "MEDIA"
    else:
        estado = "ADVERTENCIA"
        severidad_maxima = "MEDIA"

    # =========================
    # CONTROL
    # =========================
    control_stub = argo_control_validar({
        "version": "1.0",
        "modulo": "ARGO_ENTRADA",
        "entrada": entrada
    })

    # =========================
    # EXCEL
    # =========================
    wb = load_workbook(TEMPLATE_FILE)
    ws = wb["Entrada"]

    ws["B2"] = shipment_id
    ws["B3"] = fecha_recepcion
    ws["B4"] = cliente
    ws["B5"] = proveedor
    ws["B6"] = paqueteria

    ws["B7"].number_format = "@"
    ws["B7"] = tracking

    ws["B8"] = descripcion
    ws["B9"] = entrada["marca"]
    ws["B10"] = entrada["modelo"]
    ws["B11"] = entrada["no_parte"]
    ws["B12"] = entrada["no_lote"]
    ws["B13"] = entrada["no_serie"]
    ws["B14"] = cantidad
    ws["B15"] = peso_unidad_norm
    ws["B16"] = peso_total_str
    ws["B17"] = entrada["pais_origen"]

    # =========================
    # DATOS FALTANTES
    # =========================
    ws_df = wb["Datos faltantes"]
    _clear_sheet(ws_df)
    ws_df["A1"] = "Campo"
    ws_df["B1"] = "Valor"

    fila = 2
    for item in faltantes:
        ws_df[f"A{fila}"] = item["campo"]
        ws_df[f"B{fila}"] = item["valor"]
        fila += 1

    # =========================
    # ALERTAS
    # =========================
    alertas = []

    if accion == "CONTINUAR_CON_ALERTA":
        alertas.append({
            "alerta": "Generación con advertencia",
            "detalle": decision.get("razon", "OCR indicó continuar con alerta"),
            "severidad": "MEDIA"
        })

    ws_al = wb["Alertas"]
    _clear_sheet(ws_al)
    ws_al["A1"] = "Alerta"
    ws_al["B1"] = "Detalle"
    ws_al["C1"] = "Severidad"

    fila = 2
    for alerta in alertas:
        ws_al[f"A{fila}"] = alerta.get("alerta", "")
        ws_al[f"B{fila}"] = alerta.get("detalle", "")
        ws_al[f"C{fila}"] = alerta.get("severidad", "")
        fila += 1

    # =========================
    # RESUMEN OPERATIVO
    # =========================
    ws_res = wb["Resumen operativo"]
    _clear_sheet(ws_res)
    ws_res["A1"] = "Item"
    ws_res["B1"] = "Valor"

    resumen = [
        ("Fecha recepción", fecha_recepcion),
        ("Cliente", cliente),
        ("Proveedor", proveedor),
        ("Paquetería", paqueteria),
        ("Tracking", tracking),
        ("Shipment ID", shipment_id),
        ("País origen", entrada["pais_origen"]),
        ("Faltantes (#)", len(faltantes)),
        ("Alertas (#)", len(alertas)),
        ("Estado", estado),
        ("Severidad máxima", severidad_maxima),
    ]

    fila = 2
    for item, valor in resumen:
        ws_res[f"A{fila}"] = item
        ws_res[f"B{fila}"] = valor
        fila += 1

    # =========================
    # GUARDADO
    # =========================
    cliente_archivo = _safe_filename(cliente)
    ult4 = (tracking[-4:] if len(tracking) >= 4 else tracking) or "XXXX"

    output_name = f"ENTRADA_OCR_{cliente_archivo}_{ult4}.xlsx"
    output_path = os.path.join(OUTPUT_FOLDER, output_name)

    wb.save(output_path)

    # =========================
    # RESPUESTA
    # =========================
    return {
        "ok": True,
        "modulo": "ARGO_GENERAR_DESDE_OCR",
        "estado": estado,
        "severidad_maxima": severidad_maxima,
        "decision": decision,
        "conteo": {
            "faltantes": len(faltantes),
            "alertas": len(alertas)
        },
        "faltantes": faltantes,
        "alertas": alertas,
        "entrada": entrada,
        "control": control_stub,
        "archivo_generado": output_name,
        "ruta_archivo": output_path,
        "descarga": f"/descargar/{output_name}"
    }


@app.post("/argo/procesar_desde_ocr")
async def procesar_desde_ocr(
    request: Request,
    payload: dict,
    x_cliente_id: str = Header(default=None)
):
    try:

        x_usuario_email = request.headers.get("x-usuario-email")

        usuario_actual = obtener_usuario_rbac(x_usuario_email)

        if usuario_actual:

            validacion_plan = validar_limite_operaciones_plan(usuario_actual)

            if not validacion_plan.get("ok"):

                return JSONResponse(
                    status_code=403,
                    content=validacion_plan
                )

        ocr = payload or {}
        consolidado = ocr.get("consolidado", {}) or {}

        if not consolidado.get("cliente"):
            consolidado["cliente"] = "Fives Cinetic Mexico S A De C V"

        ocr["consolidado"] = consolidado

        tracking = consolidado.get("tracking") or generar_id_operacion()
        id_operacion = generar_id_operacion()

        cliente_id = (
            usuario_actual.get("id_cliente")
            if usuario_actual
            else (x_cliente_id or payload.get("cliente_id"))
        )

        if not cliente_id:
            return {
                "ok": False,
                "error": "cliente_id requerido"
            }

        cliente_nombre_login = (
            payload.get("cliente_nombre")
            or consolidado.get("cliente")
            or "SIN_CLIENTE"
         )
        salida_class = {}

        try:
            payload_master = {
                "meta": {
                    "id_operacion": id_operacion,
                    "id_shipment": tracking,
                    "id_item": None,
                },
                "descripcion": consolidado.get("descripcion") or "",
                "documentos": ocr.get("documentos", []),
                "control": {
                    "resumen": {}
                },
            }

            salida_class = build_output(payload_master) or {}

        except Exception as class_err:
            print(f"WARNING ARGO CLASS OCR [{id_operacion}]: {class_err}")
            salida_class = {}

        class_salida = salida_class.get("salida", {}) if isinstance(salida_class, dict) else {}
        class_score = class_salida.get("score_documental", {}) or {}
        class_clasificacion = class_salida.get("clasificacion", {}) or {}
        class_riesgo = class_salida.get("certeza_y_riesgo", {}) or {}

        entrada_control = {
            "cliente": consolidado.get("cliente"),
            "shipment_id": tracking,
            "tracking": tracking,
            "proveedor": consolidado.get("proveedor"),
            "paqueteria": consolidado.get("paqueteria"),
            "descripcion": consolidado.get("descripcion"),
            "peso_total": consolidado.get("peso_total"),
            "cantidad": consolidado.get("cantidad") or consolidado.get("cantidad_bultos"),
            "cantidad_bultos": consolidado.get("cantidad_bultos"),
            "marca": consolidado.get("marca"),
            "modelo": consolidado.get("modelo"),
            "no_parte": consolidado.get("no_parte"),
            "no_lote": consolidado.get("no_lote"),
            "no_serie": consolidado.get("no_serie"),
            "pais_origen": consolidado.get("pais_origen"),
        }

        control_generado = argo_control_validar({
            "version": "2.0",
            "modulo": "ARGO_PROCESAR_DESDE_OCR",
            "entrada": entrada_control
        })

        timestamp_operacion = datetime.now().isoformat()

        usuario_email_operacion = (
            (usuario_actual or {}).get("email")
            or x_usuario_email
            or payload.get("usuario_email")
            or payload.get("creado_por")
            or "sistema"
        )

        operador_operacion = (
            (usuario_actual or {}).get("nombre")
            or payload.get("operador")
            or usuario_email_operacion
        )

        rol_operador = str(
            (usuario_actual or {}).get("rol")
            or payload.get("rol_operador")
            or "operador"
        ).lower()

        operacion = {
            "cliente_id": cliente_id,
            "cliente_nombre": cliente_nombre_login,
            "id_operacion": id_operacion,
            "usuario_email": usuario_email_operacion,
            "operador": operador_operacion,
            "creado_por": usuario_email_operacion,
            "rol_operador": rol_operador,
            "fecha": timestamp_operacion,
            "timestamp_local": timestamp_operacion,
            "ocr": ocr,
            "class": salida_class,
            "control": control_generado,
            "semaforo_operacion": ocr.get("severidad_maxima") or "MEDIA",
            "decision": {
                "accion": "CONTINUAR_CON_ALERTA"
            },
            "generacion": {
                "entrada": entrada_control,
                "conteo": ocr.get("conteo", {}),
            },
        }

        control_info = (
            operacion.get("control")
            or ocr.get("control")
            or {}
        )

        resumen_operativo = {
            "semaforo": control_info.get("semaforo") or control_info.get("estado") or "SIN_CONTROL",
            "icono": control_info.get("icono") or "",
            "estado": control_info.get("estado") or "SIN_CONTROL",
            "cobertura_validacion_pct": control_info.get("cobertura_validacion_pct", 0),
            "dictamen_operativo": control_info.get("dictamen_operativo") or "Sin dictamen operativo.",
            "campos_totales": (control_info.get("conteo") or {}).get("campos_totales", 0),
            "campos_disponibles": (control_info.get("conteo") or {}).get("campos_disponibles", 0),
            "campos_no_verificables": (control_info.get("conteo") or {}).get("campos_no_verificables", 0),
            "alertas": control_info.get("alertas", []),
            "validaciones_operativas": control_info.get("validaciones_operativas", []),
        }

        operacion["resumen_operativo"] = resumen_operativo

        data_reporte = {
            "cliente": consolidado.get("cliente"),
            "shipment_id": tracking,
            "tracking": tracking,
            "proveedor": consolidado.get("proveedor"),
            "paqueteria": consolidado.get("paqueteria"),
            "descripcion": consolidado.get("descripcion"),
            "peso_total": consolidado.get("peso_total"),
            "cantidad_bultos": consolidado.get("cantidad_bultos"),
            "riesgo_automatico": class_riesgo.get("riesgo_automatico") or resumen_operativo.get("estado") or ocr.get("severidad_maxima") or "MEDIA",
            "score_documental": class_score.get("score_total_0_100") or ocr.get("score_documental_global") or 0,
            "fraccion_sugerida": class_clasificacion.get("fraccion_sugerida") or ocr.get("fraccion_sugerida") or "7318.15.99",
            "confianza_fraccion_pct": class_clasificacion.get("confianza_fraccion_pct") or ocr.get("confianza_fraccion_pct") or 0,
            "certeza_final_pct": class_riesgo.get("certeza_final_pct") or ocr.get("certeza_final_pct") or 0,
            "nivel_debida_diligencia": class_score.get("nivel_debida_diligencia") or ocr.get("nivel_debida_diligencia") or "BASICA",
            "semaforo_operativo": resumen_operativo.get("semaforo"),
            "icono_operativo": resumen_operativo.get("icono"),
            "cobertura_validacion_pct": resumen_operativo.get("cobertura_validacion_pct"),
            "dictamen_operativo": resumen_operativo.get("dictamen_operativo"),
            "campos_totales": resumen_operativo.get("campos_totales"),
            "campos_disponibles": resumen_operativo.get("campos_disponibles"),
            "campos_no_verificables": resumen_operativo.get("campos_no_verificables"),
            "validaciones_operativas": resumen_operativo.get("validaciones_operativas", []),
        }
        if usuario_actual:
            validacion_pdf = validar_feature_plan(usuario_actual, "export_pdf")

            if not validacion_pdf.get("ok"):
                return JSONResponse(
                    status_code=403,
                    content=validacion_pdf
                )

        ruta_reporte = generar_reporte_ejecutivo(
            "PLANTILLA_OFICIAL_ARGO_DOCUMENT_MEJORADA_v2026.xlsx",
            data_reporte,
            "outputs"
        )

        nombre_reporte = os.path.basename(ruta_reporte)

        reporte_storage = None

        try:
            reporte_storage = subir_archivo_a_supabase(ruta_reporte)
        except Exception as supa_err:
            print(f"WARNING SUPABASE REPORTE EJECUTIVO [{tracking}]: {supa_err}")

        operacion["reporte_ejecutivo"] = {
            "archivo": nombre_reporte,
            "ruta": ruta_reporte,
            "descarga": f"/descargar/{nombre_reporte}",
            "storage": reporte_storage,
        }
        guardado = guardar_operacion_supabase(operacion)

        try:
            from argo_historial import guardar_operacion_historial
            guardar_operacion_historial(operacion)
        except Exception as hist_err:
            print(f"WARNING HISTORIAL LOCAL [{tracking}]: {hist_err}")

        if isinstance(guardado, dict) and guardado.get("ok") is False:
            return guardado

        return {
            "ok": True,
            "mensaje": "Operación guardada desde OCR",
            "debug_guardado": guardado,
            "operacion": guardado,
            "resumen_operativo": resumen_operativo,
            "reporte_ejecutivo": {
                "archivo": nombre_reporte,
                "descarga": f"/descargar/{nombre_reporte}"
            }
        }

    except Exception as e:
        print("ERROR PROCESAR OCR:", str(e))
        return {
            "ok": False,
            "error": str(e),
        }


# =========================================================
# MASTER ADMIN EXECUTIVE EXPORT
# =========================================================

@app.get("/argo/master/export")
async def argo_master_export(request: Request):
    try:
        dashboard = await argo_master_dashboard(request)

        if not isinstance(dashboard, dict) or not dashboard.get("ok"):
            return JSONResponse(
                status_code=403,
                content=dashboard if isinstance(dashboard, dict) else {
                    "ok": False,
                    "error": "No autorizado o dashboard inválido"
                }
            )

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_name = f"ARGO_MASTER_EXECUTIVE_EXPORT_{timestamp}.xlsx"
        output_path = os.path.join("outputs", output_name)

        wb = Workbook()
        ws = wb.active
        ws.title = "Resumen Ejecutivo"

        dark = "0F172A"
        blue = "2563EB"
        green = "16A34A"
        amber = "F59E0B"
        red = "DC2626"
        light = "F8FAFC"
        white = "FFFFFF"
        border_color = "CBD5E1"

        title_font = Font(bold=True, size=18, color=white)
        subtitle_font = Font(bold=True, size=12, color=dark)
        header_font = Font(bold=True, color=white)
        normal_font = Font(size=11, color=dark)
        white_fill = PatternFill("solid", fgColor=white)
        dark_fill = PatternFill("solid", fgColor=dark)
        blue_fill = PatternFill("solid", fgColor=blue)
        light_fill = PatternFill("solid", fgColor=light)
        thin = Side(style="thin", color=border_color)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def style_title(sheet, title, subtitle):
            sheet.merge_cells("A1:H1")
            sheet["A1"] = title
            sheet["A1"].font = Font(bold=True, size=20, color=white)
            sheet["A1"].fill = dark_fill
            sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
            sheet.row_dimensions[1].height = 30

            sheet.merge_cells("A2:H2")
            sheet["A2"] = subtitle
            sheet["A2"].font = Font(size=11, color=white)
            sheet["A2"].fill = dark_fill
            sheet["A2"].alignment = Alignment(horizontal="center", vertical="center")
            sheet.row_dimensions[2].height = 22

            for col in range(1, 9):
                sheet.cell(row=1, column=col).border = border
                sheet.cell(row=2, column=col).border = border

        def write_table(sheet, start_row, headers, rows, freeze=True, filters=True):
            end_col = max(len(headers), 1)
            for idx, h in enumerate(headers, start=1):
                c = sheet.cell(row=start_row, column=idx, value=h)
                c.font = header_font
                c.fill = blue_fill
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = border

            for r_idx, row in enumerate(rows, start=start_row + 1):
                for c_idx, value in enumerate(row, start=1):
                    c = sheet.cell(row=r_idx, column=c_idx, value=value)
                    c.font = normal_font
                    c.fill = white_fill if r_idx % 2 else light_fill
                    c.border = border
                    c.alignment = Alignment(vertical="center", wrap_text=True)

            last_row = start_row + max(len(rows), 1)

            if filters:
                sheet.auto_filter.ref = f"A{start_row}:{sheet.cell(row=start_row, column=end_col).column_letter}{last_row}"

            if freeze:
                sheet.freeze_panes = f"A{start_row + 1}"

            for col in sheet.columns:
                max_len = 0
                col_letter = None

                for cell in col:
                    if not hasattr(cell, "column_letter"):
                        continue

                    if col_letter is None:
                        col_letter = cell.column_letter

                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))

                if col_letter:
                    sheet.column_dimensions[col_letter].width = min(max(max_len + 3, 14), 46)

        def metric_card(sheet, cell, title, value, fill_color):
            sheet[cell] = title
            sheet[cell].font = Font(bold=True, size=11, color=white)
            sheet[cell].fill = PatternFill("solid", fgColor=fill_color)
            sheet[cell].alignment = Alignment(horizontal="center")

            value_cell = sheet.cell(row=sheet[cell].row + 1, column=sheet[cell].column)
            value_cell.value = value
            value_cell.font = Font(bold=True, size=16, color=dark)
            value_cell.fill = light_fill
            value_cell.alignment = Alignment(horizontal="center")
            value_cell.border = border
            sheet[cell].border = border

        saas = dashboard.get("saas", {}) or {}
        resumen = dashboard.get("resumen_ejecutivo", {}) or {}
        analytics = dashboard.get("analytics", {}) or {}

        style_title(
            ws,
            "ARGO ENTERPRISE SaaS — EXECUTIVE REPORT",
            f"Reporte ejecutivo generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        )

        ws["A4"] = "Resumen CEO"
        ws["A4"].font = Font(bold=True, size=15, color=dark)

        metric_card(ws, "A6", "Tenants", saas.get("tenants_totales", 0), blue)
        metric_card(ws, "B6", "Usuarios", saas.get("usuarios_totales", 0), blue)
        metric_card(ws, "C6", "Operaciones", saas.get("operaciones_totales", 0), green)
        metric_card(ws, "D6", "Revenue USD", saas.get("revenue_estimado_usd", 0), green)
        metric_card(ws, "E6", "Riesgo Global", saas.get("riesgo_global", "BAJO"), amber if saas.get("riesgo_global") == "MEDIO" else (red if saas.get("riesgo_global") == "ALTO" else green))
        metric_card(ws, "F6", "Aprobaciones", saas.get("aprobaciones_total", 0), blue)

        ws["A10"] = "Detalle Ejecutivo"
        ws["A10"].font = Font(bold=True, size=13, color=dark)

        resumen_rows = [
            ["Tenants totales", saas.get("tenants_totales", 0)],
            ["Tenants activos", saas.get("tenants_activos", 0)],
            ["Tenants suspendidos", saas.get("tenants_suspendidos", 0)],
            ["Usuarios totales", saas.get("usuarios_totales", 0)],
            ["Operaciones totales", saas.get("operaciones_totales", 0)],
            ["Revenue estimado USD", saas.get("revenue_estimado_usd", 0)],
            ["Riesgo global", saas.get("riesgo_global", "BAJO")],
            ["Operaciones críticas", saas.get("operaciones_criticas", 0)],
            ["Aprobaciones total", saas.get("aprobaciones_total", 0)],
            ["Licencias por vencer", saas.get("licencias_por_vencer_total", 0)],
            ["Tenants en riesgo", saas.get("tenants_en_riesgo_total", 0)],
            ["Activity Feed eventos", resumen.get("activity_feed_total", 0)],
            ["Ticket promedio tenant USD", resumen.get("ticket_promedio_tenant_usd", 0)],
        ]

        write_table(ws, 12, ["Métrica", "Valor"], resumen_rows)

        # Tenants
        ws_tenants = wb.create_sheet("Tenants")
        style_title(ws_tenants, "TENANTS & LICENSES", "Uso, licencias, plan SaaS y operaciones por tenant")
        tenant_rows = [
            [
                t.get("tenant"),
                t.get("plan"),
                t.get("usuarios"),
                t.get("estado_licencia"),
                t.get("dias_restantes"),
                t.get("fecha_vencimiento"),
                t.get("operaciones_mes"),
            ]
            for t in dashboard.get("top_tenants", []) or []
        ]
        write_table(
            ws_tenants,
            4,
            ["Tenant", "Plan", "Usuarios", "Licencia", "Días restantes", "Vencimiento", "Operaciones mes"],
            tenant_rows
        )

        # Analytics
        ws_an = wb.create_sheet("Analytics")
        style_title(ws_an, "ANALYTICS EXECUTIVE", "Revenue, operaciones por día, riesgos y aprobaciones")

        revenue_rows = [
            [r.get("mes"), r.get("revenue")]
            for r in analytics.get("revenue_historico", []) or []
        ]
        write_table(ws_an, 4, ["Mes", "Revenue"], revenue_rows)

        ops_rows = [
            [o.get("dia"), o.get("ops")]
            for o in analytics.get("operaciones_por_dia", []) or []
        ]
        write_table(ws_an, 10, ["Día", "Operaciones"], ops_rows)

        risk_rows = [
            [r.get("riesgo"), r.get("total")]
            for r in analytics.get("riesgos", []) or []
        ]
        write_table(ws_an, 20, ["Riesgo", "Total"], risk_rows)

        aprob = analytics.get("aprobaciones", {}) or {}
        write_table(
            ws_an,
            28,
            ["Aprobadas", "Pendientes", "Total"],
            [[aprob.get("aprobadas", 0), aprob.get("pendientes", 0), aprob.get("total", 0)]]
        )

        # Activity Feed
        ws_feed = wb.create_sheet("Activity Feed")
        style_title(ws_feed, "EXECUTIVE ACTIVITY FEED", "Eventos recientes, riesgos y estado de aprobación")
        feed_rows = [
            [
                a.get("tenant"),
                a.get("operacion"),
                a.get("riesgo"),
                "SI" if a.get("aprobada") else "NO",
                a.get("fecha"),
            ]
            for a in dashboard.get("activity_feed", []) or []
        ]
        write_table(
            ws_feed,
            4,
            ["Tenant", "Operación", "Riesgo", "Aprobada", "Fecha"],
            feed_rows
        )

        # Últimas aprobaciones
        ws_ap = wb.create_sheet("Executive Audit")
        style_title(ws_ap, "EXECUTIVE AUDIT", "Últimas aprobaciones y trazabilidad ejecutiva")
        aprob_rows = [
            [
                a.get("tenant"),
                a.get("operacion"),
                a.get("riesgo"),
                a.get("fecha"),
            ]
            for a in dashboard.get("ultimas_aprobaciones", []) or []
        ]
        write_table(
            ws_ap,
            4,
            ["Tenant", "Operación", "Riesgo", "Fecha aprobación"],
            aprob_rows
        )

        ws_risk = wb.create_sheet("Risk & Licenses")
        style_title(ws_risk, "RISK & LICENSE CONTROL", "Licencias por vencer, tenants en riesgo y exposición SaaS")

        risk_license_rows = [
            ["Riesgo global", saas.get("riesgo_global", "BAJO")],
            ["Operaciones críticas", saas.get("operaciones_criticas", 0)],
            ["Licencias por vencer", saas.get("licencias_por_vencer_total", 0)],
            ["Tenants en riesgo", saas.get("tenants_en_riesgo_total", 0)],
            ["Activity Feed eventos", resumen.get("activity_feed_total", 0)],
        ]

        write_table(ws_risk, 4, ["Indicador", "Valor"], risk_license_rows)

        lic_rows = [
            [
                t.get("tenant"),
                t.get("plan"),
                t.get("estado_licencia"),
                t.get("dias_restantes"),
                t.get("fecha_vencimiento"),
            ]
            for t in dashboard.get("licencias_por_vencer", []) or []
        ]

        write_table(
            ws_risk,
            13,
            ["Tenant", "Plan", "Estado licencia", "Días restantes", "Vencimiento"],
            lic_rows
        )

        tenant_risk_rows = [
            [
                t.get("tenant"),
                t.get("plan"),
                t.get("estado_licencia"),
                t.get("dias_restantes"),
                t.get("operaciones_mes"),
            ]
            for t in dashboard.get("tenants_en_riesgo", []) or []
        ]

        write_table(
            ws_risk,
            22,
            ["Tenant", "Plan", "Estado licencia", "Días restantes", "Operaciones mes"],
            tenant_risk_rows
        )

        for sheet in wb.worksheets:
            sheet.sheet_view.showGridLines = False

        wb.save(output_path)

        return FileResponse(
            output_path,
            filename=output_name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={
                "ok": False,
                "error": str(e)
            }
        )


# =========================================================
# ARGO CONNECT - LAYOUT BUILDER / EXPORTADOR UNIVERSAL
# =========================================================

ARGO_CONNECT_CATALOGO = [
    # Operación
    {"campo": "id_operacion", "etiqueta": "ID Operación", "grupo": "Operación"},
    {"campo": "timestamp_local", "etiqueta": "Fecha / Hora", "grupo": "Operación"},
    {"campo": "fecha", "etiqueta": "Fecha", "grupo": "Operación"},
    {"campo": "hora", "etiqueta": "Hora", "grupo": "Operación"},
    {"campo": "referencia_operacion", "etiqueta": "Referencia Operación", "grupo": "Operación"},
    {"campo": "operador", "etiqueta": "Operador", "grupo": "Operación"},
    {"campo": "supervisor", "etiqueta": "Supervisor", "grupo": "Operación"},

    # Cliente / Comercial
    {"campo": "cliente_id", "etiqueta": "Cliente ID / Tenant", "grupo": "Cliente"},
    {"campo": "cliente_nombre", "etiqueta": "Cliente", "grupo": "Cliente"},
    {"campo": "cliente", "etiqueta": "Cliente documental", "grupo": "Cliente"},
    {"campo": "proveedor", "etiqueta": "Proveedor", "grupo": "Comercial"},
    {"campo": "shipper", "etiqueta": "Shipper", "grupo": "Comercial"},
    {"campo": "consignee", "etiqueta": "Consignee", "grupo": "Comercial"},

    # Logística
    {"campo": "shipment_id", "etiqueta": "Shipment ID", "grupo": "Logística"},
    {"campo": "tracking", "etiqueta": "Tracking", "grupo": "Logística"},
    {"campo": "paqueteria", "etiqueta": "Paquetería", "grupo": "Logística"},
    {"campo": "bl_awb_no", "etiqueta": "BL / AWB No.", "grupo": "Logística"},
    {"campo": "tipo_transporte", "etiqueta": "Tipo transporte", "grupo": "Logística"},
    {"campo": "fecha_embarque", "etiqueta": "Fecha embarque", "grupo": "Logística"},
    {"campo": "peso_total", "etiqueta": "Peso total", "grupo": "Logística"},
    {"campo": "peso", "etiqueta": "Peso", "grupo": "Logística"},
    {"campo": "cantidad_bultos", "etiqueta": "Cantidad bultos", "grupo": "Logística"},
    {"campo": "total_bultos", "etiqueta": "Total bultos", "grupo": "Logística"},
    {"campo": "cantidad", "etiqueta": "Cantidad", "grupo": "Logística"},
    {"campo": "unidad", "etiqueta": "Unidad", "grupo": "Logística"},

    # Mercancía
    {"campo": "descripcion", "etiqueta": "Descripción mercancía", "grupo": "Mercancía"},
    {"campo": "descripcion_soporte", "etiqueta": "Descripción soporte", "grupo": "Mercancía"},
    {"campo": "marca", "etiqueta": "Marca", "grupo": "Mercancía"},
    {"campo": "modelo", "etiqueta": "Modelo", "grupo": "Mercancía"},
    {"campo": "serie", "etiqueta": "Serie", "grupo": "Mercancía"},
    {"campo": "lote", "etiqueta": "Lote", "grupo": "Mercancía"},
    {"campo": "sku", "etiqueta": "SKU", "grupo": "Mercancía"},
    {"campo": "numero_parte", "etiqueta": "Número de parte", "grupo": "Mercancía"},

    # Comercio Exterior / Aduana
    {"campo": "fraccion_tigie", "etiqueta": "Fracción TIGIE", "grupo": "Aduana"},
    {"campo": "pais_origen", "etiqueta": "País origen", "grupo": "Aduana"},
    {"campo": "pais_procedencia", "etiqueta": "País procedencia", "grupo": "Aduana"},
    {"campo": "aduana", "etiqueta": "Aduana", "grupo": "Aduana"},
    {"campo": "regimen", "etiqueta": "Régimen", "grupo": "Aduana"},
    {"campo": "incoterm", "etiqueta": "Incoterm", "grupo": "Aduana"},
    {"campo": "moneda", "etiqueta": "Moneda", "grupo": "Aduana"},
    {"campo": "valor_aduana", "etiqueta": "Valor aduana", "grupo": "Aduana"},
    {"campo": "valor_comercial", "etiqueta": "Valor comercial", "grupo": "Aduana"},
    {"campo": "igi", "etiqueta": "IGI %", "grupo": "Aduana"},
    {"campo": "iva", "etiqueta": "IVA %", "grupo": "Aduana"},
    {"campo": "nom", "etiqueta": "NOM", "grupo": "Aduana"},
    {"campo": "rrna", "etiqueta": "RRNA", "grupo": "Aduana"},
    {"campo": "nico", "etiqueta": "NICO", "grupo": "Aduana"},

    # Documentos
    {"campo": "invoice_no", "etiqueta": "Invoice No.", "grupo": "Documentos"},
    {"campo": "packing_list_no", "etiqueta": "Packing List No.", "grupo": "Documentos"},
    {"campo": "control_output_path", "etiqueta": "Archivo control", "grupo": "Documentos"},
    {"campo": "document_output_path", "etiqueta": "Archivo documental", "grupo": "Documentos"},

    # ARGO CLASS
    {"campo": "fraccion_sugerida", "etiqueta": "Fracción sugerida", "grupo": "ARGO CLASS"},
    {"campo": "confianza_fraccion_pct", "etiqueta": "Confianza fracción %", "grupo": "ARGO CLASS"},
    {"campo": "certeza_final_pct", "etiqueta": "Certeza final %", "grupo": "ARGO CLASS"},
    {"campo": "score_documental", "etiqueta": "Score documental", "grupo": "ARGO CLASS"},
    {"campo": "score_documental_global", "etiqueta": "Score documental global", "grupo": "ARGO CLASS"},
    {"campo": "nivel_debida_diligencia", "etiqueta": "Debida diligencia", "grupo": "ARGO CLASS"},
    {"campo": "obs_class", "etiqueta": "Observaciones ARGO CLASS", "grupo": "ARGO CLASS"},

    # Riesgo / Aprobaciones
    {"campo": "riesgo_automatico", "etiqueta": "Riesgo automático", "grupo": "Riesgo"},
    {"campo": "riesgo_global", "etiqueta": "Riesgo global", "grupo": "Riesgo"},
    {"campo": "semaforo_operacion", "etiqueta": "Semáforo operación", "grupo": "Riesgo"},
    {"campo": "estatus_global", "etiqueta": "Estatus global", "grupo": "Riesgo"},
    {"campo": "alertas_totales", "etiqueta": "Alertas totales", "grupo": "Riesgo"},
    {"campo": "obs_control", "etiqueta": "Observaciones ARGO CONTROL", "grupo": "Riesgo"},
    {"campo": "aprobada", "etiqueta": "Aprobada", "grupo": "Aprobaciones"},
    {"campo": "aprobada_por", "etiqueta": "Aprobada por", "grupo": "Aprobaciones"},
    {"campo": "fecha_aprobacion", "etiqueta": "Fecha aprobación", "grupo": "Aprobaciones"},
]

def _argo_connect_store_path():
    os.makedirs("outputs", exist_ok=True)
    return os.path.join("outputs", "argo_connect_templates.json")

def _argo_connect_read_templates():
    import json
    path = _argo_connect_store_path()
    if not os.path.exists(path):
        return []
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
            return data if isinstance(data, list) else []
    except Exception:
        return []

def _argo_connect_write_templates(items):
    import json
    path = _argo_connect_store_path()
    with open(path, "w", encoding="utf-8") as f:
        json.dump(items, f, ensure_ascii=False, indent=2)

def _argo_connect_get_valor(op, campo):
    if not isinstance(op, dict):
        return ""

    aliases = {
        "tracking": ["tracking", "tracking_number", "shipment_id"],
        "cliente": ["cliente", "cliente_nombre", "customer", "consignee"],
        "cliente_nombre": ["cliente_nombre", "cliente", "customer"],
        "proveedor": ["proveedor", "supplier", "vendor", "shipper"],
        "descripcion": ["descripcion", "mercancia", "descripcion_soporte", "product_description"],
        "fraccion_tigie": ["fraccion_tigie", "fraccion_sugerida", "fraccion", "fraccion_arancelaria"],
        "pais_origen": ["pais_origen", "country_of_origin", "origin"],
        "peso": ["peso", "peso_total", "weight", "gross_weight"],
        "cantidad": ["cantidad", "cantidad_bultos", "total_bultos", "qty", "quantity"],
        "invoice_no": ["invoice_no", "factura", "invoice", "invoice_number"],
        "packing_list_no": ["packing_list_no", "packing_list", "packing_number"],
        "bl_awb_no": ["bl_awb_no", "awb", "bl", "bill_of_lading"],
        "valor_comercial": ["valor_comercial", "valor_aduana", "invoice_value", "commercial_value"],
        "score_documental_global": ["score_documental_global", "score_documental"],
        "operador": ["operador", "created_by", "usuario", "usuario_email"],
    }

    posibles = [campo] + aliases.get(campo, [])

    for key in posibles:
        if key in op and op.get(key) not in [None, ""]:
            v = op.get(key)
            break
    else:
        v = ""

    if isinstance(v, dict):
        v = v.get("valor") or v.get("value") or v.get("texto") or ""
    if isinstance(v, list):
        v = ", ".join(str(x) for x in v if x not in [None, ""])
    if isinstance(v, bool):
        return "SI" if v else "NO"
    if v is None:
        return ""
    return v

def _argo_connect_user(request):
    email = request.headers.get("x-usuario-email")
    usuario = obtener_usuario_rbac(email)
    if not usuario:
        return None
    return usuario

def _argo_connect_tenant(usuario):
    return usuario.get("id_cliente") or "SIN_TENANT"

@app.get("/argo/connect/catalogo")
async def argo_connect_catalogo(request: Request):
    try:
        usuario = _argo_connect_user(request)
        if not usuario:
            return JSONResponse(status_code=403, content={"ok": False, "error": "Usuario requerido"})

        grupos = {}
        for c in ARGO_CONNECT_CATALOGO:
            grupos.setdefault(c["grupo"], []).append(c)

        return {
            "ok": True,
            "catalogo": ARGO_CONNECT_CATALOGO,
            "grupos": grupos,
            "formatos": ["xlsx", "csv", "txt"],
            "separadores": [
                {"label": "Coma (,)", "value": ","},
                {"label": "Punto y coma (;)", "value": ";"},
                {"label": "Tabulador", "value": "\\t"},
                {"label": "Pipe (|)", "value": "|"},
            ],
        }
    except Exception as e:
        return JSONResponse(status_code=500, content={"ok": False, "error": str(e)})

@app.get("/argo/connect/plantillas")
async def argo_connect_listar_plantillas(request: Request):
    try:
        usuario = _argo_connect_user(request)
        if not usuario:
            return JSONResponse(status_code=403, content={"ok": False, "error": "Usuario requerido"})

        rol = str(usuario.get("rol") or "").lower()
        tenant = _argo_connect_tenant(usuario)

        items = _argo_connect_read_templates()
        if rol != "master_admin":
            items = [x for x in items if x.get("tenant_id") == tenant]

        return {"ok": True, "plantillas": items, "total": len(items)}
    except Exception as e:
        return JSONResponse(status_code=500, content={"ok": False, "error": str(e)})

@app.post("/argo/connect/plantillas")
async def argo_connect_crear_plantilla(request: Request, payload: dict):
    try:
        usuario = _argo_connect_user(request)
        if not usuario:
            return JSONResponse(status_code=403, content={"ok": False, "error": "Usuario requerido"})

        tenant = _argo_connect_tenant(usuario)
        nombre = str(payload.get("nombre") or "").strip()
        formato = str(payload.get("formato") or "xlsx").lower().strip()
        separador = payload.get("separador") or ","
        columnas = payload.get("columnas") or []

        if not nombre:
            return {"ok": False, "error": "Nombre requerido"}
        if formato not in ["xlsx", "csv", "txt"]:
            return {"ok": False, "error": "Formato inválido"}
        if not isinstance(columnas, list) or not columnas:
            return {"ok": False, "error": "Debe seleccionar al menos una columna"}

        import uuid
        from datetime import datetime

        items = _argo_connect_read_templates()
        plantilla = {
            "id": str(uuid.uuid4()),
            "tenant_id": tenant,
            "nombre": nombre,
            "descripcion": payload.get("descripcion") or "",
            "formato": formato,
            "separador": separador,
            "columnas": columnas,
            "created_at": datetime.now().replace(microsecond=0).isoformat(),
            "created_by": usuario.get("email"),
        }
        items.append(plantilla)
        _argo_connect_write_templates(items)

        return {"ok": True, "plantilla": plantilla}
    except Exception as e:
        return JSONResponse(status_code=500, content={"ok": False, "error": str(e)})

@app.patch("/argo/connect/plantillas/{plantilla_id}")
async def argo_connect_actualizar_plantilla(plantilla_id: str, request: Request, payload: dict):
    try:
        usuario = _argo_connect_user(request)
        if not usuario:
            return JSONResponse(status_code=403, content={"ok": False, "error": "Usuario requerido"})

        rol = str(usuario.get("rol") or "").lower()
        tenant = _argo_connect_tenant(usuario)
        items = _argo_connect_read_templates()

        encontrada = None
        for x in items:
            if x.get("id") == plantilla_id:
                encontrada = x
                break

        if not encontrada:
            return {"ok": False, "error": "Plantilla no encontrada"}

        if rol != "master_admin" and encontrada.get("tenant_id") != tenant:
            return JSONResponse(status_code=403, content={"ok": False, "error": "Acceso denegado"})

        for k in ["nombre", "descripcion", "formato", "separador", "columnas"]:
            if k in payload:
                encontrada[k] = payload[k]

        _argo_connect_write_templates(items)
        return {"ok": True, "plantilla": encontrada}
    except Exception as e:
        return JSONResponse(status_code=500, content={"ok": False, "error": str(e)})

@app.delete("/argo/connect/plantillas/{plantilla_id}")
async def argo_connect_eliminar_plantilla(plantilla_id: str, request: Request):
    try:
        usuario = _argo_connect_user(request)
        if not usuario:
            return JSONResponse(status_code=403, content={"ok": False, "error": "Usuario requerido"})

        rol = str(usuario.get("rol") or "").lower()
        tenant = _argo_connect_tenant(usuario)
        items = _argo_connect_read_templates()

        nueva = []
        eliminado = None

        for x in items:
            if x.get("id") == plantilla_id:
                if rol != "master_admin" and x.get("tenant_id") != tenant:
                    return JSONResponse(status_code=403, content={"ok": False, "error": "Acceso denegado"})
                eliminado = x
                continue
            nueva.append(x)

        if not eliminado:
            return {"ok": False, "error": "Plantilla no encontrada"}

        _argo_connect_write_templates(nueva)
        return {"ok": True, "eliminado": eliminado}
    except Exception as e:
        return JSONResponse(status_code=500, content={"ok": False, "error": str(e)})

@app.post("/argo/connect/exportar")
async def argo_connect_exportar(request: Request, payload: dict):
    try:
        usuario = _argo_connect_user(request)
        if not usuario:
            return JSONResponse(status_code=403, content={"ok": False, "error": "Usuario requerido"})

        rol = str(usuario.get("rol") or "").lower()
        tenant = _argo_connect_tenant(usuario)

        plantilla_id = payload.get("plantilla_id")
        operaciones_ids = payload.get("operaciones") or []
        formato_override = str(payload.get("formato") or "").lower().strip()

        items = _argo_connect_read_templates()
        plantilla = next((x for x in items if x.get("id") == plantilla_id), None)

        if not plantilla:
            return {"ok": False, "error": "Plantilla no encontrada"}

        if rol != "master_admin" and plantilla.get("tenant_id") != tenant:
            return JSONResponse(status_code=403, content={"ok": False, "error": "Acceso denegado"})

        formato = formato_override or str(plantilla.get("formato") or "xlsx").lower()
        if formato not in ["xlsx", "csv", "txt"]:
            formato = "xlsx"

        orientacion = str(plantilla.get("orientacion") or "horizontal").lower().strip()
        if orientacion not in ["horizontal", "vertical"]:
            orientacion = "horizontal"

        separador = plantilla.get("separador") or ","
        separador_real = "\t" if separador == "\\t" else separador

        columnas = plantilla.get("columnas") or []
        if not columnas:
            return {"ok": False, "error": "Plantilla sin columnas"}

        historial_raw = obtener_historial(tenant if rol != "master_admin" else None)
        if isinstance(historial_raw, dict):
            operaciones = historial_raw.get("operaciones") or historial_raw.get("historial") or historial_raw.get("data") or []
        elif isinstance(historial_raw, list):
            operaciones = historial_raw
        else:
            operaciones = []

        if operaciones_ids:
            operaciones = [op for op in operaciones if op.get("id_operacion") in operaciones_ids]

        if not operaciones:
            return {"ok": False, "error": "No hay operaciones para exportar"}

        def resolver_valor(col, op, idx_op):
            tipo = str(col.get("tipo") or "campo").lower().strip()

            if tipo == "vacio":
                return ""

            if tipo == "texto_fijo":
                return col.get("valor_fijo") or ""

            if tipo == "fecha_actual":
                return datetime.now().strftime("%Y-%m-%d")

            if tipo == "usuario_actual":
                return usuario.get("email") or usuario.get("nombre") or ""

            if tipo == "secuencia":
                inicio = col.get("inicio", 1)
                try:
                    inicio = int(inicio)
                except Exception:
                    inicio = 1
                return inicio + idx_op

            if tipo in ["formula", "concatenacion"]:
                plantilla_formula = str(col.get("valor_fijo") or "")
                resultado = plantilla_formula
                for campo_def in ARGO_CONNECT_CATALOGO:
                    key = campo_def.get("campo")
                    if not key:
                        continue
                    resultado = resultado.replace("{" + key + "}", str(_argo_connect_get_valor(op, key)))
                resultado = resultado.replace("{secuencia}", str(idx_op + 1))
                resultado = resultado.replace("{fecha_actual}", datetime.now().strftime("%Y-%m-%d"))
                resultado = resultado.replace("{usuario_actual}", usuario.get("email") or usuario.get("nombre") or "")
                return resultado

            return _argo_connect_get_valor(op, col.get("campo"))

        headers = [c.get("titulo") or c.get("etiqueta") or c.get("campo") or c.get("tipo") or "" for c in columnas]
        rows = []

        if orientacion == "vertical":
            headers = ["Campo", "Valor"]
            for idx_op, op in enumerate(operaciones):
                if idx_op:
                    rows.append(["", ""])
                rows.append(["OPERACION", op.get("id_operacion") or f"OPERACION_{idx_op + 1}"])
                for c in columnas:
                    rows.append([
                        c.get("titulo") or c.get("etiqueta") or c.get("campo") or c.get("tipo") or "",
                        resolver_valor(c, op, idx_op),
                    ])
        else:
            for idx_op, op in enumerate(operaciones):
                rows.append([resolver_valor(c, op, idx_op) for c in columnas])

        safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", plantilla.get("nombre") or "ARGO_CONNECT").strip("_")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"ARGO_CONNECT_{safe_name}_{timestamp}.{formato}"
        output_path = os.path.join("outputs", filename)
        os.makedirs("outputs", exist_ok=True)

        if formato == "xlsx":
            wb = Workbook()
            ws = wb.active
            ws.title = "ARGO Connect"
            ws.append(headers)
            for row in rows:
                ws.append(row)

            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 48)

            wb.save(output_path)
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        else:
            import csv
            with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f, delimiter=separador_real)
                writer.writerow(headers)
                writer.writerows(rows)

            media_type = "text/csv" if formato == "csv" else "text/plain"

        return FileResponse(output_path, filename=filename, media_type=media_type)

    except Exception as e:
        return JSONResponse(status_code=500, content={"ok": False, "error": str(e)})



app.mount("/", StaticFiles(directory="dist", html=True), name="frontend")

