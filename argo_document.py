# argo_document.py
from __future__ import annotations

import os
import re
import json
import hashlib
import shutil
from dataclasses import dataclass, asdict
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


# =========================================================
# CONFIG
# =========================================================

SCHEMA_NAME = "ARGO_DOCUMENT_OUTPUT_V2026"
VERSION_MOTOR = "2026.1"
MODO = "TRANSACCIONAL"

ESTATUS_ICONOS = {
    "GENERADO_OK": "🟢",
    "CON_OBSERVACIONES": "🟡",
    "REQUIERE_REVISION": "🟠",
    "ERROR_TECNICO": "🔴",
}


# =========================================================
# DATACLASSES
# =========================================================

@dataclass
class InventarioDocumento:
    tipo_documento: str
    presente: bool
    archivo: str
    referencia: str
    fecha: str
    observaciones: str


@dataclass
class MapeoCampo:
    campo: str
    valor: str
    fuente: str
    validacion: str
    observaciones: str
    evidencia: Dict[str, str]


@dataclass
class AlertaDocumento:
    codigo: str
    severidad: str
    campo: str
    mensaje: str
    fuente: str
    impacto: str
    accion_recomendada: str


@dataclass
class ScoreDocumentalDoc:
    score_total: int
    nivel: str
    componentes: Dict[str, int]
    impacto_en_score_global: Dict[str, Any]


@dataclass
class ResumenDocumento:
    documentos_detectados: int
    documentos_requeridos: int
    faltantes: int
    campos_total: int
    campos_ok: int
    campos_error: int
    campos_no_verificables: int
    severidad_maxima: str
    dictamen: str


@dataclass
class MetaDocumento:
    schema: str
    version_motor: str
    modo: str
    id_operacion: str
    timestamp_local: str
    hash_input: str
    input_filename: str
    plantilla_filename: str
    expediente_resumen: Dict[str, Any]


@dataclass
class ArgoDocumentSalida:
    ok: bool
    modulo: str
    meta: Dict[str, Any]
    estatus: str
    icono: str
    output_path: str
    resumen: Dict[str, Any]
    inventario_documental: List[Dict[str, Any]]
    mapeo_campos: List[Dict[str, Any]]
    alertas: List[Dict[str, Any]]
    score_documental_doc: Dict[str, Any]
    advertencia_juridica: str
    errores_tecnicos: List[Dict[str, Any]]


# =========================================================
# HELPERS
# =========================================================

def _ensure_dir(path: str) -> None:
    os.makedirs(path, exist_ok=True)


def _sha256_file(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def _build_id_operacion() -> str:
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    rand = hashlib.md5(f"{datetime.now().isoformat()}".encode()).hexdigest()[:8].upper()
    return f"OP-{stamp}-{rand}"


def _extract_id_operacion_from_filename(filename: str) -> Optional[str]:
    if not filename:
        return None
    m = re.search(r"(OP-\d{8}-\d{6}-[A-Z0-9]+)", filename)
    return m.group(1) if m else None


def _now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _safe_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", _safe_str(value)).strip().lower()


def _score_to_level(score: int) -> str:
    if score <= 39:
        return "BAJO"
    if score <= 69:
        return "MEDIO"
    return "ALTO"


def _severidad_rank(sev: str) -> int:
    order = {"BAJA": 1, "MEDIA": 2, "ALTA": 3, "CRITICA": 4}
    return order.get((_safe_str(sev).upper()), 0)


def _max_severidad(values: List[str]) -> str:
    if not values:
        return "BAJA"
    return max(values, key=_severidad_rank)


def _pick_first_nonempty(*values: Any) -> str:
    for v in values:
        s = _safe_str(v)
        if s:
            return s
    return ""


# =========================================================
# LECTURA BEST-EFFORT DEL XLSX DE ENTRADA
# =========================================================

def _find_value_by_labels(wb, labels: List[str]) -> Tuple[str, str]:
    """
    Busca por etiqueta en las primeras columnas del workbook de entrada.
    Regresa (valor, hoja_fuente)
    """
    labels_norm = {_normalize_text(x) for x in labels}

    for ws in wb.worksheets:
        max_row = min(ws.max_row, 250)
        max_col = min(ws.max_column, 8)

        # Caso 1: etiqueta en col A, valor en col B
        for r in range(1, max_row + 1):
            a = _normalize_text(ws.cell(r, 1).value)
            b = _safe_str(ws.cell(r, 2).value)
            if a in labels_norm and b:
                return b, ws.title

        # Caso 2: buscar etiqueta en matriz y tomar celda derecha
        for r in range(1, max_row + 1):
            for c in range(1, max_col):
                current = _normalize_text(ws.cell(r, c).value)
                if current in labels_norm:
                    right = _safe_str(ws.cell(r, c + 1).value)
                    if right:
                        return right, ws.title

    return "", ""


def _extract_input_data(input_xlsx_path: str) -> Dict[str, Dict[str, str]]:
    """
    Extrae datos best-effort desde el XLSX del pipeline.
    Formato:
    {
      "cliente": {"valor": "...", "fuente": "..."},
      ...
    }
    """
    wb = load_workbook(input_xlsx_path, data_only=True)

    mappings = {
        "shipment_id": ["shipment id", "id shipment", "id_shipment", "shipment_id"],
        "cliente": ["cliente", "customer", "consignee"],
        "proveedor": ["proveedor", "supplier", "vendor", "shipper"],
        "fraccion_tigie": ["fracción tigie", "fraccion tigie", "fracción", "fraccion", "tigie", "fraccion arancelaria"],
        "descripcion": ["descripción", "descripcion", "descripción mercancía", "descripcion mercancia", "product description"],
        "valor_aduana": ["valor aduana", "customs value", "valor"],
        "pais_origen": ["país origen", "pais origen", "country of origin", "origin"],
        "cantidad": ["cantidad", "qty", "quantity"],
        "unidad": ["unidad", "unit", "uom"],
        "peso": ["peso", "weight", "gross weight"],
        "regimen": ["régimen", "regimen"],
        "aduana": ["aduana", "customs office"],
        "fecha": ["fecha", "date", "fecha operación", "fecha operacion"],
        "referencia_operacion": ["referencia operación", "referencia operacion", "reference", "operation reference"],
        "invoice_no": ["invoice no.", "invoice no", "invoice number", "factura", "número factura", "numero factura"],
        "packing_list_no": ["packing list no.", "packing list no", "packing number", "packing list"],
        "bl_awb_no": ["bl/awb no.", "bl/awb", "bill of lading", "air waybill", "awb", "bl"],
        "moneda": ["moneda", "currency"],
        "incoterm": ["incoterm"],
        "pais_procedencia": ["país procedencia", "pais procedencia", "country of export", "country of shipment"],
        "shipper": ["shipper"],
        "consignee": ["consignee"],
        "fecha_embarque": ["fecha embarque", "shipping date", "shipment date"],
        "tipo_transporte": ["tipo transporte", "transport mode", "modo transporte"],
        "total_bultos": ["total bultos", "packages", "total packages"],
        "valor_comercial": ["valor comercial", "commercial value", "invoice value"],
        "descripcion_soporte": ["descripción mercancía (soporte)", "descripcion mercancia (soporte)", "commercial description"],
        "obs_control": ["observaciones argo control", "observaciones control", "control observations"],
        "obs_class": ["observaciones argo class", "observaciones class", "class observations"],
        "score_documental_global": ["score documental", "score documental (global)", "document score"],
        "riesgo_automatico": ["riesgo automático", "riesgo automatico", "automatic risk"],
        "nivel_debida_diligencia": ["nivel debida diligencia", "due diligence level"],
    }

    extracted: Dict[str, Dict[str, str]] = {}

    for key, labels in mappings.items():
        value, source_sheet = _find_value_by_labels(wb, labels)
        extracted[key] = {
            "valor": value,
            "fuente": source_sheet or "",
        }

    wb.close()
    return extracted


# =========================================================
# INVENTARIO DOCUMENTAL BLOQUE 1
# =========================================================

def _build_inventario_documental(input_filename: str, data: Dict[str, Dict[str, str]]) -> List[InventarioDocumento]:
    invoice_ref = data.get("invoice_no", {}).get("valor", "")
    packing_ref = data.get("packing_list_no", {}).get("valor", "")
    bl_awb_ref = data.get("bl_awb_no", {}).get("valor", "")

    tipo_transporte = _normalize_text(data.get("tipo_transporte", {}).get("valor", ""))
    has_awb_hint = "air" in tipo_transporte or "aereo" in tipo_transporte or "air waybill" in tipo_transporte
    has_bl_hint = "sea" in tipo_transporte or "ocean" in tipo_transporte or "mar" in tipo_transporte or "bill of lading" in tipo_transporte

    inventario = [
        InventarioDocumento(
            tipo_documento="INVOICE",
            presente=bool(invoice_ref),
            archivo=input_filename if invoice_ref else "",
            referencia=invoice_ref,
            fecha="",
            observaciones="" if invoice_ref else "No se localizó referencia de invoice en el XLSX de entrada."
        ),
        InventarioDocumento(
            tipo_documento="PACKING_LIST",
            presente=bool(packing_ref),
            archivo=input_filename if packing_ref else "",
            referencia=packing_ref,
            fecha="",
            observaciones="" if packing_ref else "No se localizó referencia de packing list en el XLSX de entrada."
        ),
        InventarioDocumento(
            tipo_documento="BILL_OF_LADING",
            presente=bool(bl_awb_ref) and not has_awb_hint,
            archivo=input_filename if bl_awb_ref and not has_awb_hint else "",
            referencia=bl_awb_ref if not has_awb_hint else "",
            fecha="",
            observaciones="" if (bl_awb_ref and not has_awb_hint) else "No verificable en Bloque 1."
        ),
        InventarioDocumento(
            tipo_documento="AWB",
            presente=bool(bl_awb_ref) and has_awb_hint,
            archivo=input_filename if bl_awb_ref and has_awb_hint else "",
            referencia=bl_awb_ref if has_awb_hint else "",
            fecha="",
            observaciones="" if (bl_awb_ref and has_awb_hint) else "No verificable en Bloque 1."
        ),
        InventarioDocumento(
            tipo_documento="CERTIFICATE_OF_ORIGIN",
            presente=False,
            archivo="",
            referencia="",
            fecha="",
            observaciones="No verificable en Bloque 1."
        ),
        InventarioDocumento(
            tipo_documento="PEDIMENTO",
            presente=False,
            archivo="",
            referencia="",
            fecha="",
            observaciones="No verificable en Bloque 1."
        ),
        InventarioDocumento(
            tipo_documento="FICHA_TECNICA",
            presente=False,
            archivo="",
            referencia="",
            fecha="",
            observaciones="No verificable en Bloque 1."
        ),
        InventarioDocumento(
            tipo_documento="CARTA_PROVEEDOR",
            presente=False,
            archivo="",
            referencia="",
            fecha="",
            observaciones="No verificable en Bloque 1."
        ),
        InventarioDocumento(
            tipo_documento="ARGO_CONTROL_OUTPUT",
            presente=True,
            archivo=input_filename,
            referencia="OUTPUT_XLSX",
            fecha="",
            observaciones=""
        ),
        InventarioDocumento(
            tipo_documento="ARGO_CLASS_OUTPUT",
            presente=True,
            archivo=input_filename,
            referencia="OUTPUT_XLSX",
            fecha="",
            observaciones=""
        ),
    ]

    return inventario


# =========================================================
# MAPEOS DE LA PLANTILLA
# =========================================================

EXPEDIENTE_ROW_MAP = {
    "Shipment ID": "shipment_id",
    "Cliente": "cliente",
    "Proveedor": "proveedor",
    "Fracción TIGIE": "fraccion_tigie",
    "Descripción": "descripcion",
    "Valor Aduana": "valor_aduana",
    "País Origen": "pais_origen",
    "Cantidad": "cantidad",
    "Unidad": "unidad",
    "Peso": "peso",
    "Régimen": "regimen",
    "Aduana": "aduana",
    "Fecha": "fecha",
    "Referencia Operación": "referencia_operacion",
    "Invoice No.": "invoice_no",
    "Packing List No.": "packing_list_no",
    "BL/AWB No.": "bl_awb_no",
    "Moneda": "moneda",
    "Incoterm": "incoterm",
    "País Procedencia": "pais_procedencia",
    "Shipper": "shipper",
    "Consignee": "consignee",
    "Fecha Embarque": "fecha_embarque",
    "Tipo Transporte": "tipo_transporte",
    "Total Bultos": "total_bultos",
    "Valor Comercial": "valor_comercial",
    "Descripción Mercancía (soporte)": "descripcion_soporte",
    "Observaciones ARGO CONTROL": "obs_control",
    "Observaciones ARGO CLASS": "obs_class",
    "Score Documental (global)": "score_documental_global",
    "Riesgo Automático": "riesgo_automatico",
    "Nivel Debida Diligencia": "nivel_debida_diligencia",
}

DOCUMENTOS_ROW_MAP = {
    "Commercial Invoice": "INVOICE",
    "Packing List": "PACKING_LIST",
    "Bill of Lading": "BILL_OF_LADING",
    "Air Waybill (AWB)": "AWB",
    "Certificate of Origin (COO)": "CERTIFICATE_OF_ORIGIN",
    "Pedimento (si aplica)": "PEDIMENTO",
    "Ficha técnica": "FICHA_TECNICA",
    "Carta del proveedor": "CARTA_PROVEEDOR",
    "ARGO CONTROL Output": "ARGO_CONTROL_OUTPUT",
    "ARGO CLASS Output": "ARGO_CLASS_OUTPUT",
}

VALIDACIONES_BASE = [
    ("Referencia Operación", "MEDIA"),
    ("Cliente", "ALTA"),
    ("Proveedor", "ALTA"),
    ("Descripción Mercancía", "CRITICA"),
    ("Cantidad vs Packing", "ALTA"),
    ("Peso vs Packing", "MEDIA"),
    ("Valor vs Invoice", "ALTA"),
    ("Moneda", "MEDIA"),
    ("País Origen", "ALTA"),
    ("Incoterm", "MEDIA"),
    ("BL/AWB vs Referencias", "MEDIA"),
    ("Fechas consistentes", "MEDIA"),
]


# =========================================================
# ESCRITURA EN PLANTILLA
# =========================================================

def _find_row_by_label(ws, label: str, label_col: int = 1, start_row: int = 1, end_row: int = 400) -> Optional[int]:
    target = _normalize_text(label)
    for r in range(start_row, min(end_row, ws.max_row) + 1):
        if _normalize_text(ws.cell(r, label_col).value) == target:
            return r
    return None


def _fill_expediente(ws, data: Dict[str, Dict[str, str]]) -> List[MapeoCampo]:
    mapeo_campos: List[MapeoCampo] = []

    for label, data_key in EXPEDIENTE_ROW_MAP.items():
        row = _find_row_by_label(ws, label, label_col=1, start_row=4)
        if not row:
            continue

        valor = _safe_str(data.get(data_key, {}).get("valor", ""))
        fuente_hoja = _safe_str(data.get(data_key, {}).get("fuente", ""))

        if valor:
            fuente = "Expediente" if not fuente_hoja else fuente_hoja
            validacion = "OK"
            observaciones = ""
            evidencia_archivo = fuente_hoja or "XLSX entrada"
            evidencia_ubicacion = "No verificable"
        else:
            valor = "No verificable"
            fuente = "No verificable"
            validacion = "No verificable"
            observaciones = "No se localizó evidencia suficiente en el XLSX de entrada."
            evidencia_archivo = "No verificable"
            evidencia_ubicacion = "No verificable"

        ws.cell(row, 2).value = valor
        ws.cell(row, 3).value = fuente
        ws.cell(row, 4).value = evidencia_archivo

        mapeo_campos.append(
            MapeoCampo(
                campo=label,
                valor=valor,
                fuente=fuente,
                validacion=validacion,
                observaciones=observaciones,
                evidencia={
                    "archivo": evidencia_archivo,
                    "ubicacion": evidencia_ubicacion,
                }
            )
        )

    return mapeo_campos


def _fill_documentos(ws, inventario: List[InventarioDocumento]) -> None:
    by_tipo = {x.tipo_documento: x for x in inventario}

    for label, tipo in DOCUMENTOS_ROW_MAP.items():
        row = _find_row_by_label(ws, label, label_col=1, start_row=4)
        if not row:
            continue

        doc = by_tipo.get(tipo)
        if not doc:
            continue

        ws.cell(row, 2).value = doc.referencia
        ws.cell(row, 3).value = "SI" if doc.presente else "NO"
        ws.cell(row, 4).value = doc.archivo
        if doc.observaciones:
            ws.cell(row, 7).value = doc.observaciones


def _fill_validaciones(ws, data: Dict[str, Dict[str, str]]) -> Tuple[int, int, int, List[AlertaDocumento], List[str]]:
    campos_ok = 0
    campos_error = 0
    campos_no_verificables = 0
    alertas: List[AlertaDocumento] = []
    severidades_detectadas: List[str] = []

    # Reglas simples Bloque 1
    rules = {
        "Referencia Operación": bool(data.get("referencia_operacion", {}).get("valor")),
        "Cliente": bool(data.get("cliente", {}).get("valor")),
        "Proveedor": bool(data.get("proveedor", {}).get("valor")),
        "Descripción Mercancía": bool(_pick_first_nonempty(
            data.get("descripcion_soporte", {}).get("valor"),
            data.get("descripcion", {}).get("valor")
        )),
        "Cantidad vs Packing": bool(data.get("cantidad", {}).get("valor")) and bool(data.get("packing_list_no", {}).get("valor")),
        "Peso vs Packing": bool(data.get("peso", {}).get("valor")) and bool(data.get("packing_list_no", {}).get("valor")),
        "Valor vs Invoice": bool(_pick_first_nonempty(
            data.get("valor_comercial", {}).get("valor"),
            data.get("valor_aduana", {}).get("valor")
        )) and bool(data.get("invoice_no", {}).get("valor")),
        "Moneda": bool(data.get("moneda", {}).get("valor")),
        "País Origen": bool(data.get("pais_origen", {}).get("valor")),
        "Incoterm": bool(data.get("incoterm", {}).get("valor")),
        "BL/AWB vs Referencias": bool(data.get("bl_awb_no", {}).get("valor")),
        "Fechas consistentes": bool(_pick_first_nonempty(
            data.get("fecha_embarque", {}).get("valor"),
            data.get("fecha", {}).get("valor")
        )),
    }

    recomendaciones = {
        "Referencia Operación": "Validar referencia operativa contra salida del pipeline.",
        "Cliente": "Confirmar cliente contra invoice o packing list.",
        "Proveedor": "Confirmar proveedor contra invoice o carta proveedor.",
        "Descripción Mercancía": "Validar descripción contra invoice y ARGO CLASS.",
        "Cantidad vs Packing": "Adjuntar packing list y confirmar cantidades.",
        "Peso vs Packing": "Adjuntar packing list y confirmar pesos.",
        "Valor vs Invoice": "Validar valor contra invoice final.",
        "Moneda": "Confirmar moneda comercial del invoice.",
        "País Origen": "Adjuntar COO o ficha técnica.",
        "Incoterm": "Confirmar incoterm con invoice o proveedor.",
        "BL/AWB vs Referencias": "Adjuntar documento de transporte.",
        "Fechas consistentes": "Validar fecha de operación y fecha de embarque.",
    }

    fuentes = {
        "Referencia Operación": data.get("referencia_operacion", {}).get("fuente", ""),
        "Cliente": data.get("cliente", {}).get("fuente", ""),
        "Proveedor": data.get("proveedor", {}).get("fuente", ""),
        "Descripción Mercancía": _pick_first_nonempty(data.get("descripcion_soporte", {}).get("fuente", ""), data.get("descripcion", {}).get("fuente", "")),
        "Cantidad vs Packing": data.get("packing_list_no", {}).get("fuente", ""),
        "Peso vs Packing": data.get("packing_list_no", {}).get("fuente", ""),
        "Valor vs Invoice": data.get("invoice_no", {}).get("fuente", ""),
        "Moneda": data.get("moneda", {}).get("fuente", ""),
        "País Origen": data.get("pais_origen", {}).get("fuente", ""),
        "Incoterm": data.get("incoterm", {}).get("fuente", ""),
        "BL/AWB vs Referencias": data.get("bl_awb_no", {}).get("fuente", ""),
        "Fechas consistentes": _pick_first_nonempty(data.get("fecha_embarque", {}).get("fuente", ""), data.get("fecha", {}).get("fuente", "")),
    }

    for campo, severidad_base in VALIDACIONES_BASE:
        row = _find_row_by_label(ws, campo, label_col=1, start_row=4)
        if not row:
            continue

        ok = rules.get(campo, False)
        fuente = _safe_str(fuentes.get(campo, "")) or "No verificable"

        if ok:
            estado = "OK"
            observaciones = ""
            accion = ""
            campos_ok += 1
        else:
            estado = "No verificable"
            observaciones = "No se encontró soporte suficiente en el Bloque 1."
            accion = recomendaciones.get(campo, "Validar manualmente.")
            campos_no_verificables += 1
            severidades_detectadas.append(severidad_base)

            alertas.append(
                AlertaDocumento(
                    codigo="DOC-VAL-001",
                    severidad=severidad_base,
                    campo=campo,
                    mensaje=f"No fue posible validar '{campo}' con soporte suficiente en Bloque 1.",
                    fuente=fuente,
                    impacto="Afecta certeza documental",
                    accion_recomendada=accion,
                )
            )

        ws.cell(row, 2).value = estado
        ws.cell(row, 3).value = observaciones
        ws.cell(row, 4).value = fuente
        # columna E ya trae severidad base
        ws.cell(row, 6).value = accion

    return campos_ok, campos_error, campos_no_verificables, alertas, severidades_detectadas


def _fill_resumen(
    ws,
    estatus: str,
    output_path: str,
    id_operacion: str,
    score_total: int,
    delta_global: int,
    severidad_maxima: str,
) -> None:
    icono = ESTATUS_ICONOS.get(estatus, "🟡")
    ws["B3"] = estatus
    ws["E3"] = icono
    ws["B4"] = estatus
    ws["E4"] = severidad_maxima
    ws["B8"] = score_total
    ws["E8"] = delta_global
    ws["B9"] = output_path
    ws["E9"] = id_operacion


# =========================================================
# SCORE
# =========================================================

def _build_score_documental(
    resumen: ResumenDocumento,
    inventario: List[InventarioDocumento],
) -> ScoreDocumentalDoc:
    documentos_requeridos = max(resumen.documentos_requeridos, 1)
    campos_total = max(resumen.campos_total, 1)

    completitud = max(0, round(100 - ((resumen.faltantes / documentos_requeridos) * 100)))
    consistencia = max(0, round(100 - ((resumen.campos_error / campos_total) * 100)))
    verificabilidad = max(0, round(100 - ((resumen.campos_no_verificables / campos_total) * 100)))
    trazabilidad = 65 if any(x.presente for x in inventario) else 30

    score_total = round((completitud + consistencia + verificabilidad + trazabilidad) / 4)

    # Delta capado simple
    delta = 0
    if score_total >= 85:
        delta = 5
    elif score_total >= 70:
        delta = 2
    elif score_total >= 50:
        delta = -2
    else:
        delta = -6

    return ScoreDocumentalDoc(
        score_total=score_total,
        nivel=_score_to_level(score_total),
        componentes={
            "completitud": completitud,
            "consistencia": consistencia,
            "verificabilidad": verificabilidad,
            "trazabilidad": trazabilidad,
        },
        impacto_en_score_global={
            "metodo": "ADITIVO_CAPADO",
            "delta": delta,
        }
    )


# =========================================================
# ESTATUS
# =========================================================

def _determine_estatus(
    errores_tecnicos: List[Dict[str, str]],
    resumen: ResumenDocumento,
) -> str:
    if errores_tecnicos:
        return "ERROR_TECNICO"

    if resumen.campos_error > 0 or resumen.severidad_maxima in ("ALTA", "CRITICA"):
        return "REQUIERE_REVISION"

    if resumen.campos_no_verificables > 0 or resumen.faltantes > 0:
        return "CON_OBSERVACIONES"

    return "GENERADO_OK"


# =========================================================
# PROCESO PRINCIPAL
# =========================================================

def argo_document_bloque1(
    input_xlsx_path: str,
    plantilla_path: str,
    outputs_dir: str = "outputs",
    id_operacion: Optional[str] = None,
) -> ArgoDocumentSalida:
    errores_tecnicos: List[Dict[str, str]] = []

    try:
        _ensure_dir(outputs_dir)

        input_filename = os.path.basename(input_xlsx_path)
        plantilla_filename = os.path.basename(plantilla_path)

        if not id_operacion:
            id_operacion = _extract_id_operacion_from_filename(input_filename) or _build_id_operacion()

        # Validaciones básicas
        if not os.path.exists(input_xlsx_path):
            raise FileNotFoundError(f"No existe input_xlsx_path: {input_xlsx_path}")

        if not os.path.exists(plantilla_path):
            raise FileNotFoundError(f"No existe plantilla_path: {plantilla_path}")

        # Leer input
        data = _extract_input_data(input_xlsx_path)

        # Cargar plantilla
        wb = load_workbook(plantilla_path)
        required_sheets = ["Expediente", "Documentos", "Validaciones", "Resumen"]
        for sh in required_sheets:
            if sh not in wb.sheetnames:
                raise ValueError(f"La hoja '{sh}' no existe en la plantilla.")

        ws_expediente = wb["Expediente"]
        ws_documentos = wb["Documentos"]
        ws_validaciones = wb["Validaciones"]
        ws_resumen = wb["Resumen"]

        # Llenado
        inventario = _build_inventario_documental(input_filename, data)
        mapeo_campos = _fill_expediente(ws_expediente, data)
        _fill_documentos(ws_documentos, inventario)

        campos_ok, campos_error, campos_no_verificables, alertas, severidades_detectadas = _fill_validaciones(
            ws_validaciones, data
        )

        documentos_detectados = sum(1 for x in inventario if x.presente)
        documentos_requeridos = 4  # invoice + packing + control + class (Bloque 1 base)
        faltantes = 0
        if not any(x.tipo_documento == "INVOICE" and x.presente for x in inventario):
            faltantes += 1
        if not any(x.tipo_documento == "PACKING_LIST" and x.presente for x in inventario):
            faltantes += 1

        campos_total = len(VALIDACIONES_BASE)
        severidad_maxima = _max_severidad(severidades_detectadas) if severidades_detectadas else "BAJA"

        resumen = ResumenDocumento(
            documentos_detectados=documentos_detectados,
            documentos_requeridos=documentos_requeridos,
            faltantes=faltantes,
            campos_total=campos_total,
            campos_ok=campos_ok,
            campos_error=campos_error,
            campos_no_verificables=campos_no_verificables,
            severidad_maxima=severidad_maxima,
            dictamen="",  # se llena luego
        )

        estatus = _determine_estatus(errores_tecnicos, resumen)
        resumen.dictamen = estatus

        score_doc = _build_score_documental(resumen, inventario)

        # Guardar salida
        ts = datetime.now().strftime("%m%d%Y")
        out_name = f"{id_operacion}_ARGO_DOCUMENT_{estatus}_{ts}.xlsx"
        output_path = os.path.join(outputs_dir, out_name)

        _fill_resumen(
            ws=ws_resumen,
            estatus=estatus,
            output_path=output_path,
            id_operacion=id_operacion,
            score_total=score_doc.score_total,
            delta_global=score_doc.impacto_en_score_global["delta"],
            severidad_maxima=resumen.severidad_maxima,
        )

        wb.save(output_path)
        wb.close()

        meta = MetaDocumento(
            schema=SCHEMA_NAME,
            version_motor=VERSION_MOTOR,
            modo=MODO,
            id_operacion=id_operacion,
            timestamp_local=_now_iso(),
            hash_input=_sha256_file(input_xlsx_path),
            input_filename=input_filename,
            plantilla_filename=plantilla_filename,
            expediente_resumen={
                "archivos_recibidos": 1,
                "tipos_detectados": [x.tipo_documento for x in inventario if x.presente],
            },
        )

        advertencia_juridica = (
            "La generación documental se realizó con base en la información y soportes disponibles. "
            "La responsabilidad final de revisión y uso operativo corresponde al usuario."
        )

        return ArgoDocumentSalida(
            ok=True,
            modulo="ARGO_DOCUMENT",
            meta=asdict(meta),
            estatus=estatus,
            icono=ESTATUS_ICONOS[estatus],
            output_path=output_path,
            resumen=asdict(resumen),
            inventario_documental=[asdict(x) for x in inventario],
            mapeo_campos=[asdict(x) for x in mapeo_campos],
            alertas=[asdict(x) for x in alertas],
            score_documental_doc=asdict(score_doc),
            advertencia_juridica=advertencia_juridica,
            errores_tecnicos=[],
        )

    except Exception as e:
        errores_tecnicos.append({
            "codigo": "DOC-TECH-001",
            "mensaje": "Error técnico durante procesamiento de ARGO DOCUMENT.",
            "detalle": str(e),
        })

        estatus = "ERROR_TECNICO"
        icono = ESTATUS_ICONOS[estatus]

        if not id_operacion:
            id_operacion = _build_id_operacion()

        return ArgoDocumentSalida(
            ok=False,
            modulo="ARGO_DOCUMENT",
            meta={
                "schema": SCHEMA_NAME,
                "version_motor": VERSION_MOTOR,
                "modo": MODO,
                "id_operacion": id_operacion,
                "timestamp_local": _now_iso(),
                "hash_input": "",
                "input_filename": os.path.basename(input_xlsx_path) if input_xlsx_path else "",
                "plantilla_filename": os.path.basename(plantilla_path) if plantilla_path else "",
                "expediente_resumen": {
                    "archivos_recibidos": 0,
                    "tipos_detectados": [],
                }
            },
            estatus=estatus,
            icono=icono,
            output_path="",
            resumen=asdict(ResumenDocumento(
                documentos_detectados=0,
                documentos_requeridos=0,
                faltantes=0,
                campos_total=0,
                campos_ok=0,
                campos_error=0,
                campos_no_verificables=0,
                severidad_maxima="CRITICA",
                dictamen=estatus,
            )),
            inventario_documental=[],
            mapeo_campos=[],
            alertas=[],
            score_documental_doc=asdict(ScoreDocumentalDoc(
                score_total=0,
                nivel="BAJO",
                componentes={
                    "completitud": 0,
                    "consistencia": 0,
                    "verificabilidad": 0,
                    "trazabilidad": 0,
                },
                impacto_en_score_global={
                    "metodo": "ADITIVO_CAPADO",
                    "delta": 0,
                }
            )),
            advertencia_juridica="No fue posible generar el documento por error técnico real.",
            errores_tecnicos=errores_tecnicos,
        )


def salida_to_dict(salida: ArgoDocumentSalida) -> Dict[str, Any]:
    return asdict(salida)
