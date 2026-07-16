import os
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image


def generar_reporte_ejecutivo(plantilla, datos_operacion, carpeta_salida):

    os.makedirs(carpeta_salida, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    ruta_salida = os.path.join(
        carpeta_salida,
        f"reporte_ejecutivo_argo_{timestamp}.xlsx"
    )

    wb = load_workbook(plantilla)

    if "Reporte Ejecutivo" in wb.sheetnames:
        del wb["Reporte Ejecutivo"]

    ws = wb.create_sheet("Reporte Ejecutivo", 0)

    azul_oscuro = "08162B"
    gris_claro = "EAEAEA"
    blanco = "FFFFFF"

    borde = Border(
        left=Side(style="thin", color="999999"),
        right=Side(style="thin", color="999999"),
        top=Side(style="thin", color="999999"),
        bottom=Side(style="thin", color="999999"),
    )

    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 4
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 32

    for row in range(1, 70):
        ws.row_dimensions[row].height = 24

    # =========================
    # LOGO
    # =========================
    try:
        base_dir = os.path.dirname(os.path.abspath(__file__))
        logo_path = os.path.join(base_dir, "assets", "logo_argo_excel.jpg")

        print(f"[ARGO] Buscando logo en: {logo_path}")

        if os.path.exists(logo_path):
            img = Image(logo_path)
            img.width = 210
            img.height = 85
            ws.add_image(img, "B2")
            print("[ARGO] Logo cargado correctamente")
        else:
            print(f"[ARGO] Logo no encontrado: {logo_path}")

    except Exception as e:
        print(f"[ARGO] Error cargando logo: {e}")

    # =========================
    # ENCABEZADO
    # =========================
    ws.merge_cells("B6:F8")
    ws["B6"] = "ARGO - REPORTE EJECUTIVO PREMIUM"
    ws["B6"].font = Font(size=22, bold=True, color=blanco)
    ws["B6"].fill = PatternFill(
        start_color=azul_oscuro,
        end_color=azul_oscuro,
        fill_type="solid"
    )
    ws["B6"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B9:F9")
    ws["B9"] = "Automatización inteligente de procesos aduaneros"
    ws["B9"].font = Font(size=11, italic=True, color="666666")
    ws["B9"].alignment = Alignment(horizontal="center")

    # =========================
    # DATOS GENERALES
    # =========================
    datos = [
        ("Cliente", datos_operacion.get("cliente")),
        ("Proveedor", datos_operacion.get("proveedor")),
        ("Paquetería", datos_operacion.get("paqueteria")),
        ("Tracking", datos_operacion.get("tracking")),
        ("Descripción", datos_operacion.get("descripcion")),
        ("Cantidad Bultos", datos_operacion.get("cantidad_bultos")),
        ("Peso Total", datos_operacion.get("peso_total")),
        ("Unidad Peso", datos_operacion.get("peso_unidad")),
    ]

    fila = 12

    for campo, valor in datos:
        ws[f"B{fila}"] = campo
        ws[f"B{fila}"].font = Font(size=11, bold=True)
        ws[f"B{fila}"].fill = PatternFill(
            start_color=gris_claro,
            end_color=gris_claro,
            fill_type="solid"
        )
        ws[f"B{fila}"].border = borde
        ws[f"B{fila}"].alignment = Alignment(vertical="center")

        ws[f"C{fila}"] = str(valor) if valor not in [None, ""] else "N/D"
        ws[f"C{fila}"].font = Font(size=11)
        ws[f"C{fila}"].border = borde
        ws[f"C{fila}"].alignment = Alignment(wrap_text=True, vertical="center")

        fila += 1

    ws["E12"] = "Fecha"
    ws["E12"].font = Font(size=11, bold=True)
    ws["E12"].fill = PatternFill(
        start_color=gris_claro,
        end_color=gris_claro,
        fill_type="solid"
    )
    ws["E12"].border = borde

    ws["F12"] = datetime.now().strftime("%d/%m/%Y %H:%M")
    ws["F12"].font = Font(size=11)
    ws["F12"].border = borde

    # =========================
    # VALIDACIÓN OPERACIONAL
    # Prioridad ejecutiva y comercial del reporte
    # =========================
    semaforo_operativo = (
        datos_operacion.get("semaforo_operativo") or "SIN CONTROL"
    )
    icono_operativo = datos_operacion.get("icono_operativo") or ""
    cobertura = datos_operacion.get("cobertura_validacion_pct") or 0
    dictamen = (
        datos_operacion.get("dictamen_operativo")
        or "Sin dictamen operativo."
    )
    campos_totales = datos_operacion.get("campos_totales") or 0
    campos_disponibles = datos_operacion.get("campos_disponibles") or 0
    campos_no_verificables = (
        datos_operacion.get("campos_no_verificables") or 0
    )
    validaciones = (
        datos_operacion.get("validaciones_operativas") or []
    )

    semaforo_texto = str(semaforo_operativo).upper()

    color_semaforo = {
        "VERDE": "2E7D32",
        "AMARILLO": "F9A825",
        "ROJO": "C62828",
    }.get(semaforo_texto, "666666")

    color_texto_semaforo = (
        "000000" if semaforo_texto == "AMARILLO" else "FFFFFF"
    )

    borde_estado = Border(
        left=Side(style="medium", color="333333"),
        right=Side(style="medium", color="333333"),
        top=Side(style="medium", color="333333"),
        bottom=Side(style="medium", color="333333"),
    )

    ws.merge_cells("B22:F22")
    ws["B22"] = "VALIDACIÓN OPERACIONAL"
    ws["B22"].font = Font(size=15, bold=True, color=blanco)
    ws["B22"].fill = PatternFill(
        start_color=azul_oscuro,
        end_color=azul_oscuro,
        fill_type="solid"
    )
    ws["B22"].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    ws.merge_cells("B23:C26")
    ws["B23"] = (
        f"ESTADO OPERATIVO\n"
        f"{icono_operativo} {semaforo_texto}"
    )
    ws["B23"].font = Font(
        size=18,
        bold=True,
        color=color_texto_semaforo
    )
    ws["B23"].fill = PatternFill(
        start_color=color_semaforo,
        end_color=color_semaforo,
        fill_type="solid"
    )
    ws["B23"].alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    for fila_estado in range(23, 27):
        for columna_estado in ["B", "C"]:
            ws[f"{columna_estado}{fila_estado}"].border = borde_estado

    indicadores_operativos = [
        ("Cobertura de validación", f"{cobertura}%"),
        ("Campos disponibles", campos_disponibles),
        ("Campos no verificables", campos_no_verificables),
        ("Campos evaluados", campos_totales),
    ]

    fila_indicador = 23

    for campo, valor in indicadores_operativos:
        ws[f"E{fila_indicador}"] = campo
        ws[f"E{fila_indicador}"].font = Font(size=10, bold=True)
        ws[f"E{fila_indicador}"].fill = PatternFill(
            start_color=gris_claro,
            end_color=gris_claro,
            fill_type="solid"
        )
        ws[f"E{fila_indicador}"].border = borde
        ws[f"E{fila_indicador}"].alignment = Alignment(
            vertical="center",
            wrap_text=True
        )

        ws[f"F{fila_indicador}"] = str(valor)
        ws[f"F{fila_indicador}"].font = Font(size=11, bold=True)
        ws[f"F{fila_indicador}"].border = borde
        ws[f"F{fila_indicador}"].alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        fila_indicador += 1

    ws.merge_cells("B28:F30")
    ws["B28"] = f"DICTAMEN OPERATIVO\n\n{dictamen}"
    ws["B28"].font = Font(size=12, bold=True, color="000000")
    ws["B28"].alignment = Alignment(
        wrap_text=True,
        vertical="center"
    )
    ws["B28"].border = borde

    # =========================
    # TABLA DE VALIDACIONES
    # =========================
    ws.merge_cells("B32:F32")
    ws["B32"] = "DETALLE DE VALIDACIONES OPERATIVAS"
    ws["B32"].font = Font(size=13, bold=True, color=blanco)
    ws["B32"].fill = PatternFill(
        start_color=azul_oscuro,
        end_color=azul_oscuro,
        fill_type="solid"
    )
    ws["B32"].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    encabezados = [
        ("B33", "Campo"),
        ("C33", "Valor documental"),
        ("E33", "Estado"),
        ("F33", "Severidad"),
    ]

    for celda, titulo in encabezados:
        ws[celda] = titulo
        ws[celda].font = Font(size=10, bold=True)
        ws[celda].fill = PatternFill(
            start_color=gris_claro,
            end_color=gris_claro,
            fill_type="solid"
        )
        ws[celda].border = borde
        ws[celda].alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    fila_validacion = 34

    for item in validaciones[:12]:
        etiqueta = (
            item.get("etiqueta")
            or item.get("campo")
            or "N/D"
        )
        valor_documental = item.get("valor_documental")
        estado = item.get("estado") or ""
        resultado = (
            item.get("resultado")
            or estado
            or "N/D"
        )
        severidad = item.get("severidad") or "N/D"

        if estado == "DISPONIBLE":
            estado_impresion = "OK - DISPONIBLE"
        elif estado == "NO_VERIFICABLE":
            estado_impresion = "REVISAR - NO VERIFICABLE"
        else:
            estado_impresion = resultado

        ws[f"B{fila_validacion}"] = etiqueta
        ws[f"C{fila_validacion}"] = (
            str(valor_documental)
            if valor_documental not in [None, ""]
            else "N/D"
        )
        ws[f"E{fila_validacion}"] = estado_impresion
        ws[f"F{fila_validacion}"] = severidad

        if estado == "NO_VERIFICABLE":
            ws[f"E{fila_validacion}"].font = Font(
                size=10,
                bold=True
            )
        else:
            ws[f"E{fila_validacion}"].font = Font(size=10)

        for columna in ["B", "C", "E", "F"]:
            ws[f"{columna}{fila_validacion}"].border = borde
            ws[f"{columna}{fila_validacion}"].alignment = Alignment(
                wrap_text=True,
                vertical="center"
            )

            if columna != "E":
                ws[f"{columna}{fila_validacion}"].font = Font(
                    size=10
                )

        fila_validacion += 1

    # =========================
    # ANÁLISIS DOCUMENTAL
    # Información complementaria, no protagonista
    # =========================
    riesgo = datos_operacion.get("riesgo_automatico") or "MEDIA"
    score = datos_operacion.get("score_documental") or 0
    fraccion = (
        datos_operacion.get("fraccion_sugerida")
        or "7318.15.99"
    )
    confianza = (
        datos_operacion.get("confianza_fraccion_pct") or 0
    )
    certeza = datos_operacion.get("certeza_final_pct") or 0
    diligencia = (
        datos_operacion.get("nivel_debida_diligencia")
        or "BASICA"
    )

    ws.merge_cells("B47:F47")
    ws["B47"] = "ANÁLISIS DOCUMENTAL Y CLASIFICACIÓN"
    ws["B47"].font = Font(size=13, bold=True, color=blanco)
    ws["B47"].fill = PatternFill(
        start_color=azul_oscuro,
        end_color=azul_oscuro,
        fill_type="solid"
    )
    ws["B47"].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    resumen_documental = (
        f"ARGO identificó riesgo automático {riesgo}, "
        f"score documental {score}, fracción sugerida {fraccion}, "
        f"confianza de clasificación {confianza}% y certeza final "
        f"{certeza}%. Nivel de debida diligencia recomendado: "
        f"{diligencia}."
    )

    ws.merge_cells("B48:F50")
    ws["B48"] = resumen_documental
    ws["B48"].font = Font(size=10, color="000000")
    ws["B48"].alignment = Alignment(
        wrap_text=True,
        vertical="center"
    )
    ws["B48"].border = borde

    matriz_documental = [
        ("Riesgo automático", riesgo),
        ("Score documental", score),
        ("Fracción sugerida", fraccion),
        ("Confianza fracción", f"{confianza}%"),
        ("Certeza final", f"{certeza}%"),
        ("Debida diligencia", diligencia),
    ]

    fila_matriz = 52

    for campo, valor in matriz_documental:
        ws[f"B{fila_matriz}"] = campo
        ws[f"B{fila_matriz}"].font = Font(size=10, bold=True)
        ws[f"B{fila_matriz}"].fill = PatternFill(
            start_color=gris_claro,
            end_color=gris_claro,
            fill_type="solid"
        )
        ws[f"B{fila_matriz}"].border = borde

        ws[f"C{fila_matriz}"] = str(valor)
        ws[f"C{fila_matriz}"].font = Font(size=10)
        ws[f"C{fila_matriz}"].border = borde
        ws[f"C{fila_matriz}"].alignment = Alignment(
            wrap_text=True
        )

        fila_matriz += 1

    # =========================
    # ADVERTENCIA OPERATIVA
    # =========================
    ws.merge_cells("B60:F63")
    ws["B60"] = (
        "Advertencia: ARGO procesa información con base en los "
        "documentos e imágenes proporcionados por el operador. "
        "La captura, legibilidad y validez documental son "
        "responsabilidad del usuario operativo. La clasificación "
        "sugerida debe ser validada por personal autorizado antes "
        "de su uso definitivo."
    )
    ws["B60"].font = Font(
        size=9,
        italic=True,
        color="555555"
    )
    ws["B60"].alignment = Alignment(
        wrap_text=True,
        vertical="center"
    )
    ws["B60"].border = borde

    # =========================
    # PIE
    # =========================
    ws.merge_cells("B65:F66")
    ws["B65"] = (
        "Reporte generado automáticamente por ARGO v2026"
    )
    ws["B65"].font = Font(
        size=10,
        italic=True,
        color="666666"
    )
    ws["B65"].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    wb.save(ruta_salida)

    print(f"[ARGO] Reporte ejecutivo generado: {ruta_salida}")

    return ruta_salida
