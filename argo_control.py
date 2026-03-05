from openpyxl import load_workbook
from datetime import datetime
import os


def argo_control_validar_v2(entrada_path, control_path, id_operacion=None):
    
    # Cargar archivos
    wb_entrada = load_workbook(entrada_path)
    wb_control = load_workbook(control_path)

    ws_entrada = wb_entrada.active
    ws_control = wb_control["Control"]

    errores = 0
    no_verificables = 0

    fila = 2  # Asumimos que fila 1 es encabezado

    while ws_control[f"A{fila}"].value:

        campo = ws_control[f"A{fila}"].value

        # Buscar valor en archivo entrada
        valor_detectado = None

        for row in ws_entrada.iter_rows(values_only=True):
            if campo in str(row):
                valor_detectado = campo

        if valor_detectado:
            ws_control[f"B{fila}"] = valor_detectado
            ws_control[f"C{fila}"] = "OK"
        else:
            ws_control[f"B{fila}"] = "No verificable"
            ws_control[f"C{fila}"] = "ERROR"
            ws_control[f"D{fila}"] = "Campo no encontrado en ARGO ENTRADA"
            errores += 1

        fila += 1

    # ----------------------------
    # Cálculo de porcentaje dinámico
    # ----------------------------

    total_campos = fila - 2
    total_inconsistencias = errores + no_verificables

    if total_campos > 0:
        porcentaje_error = (total_inconsistencias / total_campos) * 100
    else:
        porcentaje_error = 0

    # Determinar estatus por umbral dinámico

    if porcentaje_error < 10:
        estatus = "APROBADO"
        icono = "🟢"
    elif porcentaje_error < 30:
        estatus = "CON_OBSERVACIONES"
        icono = "🟡"
    else:
        estatus = "RECHAZADO"
        icono = "🔴"
    # ----------------------------
    # Escribir resumen en hoja Control
    # ----------------------------
    ws_control["F1"] = "% Inconsistencias"
    ws_control["G1"] = round(porcentaje_error, 2)

    ws_control["F2"] = "Dictamen"
    ws_control["G2"] = estatus
    
    fecha = datetime.now().strftime("%m%d%Y")

# Nombre del archivo con trazabilidad
if id_operacion:
    nombre_salida = f"{id_operacion}__ARGO_CONTROL__{estatus}_{fecha}.xlsx"
else:
    nombre_salida = f"ARGO_CONTROL_{estatus}_{fecha}.xlsx"

    output_path = os.path.join("outputs", nombre_salida)
    os.makedirs("outputs", exist_ok=True)
    wb_control.save(output_path)
    return output_path, icono, estatus

def extraer_resumen_control_desde_excel(output_path: str) -> dict:
    """
    Resumen mínimo para ARGO CLASS.
    Lee el Excel generado por ARGO CONTROL y calcula:
      - observaciones_total: cantidad de filas con Estado != OK (col C)
      - severidad_maxima: BAJA/MEDIA/ALTA según dictamen (G2) o por observaciones
      - dictamen: APROBADO / CON_OBSERVACIONES / RECHAZADO (si existe en G2)
    """
    from openpyxl import load_workbook

    resumen = {
        "observaciones_total": 0,
        "severidad_maxima": "BAJA",
        "dictamen": None
    }

    try:
        wb = load_workbook(output_path, data_only=True)

        if "Control" not in wb.sheetnames:
            return resumen

        ws = wb["Control"]

        # Dictamen que tú ya escribes en F2/G2
        dictamen = ws["G2"].value
        if isinstance(dictamen, str) and dictamen.strip():
            dictamen = dictamen.strip().upper()
            resumen["dictamen"] = dictamen

            if dictamen == "RECHAZADO":
                resumen["severidad_maxima"] = "ALTA"
            elif dictamen == "CON_OBSERVACIONES":
                resumen["severidad_maxima"] = "MEDIA"
            elif dictamen == "APROBADO":
                resumen["severidad_maxima"] = "BAJA"

        # Contar observaciones (columna C = Estado)
        fila = 2
        obs = 0
        while ws[f"A{fila}"].value:
            estado = ws[f"C{fila}"].value
            estado_txt = (str(estado).strip().upper() if estado is not None else "")
            if estado_txt and estado_txt != "OK":
                obs += 1
            fila += 1

        resumen["observaciones_total"] = obs

        # Si no hubo dictamen, derivar severidad por cantidad
        if resumen["dictamen"] is None:
            if obs == 0:
                resumen["severidad_maxima"] = "BAJA"
            elif obs <= 3:
                resumen["severidad_maxima"] = "MEDIA"
            else:
                resumen["severidad_maxima"] = "ALTA"

        return resumen

    except Exception:
        return resumen
